import requests
import json
import logging
import os
import time
import schedule
from datetime import datetime, timedelta
from telegram import Bot, Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, CommandHandler, CallbackContext, MessageHandler, Filters
from telegram.error import TelegramError
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials
import fcntl
import threading
import atexit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import random
import hmac
import hashlib
import re
from typing import Optional, List, Dict, Set, Any, Union
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from functools import wraps

# Load environment variables
load_dotenv()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)



def rate_limit_google_sheets(calls_per_minute=60):
    """Decorator to rate limit Google Sheets API calls"""
    min_interval = 60.0 / calls_per_minute
    last_called = [0.0]
    
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            elapsed = time.time() - last_called[0]
            left_to_wait = min_interval - elapsed
            
            if left_to_wait > 0:
                time.sleep(left_to_wait)
            
            last_called[0] = time.time()
            return func(*args, **kwargs)
        return wrapper
    return decorator


class CircuitBreaker:
    def __init__(self, failure_threshold=5, recovery_timeout=60):
        self.failure_threshold = failure_threshold
        self.recovery_timeout = recovery_timeout
        self.failure_count = 0
        self.last_failure_time = 0
        self.state = "CLOSED"  # CLOSED, OPEN, HALF_OPEN
        
    def can_execute(self):
        if self.state == "OPEN":
            if time.time() - self.last_failure_time > self.recovery_timeout:
                self.state = "HALF_OPEN"
                return True
            return False
        return True
    
    def record_success(self):
        self.failure_count = 0
        self.state = "CLOSED"
        
    def record_failure(self):
        self.failure_count += 1
        self.last_failure_time = time.time()
        if self.failure_count >= self.failure_threshold:
            self.state = "OPEN"

class MEXCTracker:
    def __init__(self):
        self.bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
        self.chat_id = os.getenv('TELEGRAM_CHAT_ID')
        self.update_interval = int(os.getenv('UPDATE_INTERVAL', 60))
        self.price_check_interval = int(os.getenv('PRICE_CHECK_INTERVAL', 5))  # minutes
        
        if not self.bot_token:
            raise ValueError("TELEGRAM_BOT_TOKEN is required")
        
        self.updater = Updater(token=self.bot_token, use_context=True)
        self.dispatcher = self.updater.dispatcher
        self.bot = Bot(token=self.bot_token)
        self.setup_handlers()
        self.init_data_file()
        self.last_unique_futures = set()
        self.bybit_api_key = os.getenv('BYBIT_API_KEY', '')
        self.bybit_api_secret = os.getenv('BYBIT_API_SECRET', '')
        

        # Initialize Google Sheets attributes to None
        self.gs_client = None
        self.spreadsheet = None
        self.creds = None
        self.sheets_circuit_breaker = CircuitBreaker()
        
        self.last_sheets_update = 0
        self.sheets_update_interval = 60  # seconds
        self.sheets_batch_data = {}
        self.sheets_cache_time = 0
        self.sheets_cache_duration = 30  # seconds
        # Price tracking
        self.price_history = {}  # symbol: {timestamp: price}
        self.last_price_check = None
        self.restart_count = 0
        self.last_restart = None
        # Google Sheets setup
        self.setup_google_sheets()
        self.session = self._create_session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
        })
        self.proxies = self._get_proxies()

    def _create_session(self):
        """Create requests session with minimal headers"""
        session = requests.Session()
        
        # Simple retry strategy
        retry_strategy = Retry(
            total=2,
            backoff_factor=0.5,
            status_forcelist=[429, 500, 502, 503, 504],
        )
        
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        
        return session

    def _get_proxies(self) -> List[dict]:
        return [{}]  # Empty dict means no proxy

    def setup_google_sheets(self):
        """Setup Google Sheets connection with spreadsheet discovery"""
        try:
            logger.info("🔧 Starting Google Sheets setup...")
            
            creds_json = os.getenv('GOOGLE_CREDENTIALS_JSON')
            sheet_email = os.getenv('GOOGLE_SHEET_EMAIL')
            
            logger.info(f"📝 Environment check - CREDENTIALS_JSON: {'✅ Set' if creds_json else '❌ Missing'}")
            logger.info(f"📝 Environment check - SHEET_EMAIL: {'✅ ' + sheet_email if sheet_email else '❌ Missing'}")
            
            if not creds_json:
                logger.error("❌ GOOGLE_CREDENTIALS_JSON is missing")
                return False

            if not sheet_email:
                logger.error("❌ GOOGLE_SHEET_EMAIL is missing")
                return False

            # Parse credentials
            try:
                creds_dict = json.loads(creds_json)
                service_email = creds_dict.get('client_email', 'Unknown')
                logger.info(f"✅ Credentials parsed - Service account: {service_email}")
            except Exception as e:
                logger.error(f"❌ Error parsing credentials: {e}")
                return False

            # Setup authentication
            try:
                scope = [
                    'https://www.googleapis.com/auth/spreadsheets',
                    'https://www.googleapis.com/auth/drive'
                ]
                self.creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
                self.gs_client = gspread.authorize(self.creds)
                logger.info("✅ Google Sheets client authorized successfully")
            except Exception as e:
                logger.error(f"❌ Authentication failed: {e}")
                return False

            # Discover available spreadsheets
            logger.info("🔍 Discovering available spreadsheets...")
            try:
                all_spreadsheets = self.gs_client.list_spreadsheet_files()
                logger.info(f"📋 Found {len(all_spreadsheets)} spreadsheets accessible by the service account")
                
                if not all_spreadsheets:
                    logger.error("❌ No spreadsheets found. The service account has no access to any spreadsheets.")
                    return False
                
                # Log all available spreadsheets for debugging
                logger.info("📝 Available spreadsheets:")
                for i, spreadsheet in enumerate(all_spreadsheets[:10]):  # Show first 10
                    logger.info(f"   {i+1}. '{spreadsheet['name']}' (ID: {spreadsheet['id']})")
                
                if len(all_spreadsheets) > 10:
                    logger.info(f"   ... and {len(all_spreadsheets) - 10} more")
                
                # Try to find matching spreadsheet
                matching_spreadsheets = [
                    s for s in all_spreadsheets 
                    if sheet_email.lower() in s['name'].lower()
                ]
                
                if matching_spreadsheets:
                    if len(matching_spreadsheets) > 1:
                        logger.info(f"🔍 Found {len(matching_spreadsheets)} matching spreadsheets:")
                        for match in matching_spreadsheets:
                            logger.info(f"   - '{match['name']}' (ID: {match['id']})")
                    
                    # Use the first match
                    spreadsheet_info = matching_spreadsheets[0]
                    logger.info(f"✅ Selecting spreadsheet: '{spreadsheet_info['name']}'")
                    self.spreadsheet = self.gs_client.open_by_key(spreadsheet_info['id'])
                    logger.info(f"✅ Connected to: {self.spreadsheet.title}")
                    
                else:
                    logger.error(f"❌ No spreadsheet found with name containing: '{sheet_email}'")
                    logger.info("💡 Available spreadsheet names:")
                    for s in all_spreadsheets[:5]:
                        logger.info(f"   - '{s['name']}'")
                    
                    # Let's try to create one
                    logger.info(f"🆕 Creating new spreadsheet: '{sheet_email}'")
                    try:
                        self.spreadsheet = self.gs_client.create(sheet_email)
                        logger.info(f"✅ Created new spreadsheet: {self.spreadsheet.title}")
                        
                        # Share with the service account
                        try:
                            self.spreadsheet.share(service_email, perm_type='user', role='writer')
                            logger.info(f"✅ Shared with service account: {service_email}")
                        except Exception as share_error:
                            logger.warning(f"⚠️ Could not share spreadsheet: {share_error}")
                            
                    except Exception as create_error:
                        logger.error(f"❌ Failed to create spreadsheet: {create_error}")
                        return False
                        
            except Exception as e:
                logger.error(f"❌ Error discovering spreadsheets: {e}")
                return False

            # Test connection
            try:
                worksheet = self.spreadsheet.sheet1
                logger.info("✅ Sheet connection test successful")
                return True
            except Exception as test_error:
                logger.error(f"❌ Failed to access worksheet: {test_error}")
                return False
                
        except Exception as e:
            logger.error(f"❌ Google Sheets setup error: {e}")
            return False
        
    # ==================== PRICE MONITORING ====================






    def get_all_mexc_prices(self):
        """Get price data for MEXC futures - CONSISTENT with check command"""
        try:
            # Use the EXACT SAME approach as check command
            batch_data = self.get_mexc_prices_batch_working()
            
            # Get unique futures
            unique_futures, _ = self.find_unique_futures_robust()
            
            # Apply the same matching logic as check command
            price_data = {}
            
            for symbol in unique_futures:
                # Try exact match first
                if symbol in batch_data:
                    price_data[symbol] = batch_data[symbol]
                else:
                    # Try alternative formats (same as check command)
                    alt_formats = [
                        symbol.replace('_', ''),
                        symbol.replace('_', '-'), 
                        symbol.replace('_', '/'),
                    ]
                    
                    found = False
                    for alt_format in alt_formats:
                        if alt_format in batch_data:
                            price_data[symbol] = batch_data[alt_format].copy()
                            price_data[symbol]['symbol'] = symbol  # Fix symbol name
                            found = True
                            break
                    
                    if not found:
                        # Symbol not found in batch, add with None price
                        price_data[symbol] = {
                            'symbol': symbol,
                            'price': None,
                            'changes': {},
                            'timestamp': datetime.now(),
                            'source': 'not_found'
                        }
            
            return price_data
            
        except Exception as e:
            logger.error(f"Error in get_all_mexc_prices: {e}")
            return {}
    
    def _get_price_sources_summary(self, price_data):
        """Get summary of where prices came from"""
        sources = {}
        for symbol, data in price_data.items():
            if data.get('price') is not None:
                source = data.get('source', 'unknown')
                sources[source] = sources.get(source, 0) + 1
        return sources
   
    def get_mexc_prices_batch_working(self):
        """Get prices using working MEXC API endpoint - WITH RATE LIMITING & RETRIES"""
        try:
            url = "https://contract.mexc.com/api/v1/contract/ticker"
            
            # Add retry logic
            for attempt in range(3):
                try:
                    response = requests.get(url, timeout=15)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        if data.get('success'):
                            tickers = data.get('data', [])
                            price_data = {}
                            
                            for ticker in tickers:
                                try:
                                    symbol = ticker.get('symbol')
                                    price_str = ticker.get('lastPrice')
                                    
                                    if symbol and price_str:
                                        price = float(price_str)
                                        
                                        # FIX: ACCEPT ALL VALID PRICES, EVEN VERY SMALL ONES
                                        # Only skip negative prices
                                        if price < 0:
                                            continue
                                            
                                        change_rate = float(ticker.get('riseFallRate', 0)) * 100
                                        
                                        price_data[symbol] = {
                                            'symbol': symbol,
                                            'price': price,
                                            'changes': {
                                                '5m': change_rate,
                                                '60m': change_rate,
                                                '240m': change_rate
                                            },
                                            'timestamp': datetime.now(),
                                            'source': 'batch_ticker'
                                        }
                                except (ValueError, TypeError) as e:
                                    continue
                            
                            logger.info(f"✅ Batch prices: {len(price_data)} symbols (attempt {attempt + 1})")
                            return price_data
                    
                    # If we get here, the request failed
                    if attempt < 2:  # Don't sleep on last attempt
                        time.sleep(1)  # Wait 1 second before retry
                        
                except requests.exceptions.Timeout:
                    logger.warning(f"⚠️ Batch API timeout on attempt {attempt + 1}")
                    if attempt < 2:
                        time.sleep(2)
                except requests.exceptions.ConnectionError:
                    logger.warning(f"⚠️ Batch API connection error on attempt {attempt + 1}")
                    if attempt < 2:
                        time.sleep(2)
                except Exception as e:
                    logger.warning(f"⚠️ Batch API error on attempt {attempt + 1}: {e}")
                    if attempt < 2:
                        time.sleep(1)
            
            return {}
            
        except Exception as e:
            logger.error(f"Batch price error: {e}")
            return {}
    
    def get_mexc_price_data_working(self, symbol):
        """Get individual price data - ACCEPT MICRO-CAP PRICES"""
        try:
            endpoints = [
            f"https://contract.mexc.com/api/v1/contract/ticker?symbol={symbol}",
            f"https://contract.mexc.com/api/v1/contract/detail?symbol={symbol}",
            f"https://futures.mexc.com/api/v1/contract/ticker?symbol={symbol}"  # Alternative domain
        ]
            
            for url in endpoints:
                try:
                    response = requests.get(url, timeout=10)
                    
                    if response.status_code == 200:
                        data = response.json()
                        if data.get('success', False):
                            ticker_data = data.get('data', {})
                            
                            if isinstance(ticker_data, list) and ticker_data:
                                ticker_data = ticker_data[0]
                            
                            price_str = ticker_data.get('lastPrice') or ticker_data.get('price')
                            if price_str:
                                price = float(price_str)
                                
                                # FIX: ACCEPT ALL PRICES, EVEN MICRO-CAP
                                # Only filter out truly invalid prices (negative or None)
                                if price is None or price < 0:
                                    logger.debug(f"⚠️ Skipping {symbol} - invalid price: {price}")
                                    continue
                                    
                                change_rate = float(ticker_data.get('riseFallRate', 0)) * 100
                                
                                return {
                                    'symbol': symbol,
                                    'price': price,
                                    'changes': {
                                        '5m': change_rate,
                                        '60m': change_rate,
                                        '240m': change_rate
                                    },
                                    'timestamp': datetime.now(),
                                    'source': 'individual'
                                }
                except Exception as endpoint_error:
                    continue
            
            return None
            
        except Exception as e:
            logger.debug(f"Individual price error for {symbol}: {e}")
            return None
    
    def get_mexc_price_data(self, symbol):
        """Main price data method - use the working version"""
        return self.get_mexc_price_data_working(symbol)    

    def get_mexc_prices_batch(self):
        """Get prices in batch using ticker endpoint"""
        try:
            url = "https://contract.mexc.com/api/v1/contract/ticker"
            response = requests.get(url, timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success'):
                    tickers = data.get('data', [])
                    price_data = {}
                    
                    for ticker in tickers:
                        symbol = ticker.get('symbol')
                        price = ticker.get('lastPrice')
                        
                        if symbol and price:
                            # Add underscore for consistency
                            formatted_symbol = symbol
                            price_data[formatted_symbol] = {
                                'symbol': formatted_symbol,
                                'price': float(price),
                                'changes': {},  # No historical changes in batch
                                'timestamp': datetime.now(),
                                'source': 'batch_ticker'
                            }
                    
                    logger.info(f"✅ Batch prices: {len(price_data)} symbols")
                    return price_data
            
            return {}
            
        except Exception as e:
            logger.error(f"Batch price method error: {e}")
            return {}




    def get_consistent_price_data(self):
        """Get consistent price data with proper historical tracking"""
        try:
            current_time = datetime.now()
            cache_key = "price_data_cache"
            cache_duration = 30  # seconds
            
            # Check if we have recent cached data
            if hasattr(self, '_price_data_cache') and hasattr(self, '_price_cache_time'):
                if (current_time - self._price_cache_time).seconds < cache_duration:
                    logger.info("🔄 Using cached price data")
                    return self._price_data_cache.copy()
            
            # Get fresh data from batch API
            batch_data = self.get_mexc_prices_batch_working()
            logger.info(f"📊 Fresh batch data: {len(batch_data)} symbols")
            
            # Get unique futures
            unique_futures, _ = self.find_unique_futures_robust()
            
            # Apply matching logic
            price_data = {}
            matched_symbols = 0
            
            for symbol in unique_futures:
                # Try exact match first
                if symbol in batch_data:
                    batch_info = batch_data[symbol]
                    current_price = batch_info['price']
                    current_change = batch_info['changes'].get('5m', 0)
                    
                    # Calculate proper historical changes
                    historical_changes = self.calculate_historical_changes(symbol, current_price)
                    
                    price_data[symbol] = {
                        'symbol': symbol,
                        'price': current_price,
                        'changes': historical_changes,  # Use calculated historical changes
                        'timestamp': current_time,
                        'source': 'batch_exact'
                    }
                    matched_symbols += 1
                else:
                    # Try alternative formats
                    alt_formats = [
                        symbol.replace('_', ''),
                        symbol.replace('_', '-'), 
                        symbol.replace('_', '/'),
                    ]
                    
                    found = False
                    for alt_format in alt_formats:
                        if alt_format in batch_data:
                            batch_info = batch_data[alt_format]
                            current_price = batch_info['price']
                            current_change = batch_info['changes'].get('5m', 0)
                            
                            # Calculate proper historical changes
                            historical_changes = self.calculate_historical_changes(symbol, current_price)
                            
                            price_data[symbol] = {
                                'symbol': symbol,
                                'price': current_price,
                                'changes': historical_changes,  # Use calculated historical changes
                                'timestamp': current_time,
                                'source': f'batch_alt_{alt_format}'
                            }
                            matched_symbols += 1
                            found = True
                            break
                    
                    if not found:
                        price_data[symbol] = {
                            'symbol': symbol,
                            'price': None,
                            'changes': {},
                            'timestamp': current_time,
                            'source': 'not_found'
                        }
            
            # Cache the results
            self._price_data_cache = price_data.copy()
            self._price_cache_time = current_time
            
            logger.info(f"💰 Consistent price data: {matched_symbols}/{len(unique_futures)} matched")
            return price_data
            
        except Exception as e:
            logger.error(f"Consistent price data error: {e}")
            return {}

    def calculate_historical_changes(self, symbol, current_price):
        """Calculate proper historical price changes for different timeframes"""
        try:
            changes = {}
            
            # Get price history for this symbol
            if symbol not in self.price_history:
                self.price_history[symbol] = {}
            
            current_time = datetime.now()
            
            # Store current price in history
            self.price_history[symbol][current_time] = current_price
            
            # Clean old history (keep only last 24 hours)
            cutoff_time = current_time - timedelta(hours=24)
            self.price_history[symbol] = {
                ts: price for ts, price in self.price_history[symbol].items() 
                if ts > cutoff_time
            }
            
            # Calculate changes for different timeframes
            timeframes = [
                ('5m', timedelta(minutes=5)),
                ('15m', timedelta(minutes=15)),
                ('30m', timedelta(minutes=30)),
                ('60m', timedelta(hours=1)),
                ('240m', timedelta(hours=4))
            ]
            
            for timeframe_name, time_delta in timeframes:
                target_time = current_time - time_delta
                historical_price = self.find_historical_price(symbol, target_time)
                
                if historical_price and historical_price > 0:
                    price_change = ((current_price - historical_price) / historical_price) * 100
                    changes[timeframe_name] = price_change
                else:
                    # If no historical data, use current change or set to None
                    changes[timeframe_name] = None
            
            return changes
            
        except Exception as e:
            logger.error(f"Error calculating historical changes for {symbol}: {e}")
            return {}

    def find_historical_price(self, symbol, target_time):
        """Find the closest historical price to the target time"""
        try:
            if symbol not in self.price_history or not self.price_history[symbol]:
                return None
            
            closest_time = None
            min_time_diff = timedelta.max
            
            for timestamp in self.price_history[symbol].keys():
                time_diff = abs(timestamp - target_time)
                if time_diff < min_time_diff:
                    min_time_diff = time_diff
                    closest_time = timestamp
            
            # Only return if within reasonable time window (2x the timeframe)
            max_allowed_diff = (target_time - (target_time - timedelta(hours=2))).total_seconds()
            if min_time_diff.total_seconds() <= max_allowed_diff:
                return self.price_history[symbol][closest_time]
            
            return None
            
        except Exception as e:
            logger.error(f"Error finding historical price for {symbol}: {e}")
            return None
        
    def get_all_mexc_prices(self):
        """Get price data for MEXC futures - USE CONSISTENT APPROACH"""
        return self.get_consistent_price_data()

    def analyze_price_movements(self, price_data):
        """Analyze price movements with proper historical data"""
        try:
            logger.info(f"🔍 Analyzing price movements for {len(price_data)} symbols")
            
            symbols_with_changes = []
            
            for symbol, data in price_data.items():
                changes = data.get('changes', {})
                price = data.get('price', 0)
                
                # Calculate overall score based on available timeframes
                score = 0
                weight_total = 0
                timeframe_weights = {
                    '5m': 2.0,   # Highest weight for recent changes
                    '15m': 1.5,
                    '30m': 1.2,
                    '60m': 1.0,
                    '240m': 0.5  # Lowest weight for older changes
                }
                
                for timeframe, weight in timeframe_weights.items():
                    if timeframe in changes and changes[timeframe] is not None:
                        score += changes[timeframe] * weight
                        weight_total += weight
                
                # Normalize score if we have weights
                if weight_total > 0:
                    score = score / weight_total
                
                # Determine latest valid change for display
                latest_change = None
                for timeframe in ['5m', '15m', '30m', '60m', '240m']:
                    if timeframe in changes and changes[timeframe] is not None:
                        latest_change = changes[timeframe]
                        break
                
                symbols_with_changes.append({
                    'symbol': symbol,
                    'price': price,
                    'changes': changes,
                    'score': score,
                    'latest_change': latest_change if latest_change is not None else 0
                })
            
            # Sort by score (highest first)
            symbols_with_changes.sort(key=lambda x: x['score'], reverse=True)
            
            logger.info(f"✅ Analyzed {len(symbols_with_changes)} symbols")
            return symbols_with_changes
            
        except Exception as e:
            logger.error(f"Error analyzing price movements: {e}")
            return []
        
                
    def calculate_changes_from_history(self, symbol, current_price):
        """Calculate price changes from historical data if available"""
        try:
            if symbol not in self.price_history or len(self.price_history[symbol]) < 2:
                return {}
            
            history = self.price_history[symbol]
            timestamps = sorted(history.keys())
            current_time = datetime.now()
            
            changes = {}
            
            # Calculate 5m change
            five_min_ago = current_time - timedelta(minutes=5)
            price_5m = self.find_closest_price(history, five_min_ago)
            if price_5m:
                changes['5m'] = ((current_price - price_5m) / price_5m) * 100
            
            # Calculate 1h change
            one_hour_ago = current_time - timedelta(hours=1)
            price_1h = self.find_closest_price(history, one_hour_ago)
            if price_1h:
                changes['60m'] = ((current_price - price_1h) / price_1h) * 100
            
            # Calculate 4h change
            four_hours_ago = current_time - timedelta(hours=4)
            price_4h = self.find_closest_price(history, four_hours_ago)
            if price_4h:
                changes['240m'] = ((current_price - price_4h) / price_4h) * 100
            
            return changes
            
        except Exception as e:
            logger.error(f"Error calculating changes from history for {symbol}: {e}")
            return {}



    def find_closest_price(self, history, target_time):
        """Find closest price to target time in history"""
        try:
            closest_time = None
            min_diff = timedelta.max
            
            for timestamp in history.keys():
                diff = abs(timestamp - target_time)
                if diff < min_diff:
                    min_diff = diff
                    closest_time = timestamp
            
            # Only return if within reasonable time window
            if min_diff < timedelta(hours=1):
                return history[closest_time]
            return None
            
        except Exception as e:
            logger.error(f"Error finding closest price: {e}")
            return None
        

    # ==================== ENHANCED UNIQUE FUTURES MONITORING ====================

    def monitor_unique_futures_changes(self):
        """Monitor changes in unique futures - FIXED to send notifications"""
        try:
            logger.info("🔍 Monitoring unique futures changes...")
            
            # Get current unique futures
            current_unique, exchange_stats = self.find_unique_futures_robust()
            current_unique_set = set(current_unique)
            
            # Load previous state
            data = self.load_data()
            previous_unique = set(data.get('unique_futures', []))
            
            # Find new unique futures
            new_futures = current_unique_set - previous_unique
            lost_futures = previous_unique - current_unique_set
            
            # Send notifications only if there are changes
            if new_futures:
                self.send_new_unique_notification(new_futures, current_unique_set)
                logger.info(f"🚀 Found {len(new_futures)} new unique futures")
            
            if lost_futures:
                self.send_lost_unique_notification(lost_futures, current_unique_set)
                logger.info(f"📉 Lost {len(lost_futures)} unique futures")
            
            # Update stored data
            data['unique_futures'] = list(current_unique_set)
            data['last_check'] = datetime.now().isoformat()
            data['exchange_stats'] = exchange_stats
            self.save_data(data)
            
            self.last_unique_futures = current_unique_set
            
            logger.info(f"🔄 Changes: +{len(new_futures)}, -{len(lost_futures)}, Total: {len(current_unique_set)}")
            
            return new_futures, lost_futures
            
        except Exception as e:
            logger.error(f"Error monitoring unique futures: {e}")
            return set(), set()


            
    def format_change(self, change):
        """Format price change with color emoji"""
        if change > 0:
            return f"🟢 +{change:.2f}%"
        elif change < 0:
            return f"🔴 {change:.2f}%"
        else:
            return f"⚪ {change:.2f}%"

    # ==================== ENHANCED GOOGLE SHEETS ====================




    def format_change_for_sheet(self, change):
        """Format change for Google Sheets with proper None handling"""
        if change is None:
            return 'N/A'
        
        # Add emoji based on change value
        if change > 10:
            return f"🚀 {change:+.2f}%"
        elif change > 5:
            return f"🟢 {change:+.2f}%"
        elif change > 2:
            return f"📈 {change:+.2f}%"
        elif change < -10:
            return f"💥 {change:+.2f}%"
        elif change < -5:
            return f"🔴 {change:+.2f}%"
        elif change < -2:
            return f"📉 {change:+.2f}%"
        else:
            return f"{change:+.2f}%"

    def calculate_trend_score(self, changes):
        """Calculate overall trend score from multiple timeframes"""
        timeframes = ['5m', '15m', '30m', '60m', '240m']
        weights = [2.0, 1.5, 1.2, 1.0, 0.5]  # Higher weight for shorter timeframes
        
        total_score = 0
        total_weight = 0
        
        for i, timeframe in enumerate(timeframes):
            change = changes.get(timeframe)
            if change is not None:
                total_score += change * weights[i]
                total_weight += weights[i]
        
        return total_score / total_weight if total_weight > 0 else 0




    def create_and_send_excel(self, update: Update, context: CallbackContext):
        """Create and send Excel file via Telegram - FIXED to use same method as check"""
        try:
            update.message.reply_html("📊 <b>Creating comprehensive Excel report...</b>")
            
            # Get all the data needed
            all_futures_data = []
            exchanges = {
                'MEXC': self.get_mexc_futures,
                'Binance': self.get_binance_futures,
                'Bybit': self.get_bybit_futures,
                'OKX': self.get_okx_futures,
                'Gate.io': self.get_gate_futures,
                'KuCoin': self.get_kucoin_futures,
                'BingX': self.get_bingx_futures,
                'BitGet': self.get_bitget_futures
            }
            
            symbol_coverage = {}
            current_time = datetime.now().isoformat()
            
            # Collect data from all exchanges
            for name, method in exchanges.items():
                try:
                    futures = method()
                    for symbol in futures:
                        all_futures_data.append({
                            'symbol': symbol,
                            'exchange': name,
                            'timestamp': current_time
                        })
                        
                        # Track symbol coverage
                        normalized = self.normalize_symbol_for_comparison(symbol)
                        if normalized not in symbol_coverage:
                            symbol_coverage[normalized] = set()
                        symbol_coverage[normalized].add(name)
                        
                except Exception as e:
                    logger.error(f"Error getting {name} data: {e}")
            
            # Get unique futures
            unique_futures, exchange_stats = self.find_unique_futures_robust()
            
            # FIX: Use the EXACT SAME approach as check command
            batch_data = self.get_consistent_price_data()
            logger.info(f"📊 Excel - Batch data: {len(batch_data)} symbols")
            
            # Create price_data by matching unique symbols with batch data (SAME AS CHECK)
            price_data = {}
            matched_symbols = 0
            
            for symbol in unique_futures:
                # Try exact match first
                if symbol in batch_data:
                    price_data[symbol] = batch_data[symbol]
                    matched_symbols += 1
                else:
                    # Try alternative formats (SAME AS CHECK)
                    alt_formats = [
                        symbol.replace('_', ''),
                        symbol.replace('_', '-'), 
                        symbol.replace('_', '/'),
                    ]
                    
                    found = False
                    for alt_format in alt_formats:
                        if alt_format in batch_data:
                            price_data[symbol] = batch_data[alt_format].copy()
                            price_data[symbol]['symbol'] = symbol  # Fix symbol name
                            matched_symbols += 1
                            found = True
                            break
                    
                    if not found:
                        # Symbol not found in batch, add with None price
                        price_data[symbol] = {
                            'symbol': symbol,
                            'price': None,
                            'changes': {},
                            'timestamp': datetime.now(),
                            'source': 'not_found'
                        }
            
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # DEBUG: Log what we found
            logger.info(f"🔍 Excel - Price coverage: {matched_symbols}/{len(unique_futures)} ({matched_symbols/len(unique_futures)*100:.1f}%)")
            
            # Create Excel file
            excel_content = self.create_mexc_analysis_excel(all_futures_data, symbol_coverage, analyzed_prices)
            
            if not excel_content:
                update.message.reply_html("❌ <b>Failed to create Excel file</b>")
                return
            
            # Send file via Telegram
            filename = f"mexc_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            update.message.reply_document(
                document=io.BytesIO(excel_content),
                filename=filename,
                caption=(
                    f"📊 <b>MEXC Futures Analysis Report</b>\n\n"
                    f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                    f"🎯 Unique Futures: {len(unique_futures)}\n"
                    f"🏢 Exchanges: 8\n"
                    f"💰 Price Data: {matched_symbols}/{len(unique_futures)} ({matched_symbols/len(unique_futures)*100:.1f}%)\n\n"
                    f"<i>Sheets included: Dashboard, Unique Futures, All Futures, MEXC Analysis, Price Analysis, Exchange Stats</i>"
                ),
                parse_mode='HTML'
            )
            
            logger.info("✅ Excel file sent successfully")
            
        except Exception as e:
            error_msg = f"❌ <b>Error creating Excel file:</b>\n<code>{str(e)}</code>"
            update.message.reply_html(error_msg)
            logger.error(f"Excel creation error: {e}")

    def setup_google_sheets_historical_storage(self):
        """Setup historical data storage in Google Sheets"""
        try:
            if not self.spreadsheet:
                return False
            
            # Create or get Historical Data sheet
            try:
                self.historical_worksheet = self.spreadsheet.worksheet('Historical Data')
            except gspread.WorksheetNotFound:
                self.historical_worksheet = self.spreadsheet.add_worksheet(
                    title='Historical Data', 
                    rows=10000, 
                    cols=10
                )
                # Set up headers
                headers = [
                    'Timestamp', 'Symbol', 'Price', '5m Change %', '15m Change %', 
                    '30m Change %', '1h Change %', '4h Change %', 'Source'
                ]
                self.historical_worksheet.update([headers], 'A1')
                logger.info("✅ Created Historical Data sheet")
            
            # Create or get Price History sheet for raw price data
            try:
                self.price_history_worksheet = self.spreadsheet.worksheet('Price History')
            except gspread.WorksheetNotFound:
                self.price_history_worksheet = self.spreadsheet.add_worksheet(
                    title='Price History',
                    rows=50000,
                    cols=5
                )
                headers = ['Timestamp', 'Symbol', 'Price', 'Source', 'Batch ID']
                self.price_history_worksheet.update([headers], 'A1')
                logger.info("✅ Created Price History sheet")
            
            return True
            
        except Exception as e:
            logger.error(f"Error setting up historical storage: {e}")
            return False

    def store_price_history(self, price_data):
        """Store current price data to Google Sheets for historical analysis"""
        try:
            if not hasattr(self, 'price_history_worksheet'):
                if not self.setup_google_sheets_historical_storage():
                    return
            
            current_time = datetime.now()
            batch_id = current_time.strftime('%Y%m%d_%H%M')
            
            # Prepare data for storage
            rows_to_store = []
            for symbol, data in price_data.items():
                price = data.get('price')
                if price is None:
                    continue
                    
                rows_to_store.append([
                    current_time.isoformat(),
                    symbol,
                    price,
                    data.get('source', 'unknown'),
                    batch_id
                ])
            
            # Store in batches to avoid timeout
            if rows_to_store:
                batch_size = 100
                for i in range(0, len(rows_to_store), batch_size):
                    batch = rows_to_store[i:i + batch_size]
                    self.price_history_worksheet.append_rows(batch)
                
                logger.info(f"💾 Stored {len(rows_to_store)} price records to Google Sheets")
                
        except Exception as e:
            logger.error(f"Error storing price history: {e}")

    def calculate_historical_changes_from_sheets(self, symbol, current_price):
        """Calculate historical changes using data from Google Sheets"""
        try:
            if not hasattr(self, 'price_history_worksheet'):
                return {}
            
            current_time = datetime.now()
            changes = {}
            
            # Define timeframes to calculate
            timeframes = [
                ('5m', timedelta(minutes=5)),
                ('15m', timedelta(minutes=15)),
                ('30m', timedelta(minutes=30)),
                ('60m', timedelta(hours=1)),
                ('240m', timedelta(hours=4))
            ]
            
            # Get recent price history for this symbol from Google Sheets
            historical_prices = self.get_symbol_price_history(symbol)
            
            for timeframe_name, time_delta in timeframes:
                target_time = current_time - time_delta
                historical_price = self.find_closest_historical_price(historical_prices, target_time)
                
                if historical_price and historical_price > 0:
                    price_change = ((current_price - historical_price) / historical_price) * 100
                    changes[timeframe_name] = price_change
                else:
                    changes[timeframe_name] = None
            
            return changes
            
        except Exception as e:
            logger.error(f"Error calculating historical changes from sheets for {symbol}: {e}")
            return {}

    def get_symbol_price_history(self, symbol, hours_back=24):
        """Get price history for a specific symbol from Google Sheets"""
        try:
            if not hasattr(self, 'price_history_worksheet'):
                return []
            
            # Get all records for this symbol
            all_records = self.price_history_worksheet.get_all_records()
            
            # Filter for this symbol and recent data
            cutoff_time = datetime.now() - timedelta(hours=hours_back)
            symbol_history = []
            
            for record in all_records:
                if record.get('Symbol') == symbol:
                    try:
                        timestamp = datetime.fromisoformat(record.get('Timestamp', ''))
                        if timestamp > cutoff_time:
                            symbol_history.append({
                                'timestamp': timestamp,
                                'price': float(record.get('Price', 0)),
                                'source': record.get('Source', '')
                            })
                    except (ValueError, TypeError):
                        continue
            
            # Sort by timestamp
            symbol_history.sort(key=lambda x: x['timestamp'])
            return symbol_history
            
        except Exception as e:
            logger.error(f"Error getting price history for {symbol}: {e}")
            return []

    def find_closest_historical_price(self, historical_prices, target_time):
        """Find the closest historical price to target time"""
        if not historical_prices:
            return None
        
        closest_record = None
        min_time_diff = timedelta.max
        
        for record in historical_prices:
            time_diff = abs(record['timestamp'] - target_time)
            if time_diff < min_time_diff:
                min_time_diff = time_diff
                closest_record = record
        
        # Only return if within reasonable time window (2x the timeframe)
        max_allowed_diff = timedelta(hours=2)
        if min_time_diff < max_allowed_diff:
            return closest_record['price']
        
        return None

    def store_calculated_changes(self, analyzed_prices):
        """Store calculated changes to Historical Data sheet"""
        try:
            if not hasattr(self, 'historical_worksheet'):
                if not self.setup_google_sheets_historical_storage():
                    return
            
            current_time = datetime.now()
            rows_to_store = []
            
            for item in analyzed_prices:
                symbol = item['symbol']
                price = item.get('price')
                changes = item.get('changes', {})
                
                if price is None:
                    continue
                
                rows_to_store.append([
                    current_time.isoformat(),
                    symbol,
                    price,
                    changes.get('5m'),
                    changes.get('15m'),
                    changes.get('30m'),
                    changes.get('60m'),
                    changes.get('240m'),
                    item.get('source', 'calculated')
                ])
            
            if rows_to_store:
                # Store in batches
                batch_size = 100
                for i in range(0, len(rows_to_store), batch_size):
                    batch = rows_to_store[i:i + batch_size]
                    self.historical_worksheet.append_rows(batch)
                
                logger.info(f"💾 Stored {len(rows_to_store)} calculated changes to Historical Data")
                
        except Exception as e:
            logger.error(f"Error storing calculated changes: {e}")

    def cleanup_old_price_data(self):
        """Clean up old price data to prevent sheets from getting too large"""
        try:
            if not hasattr(self, 'price_history_worksheet'):
                return
            
            # Keep only last 7 days of data
            cutoff_time = datetime.now() - timedelta(days=7)
            
            # Get all records
            all_records = self.price_history_worksheet.get_all_records()
            records_to_keep = []
            
            for record in all_records:
                try:
                    timestamp = datetime.fromisoformat(record.get('Timestamp', ''))
                    if timestamp > cutoff_time:
                        records_to_keep.append(record)
                except (ValueError, TypeError):
                    continue
            
            # If we have records to delete, recreate the sheet
            if len(records_to_keep) < len(all_records):
                # Clear and repopulate with recent data
                self.price_history_worksheet.clear()
                headers = ['Timestamp', 'Symbol', 'Price', 'Source', 'Batch ID']
                self.price_history_worksheet.update([headers], 'A1')
                
                # Convert records back to rows
                rows_to_keep = []
                for record in records_to_keep:
                    rows_to_keep.append([
                        record.get('Timestamp'),
                        record.get('Symbol'),
                        record.get('Price'),
                        record.get('Source'),
                        record.get('Batch ID')
                    ])
                
                # Store in batches
                if rows_to_keep:
                    batch_size = 100
                    for i in range(0, len(rows_to_keep), batch_size):
                        batch = rows_to_keep[i:i + batch_size]
                        self.price_history_worksheet.append_rows(batch)
                
                deleted_count = len(all_records) - len(records_to_keep)
                logger.info(f"🧹 Cleaned up {deleted_count} old price records")
                
        except Exception as e:
            logger.error(f"Error cleaning up old price data: {e}")

    def update_exchange_stats_sheet(self, spreadsheet, exchange_stats, timestamp):
        """Update Exchange Stats sheet with proper data"""
        try:
            worksheet = spreadsheet.worksheet('Exchange Stats')
            
            # Clear existing data
            worksheet.clear()
            
            # Headers
            headers = [
                'Exchange', 'Futures Count', 'Status', 'Last Updated', 
                'Success Rate', 'API Health'
            ]
            worksheet.update('A1', [headers])
            
            # Get actual exchange data
            actual_stats = self.get_all_exchanges_futures_stats()
            
            stats_data = []
            for exchange, count in actual_stats.items():
                status = "✅ WORKING" if count > 0 else "❌ FAILED"
                success_rate = "100%" if count > 0 else "0%"
                health = "🟢 HEALTHY" if count > 0 else "🔴 OFFLINE"
                
                stats_data.append([
                    exchange,
                    count,
                    status,
                    timestamp,
                    success_rate,
                    health
                ])
            
            if stats_data:
                worksheet.update('A2', stats_data)
                logger.info(f"✅ Updated Exchange Stats with {len(stats_data)} records")
            
            # Apply formatting for better visualization
            try:
                # Format status column with colors
                worksheet.format('C2:C100', {
                    "backgroundColor": {
                        "red": 0.9, "green": 0.98, "blue": 0.9
                    }
                })
            except Exception as format_error:
                logger.warning(f"Could not format Exchange Stats: {format_error}")
                
        except Exception as e:
            logger.error(f"Error updating Exchange Stats sheet: {e}")

    def get_all_exchanges_futures_stats(self):
        """Get actual futures count from all exchanges"""
        try:
            exchange_stats = {}
            exchanges = {
                'MEXC': self.get_mexc_futures,
                'Binance': self.get_binance_futures,
                'Bybit': self.get_bybit_futures,
                'OKX': self.get_okx_futures,
                'Gate.io': self.get_gate_futures,
                'KuCoin': self.get_kucoin_futures,
                'BingX': self.get_bingx_futures,
                'BitGet': self.get_bitget_futures
            }
            
            for name, method in exchanges.items():
                try:
                    futures = method()
                    exchange_stats[name] = len(futures)
                    logger.info(f"✅ {name}: {len(futures)} futures")
                except Exception as e:
                    exchange_stats[name] = 0
                    logger.error(f"❌ {name} error: {e}")
            
            return exchange_stats
            
        except Exception as e:
            logger.error(f"Error getting exchange stats: {e}")
            return {}
        



    def format_change_for_sheets_with_colors(self, change):
        """Format change for Google Sheets with proper +/- signs"""
        if change is None:
            return 'N/A'
        
        # Always show + sign for positive changes, - sign for negative
        if change > 0:
            return f"+{change:.2f}%"
        else:
            return f"{change:.2f}%"  # Negative numbers already have - sign

    def apply_color_formatting_to_sheets(self, worksheet, data_rows_count):
        """Apply color formatting to Google Sheets for better visualization - FIXED"""
        try:
            if data_rows_count == 0:
                return
                
            # Define the range for change columns (C to G for 5m to 4h changes)
            change_columns = ['C', 'D', 'E', 'F', 'G']
            
            # Prepare all format requests
            format_requests = []
            
            for col in change_columns:
                range_str = f"{col}2:{col}{data_rows_count + 1}"
                
                # Positive changes (contains "+")
                format_requests.append({
                    'addConditionalFormatRule': {
                        'rule': {
                            'ranges': [{'sheetId': worksheet.id, 'startRowIndex': 1, 'endRowIndex': data_rows_count + 1, 
                                    'startColumnIndex': ord(col) - 65, 'endColumnIndex': ord(col) - 64}],
                            'booleanRule': {
                                'condition': {
                                    'type': 'TEXT_CONTAINS',
                                    'values': [{'userEnteredValue': '+'}]
                                },
                                'format': {
                                    'backgroundColor': {'red': 0.85, 'green': 0.98, 'blue': 0.85},
                                    'textFormat': {
                                        'foregroundColor': {'red': 0.0, 'green': 0.5, 'blue': 0.0},
                                        'bold': True
                                    }
                                }
                            }
                        }
                    }
                })
                
                # Negative changes (contains "-")
                format_requests.append({
                    'addConditionalFormatRule': {
                        'rule': {
                            'ranges': [{'sheetId': worksheet.id, 'startRowIndex': 1, 'endRowIndex': data_rows_count + 1,
                                    'startColumnIndex': ord(col) - 65, 'endColumnIndex': ord(col) - 64}],
                            'booleanRule': {
                                'condition': {
                                    'type': 'TEXT_CONTAINS',
                                    'values': [{'userEnteredValue': '-'}]
                                },
                                'format': {
                                    'backgroundColor': {'red': 1.0, 'green': 0.9, 'blue': 0.9},
                                    'textFormat': {
                                        'foregroundColor': {'red': 0.8, 'green': 0.0, 'blue': 0.0},
                                        'bold': True
                                    }
                                }
                            }
                        }
                    }
                })
                
                # Neutral (N/A values)
                format_requests.append({
                    'addConditionalFormatRule': {
                        'rule': {
                            'ranges': [{'sheetId': worksheet.id, 'startRowIndex': 1, 'endRowIndex': data_rows_count + 1,
                                    'startColumnIndex': ord(col) - 65, 'endColumnIndex': ord(col) - 64}],
                            'booleanRule': {
                                'condition': {
                                    'type': 'TEXT_CONTAINS',
                                    'values': [{'userEnteredValue': 'N/A'}]
                                },
                                'format': {
                                    'backgroundColor': {'red': 0.95, 'green': 0.95, 'blue': 0.95},
                                    'textFormat': {
                                        'foregroundColor': {'red': 0.3, 'green': 0.3, 'blue': 0.3}
                                    }
                                }
                            }
                        }
                    }
                })
            
            # Score column gradient (column H)
            score_col_index = ord('H') - 65  # H is column 7
            format_requests.append({
                'addConditionalFormatRule': {
                    'rule': {
                        'ranges': [{'sheetId': worksheet.id, 'startRowIndex': 1, 'endRowIndex': data_rows_count + 1,
                                'startColumnIndex': score_col_index, 'endColumnIndex': score_col_index + 1}],
                        'gradientRule': {
                            'minpoint': {
                                'color': {'red': 1.0, 'green': 0.9, 'blue': 0.9},  # Light red
                                'type': 'MIN'
                            },
                            'midpoint': {
                                'color': {'red': 1.0, 'green': 1.0, 'blue': 0.9},  # Light yellow
                                'type': 'NUMBER',
                                'value': '0'
                            },
                            'maxpoint': {
                                'color': {'red': 0.85, 'green': 0.98, 'blue': 0.85},  # Light green
                                'type': 'MAX'
                            }
                        }
                    }
                }
            })
            
            # Apply all formatting requests
            if format_requests:
                worksheet.spreadsheet.batch_update({'requests': format_requests})
                logger.info(f"🎨 Applied {len(format_requests)} color formatting rules")
            
        except Exception as e:
            logger.error(f"Error applying color formatting: {e}")
    
    
    def update_dashboard_with_comprehensive_stats(self, exchange_stats, unique_symbols_count, unique_futures_count, analyzed_prices):
        """Update dashboard with comprehensive statistics"""
        if not self.spreadsheet:
            return
        
        try:
            worksheet = self.spreadsheet.worksheet("Dashboard")
            
            # Get actual exchange statistics
            actual_exchange_stats = self.get_all_exchanges_futures_stats()
            working_exchanges = sum(1 for count in actual_exchange_stats.values() if count > 0)
            total_exchanges = len(actual_exchange_stats)
            total_futures = sum(actual_exchange_stats.values())
            
            # Calculate price statistics
            valid_prices = len([p for p in analyzed_prices if p.get('price') is not None]) if analyzed_prices else 0
            price_coverage = (valid_prices / unique_futures_count) * 100 if unique_futures_count > 0 else 0
            
            # Get top performers
            top_performers = []
            if analyzed_prices:
                sorted_prices = sorted(analyzed_prices, key=lambda x: x.get('score', 0), reverse=True)
                top_performers = sorted_prices[:3]
            
            stats_update = [
                ["🤖 MEXC FUTURES AUTO-UPDATE DASHBOARD", ""],
                ["Last Updated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Update Interval", f"{self.update_interval} minutes"],
                ["Price Check Interval", f"{self.price_check_interval} minutes"],
                ["", ""],
                ["📊 EXCHANGE STATISTICS", ""],
                ["Working Exchanges", f"{working_exchanges}/{total_exchanges}"],
                ["Total Futures", f"{total_futures}"],
                ["MEXC Futures", f"{actual_exchange_stats.get('MEXC', 0)}"],
                ["Unique MEXC Futures", f"{unique_futures_count} 🎯"],
                ["", ""],
                ["💰 PRICE ANALYSIS", ""],
                ["Symbols with Price Data", f"{valid_prices}/{unique_futures_count}"],
                ["Price Data Coverage", f"{price_coverage:.1f}%"],
                ["Top Performer Score", f"{top_performers[0].get('score', 0):.1f}" if top_performers else "N/A"],
                ["Market Sentiment", self.get_market_sentiment(analyzed_prices)],
                ["", ""],
                ["⚡ PERFORMANCE", ""],
                ["Next Data Update", (datetime.now() + timedelta(minutes=self.update_interval)).strftime('%H:%M:%S')],
                ["Next Price Update", (datetime.now() + timedelta(minutes=self.price_check_interval)).strftime('%H:%M:%S')],
                ["Next 4h Chart", (datetime.now() + timedelta(hours=4)).strftime('%H:%M:%S')],
                ["Status", "🟢 RUNNING"],
            ]
            
            # Update dashboard
            worksheet.clear()
            worksheet.update(stats_update, 'A1')
            
            # Apply dashboard formatting
            self.apply_dashboard_formatting(worksheet)
            
            logger.info("✅ Dashboard updated with comprehensive stats")
            
        except Exception as e:
            logger.error(f"Error updating dashboard: {e}")

    def apply_dashboard_formatting(self, worksheet):
        """Apply formatting to dashboard"""
        try:
            # Format title row
            worksheet.format('A1:B1', {
                'textFormat': {'bold': True, 'fontSize': 14},
                'backgroundColor': {'red': 0.8, 'green': 0.9, 'blue': 1.0},
                'horizontalAlignment': 'CENTER'
            })
            
            # Format section headers
            section_rows = [6, 12, 18]  # Adjust based on your actual row numbers
            for row in section_rows:
                worksheet.format(f'A{row}:B{row}', {
                    'textFormat': {'bold': True},
                    'backgroundColor': {'red': 0.95, 'green': 0.95, 'blue': 0.95}
                })
                
        except Exception as e:
            logger.warning(f"Could not format dashboard: {e}")

    def get_market_sentiment(self, analyzed_prices):
        """Calculate overall market sentiment"""
        if not analyzed_prices:
            return "N/A"
        
        positive_symbols = 0
        total_symbols = 0
        
        for item in analyzed_prices:
            changes = item.get('changes', {})
            # Use 1h change for sentiment
            change_1h = changes.get('60m')
            if change_1h is not None:
                total_symbols += 1
                if change_1h > 0:
                    positive_symbols += 1
        
        if total_symbols == 0:
            return "N/A"
        
        sentiment_ratio = positive_symbols / total_symbols
        
        if sentiment_ratio > 0.7:
            return "🟢 BULLISH"
        elif sentiment_ratio > 0.5:
            return "🟡 NEUTRAL"
        else:
            return "🔴 BEARISH"


    def apply_color_formatting_to_all_sheets(self):
        """Apply color formatting to all relevant sheets"""
        try:
            if not self.spreadsheet:
                return
                
            sheets_to_format = ['Unique Futures', 'Price Analysis', 'MEXC Analysis']
            
            for sheet_name in sheets_to_format:
                try:
                    worksheet = self.spreadsheet.worksheet(sheet_name)
                    data = worksheet.get_all_values()
                    
                    if len(data) > 1:  # Has data beyond headers
                        data_rows_count = len(data) - 1
                        self.apply_simple_color_formatting(worksheet, data_rows_count)
                        logger.info(f"🎨 Applied formatting to {sheet_name}")
                        
                except Exception as e:
                    logger.warning(f"Could not format {sheet_name}: {e}")
                    
        except Exception as e:
            logger.error(f"Error applying color formatting to all sheets: {e}")

    def apply_simple_color_formatting(self, worksheet, data_rows_count):
        """Apply simple color formatting to a worksheet"""
        try:
            if data_rows_count == 0:
                return
                
            # Get all data to check values
            all_data = worksheet.get_all_values()
            if len(all_data) < 2:
                return
                
            # Define which columns contain percentage changes for each sheet
            sheet_columns = {
                'Unique Futures': ['C', 'D', 'E', 'F', 'G'],  # 5m, 15m, 30m, 1h, 4h changes
                'Price Analysis': ['D', 'E', 'F', 'G', 'H'],  # 5m, 15m, 30m, 1h, 4h changes
                'MEXC Analysis': ['F', 'G', 'H']  # 5m, 1h, 4h changes
            }
            
            sheet_name = worksheet.title
            change_columns = sheet_columns.get(sheet_name, [])
            
            if not change_columns:
                return
                
            # Format each row
            for row_idx in range(2, data_rows_count + 2):  # Start from row 2 (after headers)
                try:
                    # Check the first change column to determine the trend
                    first_col = change_columns[0]
                    cell_value = worksheet.acell(f'{first_col}{row_idx}').value
                    
                    if cell_value and cell_value != 'N/A':
                        # Create range for all change columns in this row
                        col_range = f'{change_columns[0]}{row_idx}:{change_columns[-1]}{row_idx}'
                        
                        if '+' in cell_value:
                            # Positive - green background
                            worksheet.format(col_range, {
                                "backgroundColor": {"red": 0.85, "green": 0.98, "blue": 0.85},
                                "horizontalAlignment": "CENTER"
                            })
                        elif '-' in cell_value:
                            # Negative - red background
                            worksheet.format(col_range, {
                                "backgroundColor": {"red": 1.0, "green": 0.9, "blue": 0.9},
                                "horizontalAlignment": "CENTER"
                            })
                        else:
                            # Neutral - gray background
                            worksheet.format(col_range, {
                                "backgroundColor": {"red": 0.95, "green": 0.95, "blue": 0.95},
                                "horizontalAlignment": "CENTER"
                            })
                            
                except Exception as row_error:
                    continue
                    
        except Exception as e:
            logger.error(f"Error in simple color formatting: {e}")

    def update_unique_futures_sheet_with_prices(self, unique_futures, analyzed_prices):
        """Update Unique Futures sheet with colorful formatting"""
        try:
            worksheet = self.spreadsheet.worksheet('Unique Futures')
            
            # Clear existing data
            worksheet.clear()
            
            headers = [
                'Symbol', 'Current Price', '5m Change %', '15m Change %', 
                '30m Change %', '1h Change %', '4h Change %', 'Score', 'Status', 'Last Updated'
            ]
            worksheet.update([headers], 'A1')
            
            sheet_data = []
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            price_map = {item['symbol']: item for item in analyzed_prices}
            
            for symbol in sorted(unique_futures):
                price_info = price_map.get(symbol)
                changes = price_info.get('changes', {}) if price_info else {}
                price = price_info.get('price') if price_info else None
                
                # Format price
                if price is not None:
                    if price >= 1:
                        price_display = f"${price:.4f}"
                    elif price >= 0.01:
                        price_display = f"${price:.6f}"
                    else:
                        price_display = f"${price:.8f}"
                else:
                    price_display = 'N/A'
                
                # Format changes with clear +/-
                row = [
                    symbol,
                    price_display,
                    self.format_change_with_sign(changes.get('5m')),
                    self.format_change_with_sign(changes.get('15m')),
                    self.format_change_with_sign(changes.get('30m')),
                    self.format_change_with_sign(changes.get('60m')),
                    self.format_change_with_sign(changes.get('240m')),
                    f"{price_info.get('score', 0):.2f}" if price_info else 'N/A',
                    'UNIQUE',
                    current_time
                ]
                sheet_data.append(row)
            
            if sheet_data:
                worksheet.update(f'A2:J{len(sheet_data) + 1}', sheet_data)
                logger.info(f"✅ Updated Unique Futures with {len(sheet_data)} records")
                
                # Apply color formatting
                self.apply_simple_color_formatting(worksheet, len(sheet_data))
                
        except Exception as e:
            logger.error(f"Error updating Unique Futures sheet: {e}")



    def format_change_with_sign(self, change):
        """Format change with clear + sign for positive, - for negative"""
        if change is None:
            return 'N/A'
        
        # Always show + for positive numbers
        if change > 0:
            return f"+{change:.2f}%"
        else:
            return f"{change:.2f}%"  # Negative numbers already have -


    @rate_limit_google_sheets(calls_per_minute=45)
    def update_google_sheet_with_prices(self):
        """Update Google Sheet with prices - FIXED with better logging"""
        try:
            current_time = time.time()
            if hasattr(self, '_last_sheets_call'):
                time_since_last = current_time - self._last_sheets_call
                if time_since_last < 60:  # Wait at least 60 seconds between calls
                    logger.info(f"⏸️ Rate limiting Sheets API call ({time_since_last:.1f}s since last)")
                    return
        
            if not self.gs_client or not self.spreadsheet:
                logger.warning("❌ Google Sheets not available")
                return
            
            logger.info("🔄 Starting Google Sheet update...")
            
            # Get unique futures
            unique_futures, _ = self.find_unique_futures_robust()
            logger.info(f"🎯 Found {len(unique_futures)} unique futures")
            
            # Get current price data
            batch_data = self.get_mexc_prices_batch_working()
            logger.info(f"📊 Got {len(batch_data)} prices from batch API")
            
            # Store price history
            self.store_price_history(batch_data)
            
            # Create price_data with proper historical changes
            price_data = {}
            matched_symbols = 0
            
            for symbol in unique_futures:
                current_price_info = None
                
                # Find current price
                if symbol in batch_data:
                    current_price_info = batch_data[symbol]
                else:
                    # Try alternative formats
                    alt_formats = [
                        symbol.replace('_', ''),
                        symbol.replace('_', '-'), 
                        symbol.replace('_', '/'),
                    ]
                    for alt_format in alt_formats:
                        if alt_format in batch_data:
                            current_price_info = batch_data[alt_format].copy()
                            current_price_info['symbol'] = symbol
                            break
                
                if current_price_info and current_price_info.get('price') is not None:
                    current_price = current_price_info['price']
                    
                    # Calculate proper historical changes
                    historical_changes = self.calculate_historical_changes_from_sheets(symbol, current_price)
                    
                    price_data[symbol] = {
                        'symbol': symbol,
                        'price': current_price,
                        'changes': historical_changes,
                        'timestamp': datetime.now(),
                        'source': current_price_info.get('source', 'historical')
                    }
                    matched_symbols += 1
                else:
                    price_data[symbol] = {
                        'symbol': symbol,
                        'price': None,
                        'changes': {},
                        'timestamp': datetime.now(),
                        'source': 'not_found'
                    }
            
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Get exchange statistics
            exchange_stats = self.get_all_exchanges_futures_stats()
            
            # Update all sheets
            self.update_unique_futures_sheet_with_prices(unique_futures, analyzed_prices)
            self.update_price_analysis_sheet(analyzed_prices)
            self.update_exchange_stats_sheet(self.spreadsheet, exchange_stats, datetime.now().isoformat())
            
            # Store calculated changes
            self.store_calculated_changes(analyzed_prices)
            
            logger.info(f"✅ Google Sheets updated: {matched_symbols}/{len(unique_futures)} prices")
            
            # Send confirmation to Telegram
            if matched_symbols > 0:
                confirmation_msg = (
                    f"📊 <b>Google Sheets Updated</b>\n\n"
                    f"✅ Prices updated: {matched_symbols}/{len(unique_futures)}\n"
                    f"💰 Coverage: {(matched_symbols/len(unique_futures)*100):.1f}%\n"
                    f"🕒 Time: {datetime.now().strftime('%H:%M:%S')}\n\n"
                    f"<i>Next update in 5 minutes</i>"
                )
                self.send_broadcast_message(confirmation_msg)
            
        except Exception as e:
            logger.error(f"❌ Error updating Google Sheets: {e}")
            # Send error to Telegram
            error_msg = f"❌ <b>Google Sheets Update Failed</b>\n\nError: {str(e)}"
            self.send_broadcast_message(error_msg)

    def get_cached_sheets_data(self):
        """Get cached Google Sheets data to reduce API calls"""
        current_time = time.time()
        
        if (current_time - self.sheets_cache_time < self.sheets_cache_duration and 
            self.sheets_batch_data):
            return self.sheets_batch_data
            
        # Fetch fresh data and cache it
        try:
            # Your data fetching logic
            self.sheets_batch_data = self.fetch_all_sheets_data()
            self.sheets_cache_time = current_time
            return self.sheets_batch_data
        except Exception as e:
            logger.error(f"Failed to cache Sheets data: {e}")
            return {}


    def batch_sheets_operations(self, operations):
        """Batch multiple Sheets operations into one API call"""
        try:
            if not operations:
                return
                
            # Group operations by type and execute in batches
            batch_size = 10  # Small batches to avoid timeouts
            for i in range(0, len(operations), batch_size):
                batch = operations[i:i + batch_size]
                self.execute_sheets_batch(batch)
                time.sleep(1)  # Rate limiting between batches
                
        except Exception as e:
            logger.error(f"Batch Sheets operations failed: {e}")

    def execute_sheets_batch(self, batch_operations):
        """Execute a batch of Sheets operations"""
        try:
            if not self.spreadsheet:
                return
                
            # Use batchUpdate for multiple operations
            requests = []
            for operation in batch_operations:
                if operation['type'] == 'update':
                    requests.append({
                        'updateCells': {
                            'range': operation['range'],
                            'rows': operation['rows'],
                            'fields': 'userEnteredValue,userEnteredFormat'
                        }
                    })
            
            if requests:
                self.spreadsheet.batch_update({'requests': requests})
                
        except Exception as e:
            logger.error(f"Batch execution failed: {e}")




    def handle_sheets_api_error(self, error, operation_name):
        """Handle Google Sheets API errors with exponential backoff"""
        if '429' in str(error):
            logger.warning(f"Google Sheets quota exceeded for {operation_name}")
            # Implement exponential backoff
            wait_time = min(300, 2 ** self.retry_count)  # Max 5 minutes
            logger.info(f"Waiting {wait_time} seconds before retry")
            time.sleep(wait_time)
            self.retry_count += 1
            return True
        else:
            logger.error(f"Google Sheets error in {operation_name}: {error}")
            self.retry_count = 0
            return False

    def optimized_data_flow(self):
        """Optimized data flow to minimize API calls"""
        try:
            # Step 1: Get all data in one go
            unique_futures, exchange_stats = self.find_unique_futures_robust()
            price_data = self.get_consistent_price_data()
            
            # Step 2: Analyze locally
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Step 3: Prepare all Sheets data
            sheets_operations = self.prepare_all_sheets_operations(
                unique_futures, exchange_stats, analyzed_prices
            )
            
            # Step 4: Execute in one batch (with rate limiting)
            self.batch_sheets_operations(sheets_operations)
            
            # Step 5: Store historical data (with caching)
            self.store_historical_data_with_cache(analyzed_prices)
            
        except Exception as e:
            logger.error(f"Optimized data flow failed: {e}")


    def create_historical_trends_sheet(self, wb, historical_data):
        """Create Historical Trends sheet showing price movement patterns"""
        if not historical_data:
            return
            
        ws = wb.create_sheet("Historical Trends")
        
        # Headers for trend analysis
        headers = [
            'Symbol', 'Current Price', 'Trend Direction', 'Volatility Score',
            '5m Trend', '15m Trend', '30m Trend', '1h Trend', '4h Trend',
            'Best Timeframe', 'Worst Timeframe', 'Consistency Score'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Analyze trends for each symbol
        row = 2
        for symbol, data in sorted(historical_data.items()):
            changes = [
                data.get('change_5m'),
                data.get('change_15m'), 
                data.get('change_30m'),
                data.get('change_1h'),
                data.get('change_4h')
            ]
            
            # Filter out None values
            valid_changes = [ch for ch in changes if ch is not None]
            
            if not valid_changes:
                continue
                
            # Calculate trend metrics
            avg_change = sum(valid_changes) / len(valid_changes)
            volatility = max(valid_changes) - min(valid_changes) if len(valid_changes) > 1 else 0
            
            # Determine trend direction
            if avg_change > 5:
                trend_direction = "STRONG UP 🚀"
            elif avg_change > 2:
                trend_direction = "UP 🟢"
            elif avg_change < -5:
                trend_direction = "STRONG DOWN 🔻"
            elif avg_change < -2:
                trend_direction = "DOWN 🔴"
            else:
                trend_direction = "FLAT ⚪"
            
            # Find best and worst timeframes
            timeframes = ['5m', '15m', '30m', '1h', '4h']
            timeframe_changes = list(zip(timeframes, changes))
            
            # Filter out None values for best/worst calculation
            valid_timeframes = [(tf, ch) for tf, ch in timeframe_changes if ch is not None]
            
            if valid_timeframes:
                best_timeframe = max(valid_timeframes, key=lambda x: x[1])[0]
                worst_timeframe = min(valid_timeframes, key=lambda x: x[1])[0]
            else:
                best_timeframe = "N/A"
                worst_timeframe = "N/A"
            
            # Calculate consistency (how many timeframes show the same direction)
            positive_changes = len([ch for ch in valid_changes if ch > 0])
            negative_changes = len([ch for ch in valid_changes if ch < 0])
            consistency = max(positive_changes, negative_changes) / len(valid_changes) * 100 if valid_changes else 0
            
            # Add row data
            ws.cell(row=row, column=1).value = symbol
            ws.cell(row=row, column=2).value = data.get('current_price', 'N/A')
            ws.cell(row=row, column=3).value = trend_direction
            ws.cell(row=row, column=4).value = f"{volatility:.2f}"
            ws.cell(row=row, column=5).value = self.format_change_for_excel(data.get('change_5m'))
            ws.cell(row=row, column=6).value = self.format_change_for_excel(data.get('change_15m'))
            ws.cell(row=row, column=7).value = self.format_change_for_excel(data.get('change_30m'))
            ws.cell(row=row, column=8).value = self.format_change_for_excel(data.get('change_1h'))
            ws.cell(row=row, column=9).value = self.format_change_for_excel(data.get('change_4h'))
            ws.cell(row=row, column=10).value = best_timeframe
            ws.cell(row=row, column=11).value = worst_timeframe
            ws.cell(row=row, column=12).value = f"{consistency:.1f}%"
            
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 12



    def create_price_analysis_sheet(self, wb, analyzed_prices=None, historical_data=None):
        """Create Price Analysis sheet with historical data"""
        ws = wb.create_sheet("Price Analysis")
        
        # Headers
        headers = [
            'Rank', 'Symbol', 'Current Price', '5m %', '15m %', '30m %', 
            '1h %', '4h %', 'Score', 'Trend', 'Volume', 'Last Updated'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Combine analyzed prices with historical data for ranking
        all_data = []
        
        # Add analyzed prices first
        if analyzed_prices:
            for item in analyzed_prices:
                all_data.append({
                    'symbol': item['symbol'],
                    'price': item.get('price'),
                    'changes': item.get('changes', {}),
                    'score': item.get('score', 0),
                    'source': 'analyzed'
                })
        
        # Add historical data (avoid duplicates)
        if historical_data:
            for symbol, data in historical_data.items():
                if not any(d['symbol'] == symbol for d in all_data):
                    all_data.append({
                        'symbol': symbol,
                        'price': data.get('current_price'),
                        'changes': {
                            '5m': data.get('change_5m'),
                            '15m': data.get('change_15m'),
                            '30m': data.get('change_30m'),
                            '60m': data.get('change_1h'),
                            '240m': data.get('change_4h')
                        },
                        'score': data.get('score', 0),
                        'source': 'historical'
                    })
        
        # Sort by score and take top 50
        all_data.sort(key=lambda x: x.get('score', 0), reverse=True)
        top_performers = all_data[:50]
        
        # Add data
        row = 2
        for i, item in enumerate(top_performers, 1):
            changes = item.get('changes', {})
            
            # Determine trend based on multiple timeframes
            trend_score = self.calculate_trend_score(changes)
            if trend_score > 5:
                trend = "🚀 STRONG UP"
            elif trend_score > 2:
                trend = "🟢 UP"
            elif trend_score < -5:
                trend = "🔻 STRONG DOWN"
            elif trend_score < -2:
                trend = "🔴 DOWN"
            else:
                trend = "⚪ FLAT"
            
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = item['symbol']
            ws.cell(row=row, column=3).value = item.get('price', 'N/A')
            ws.cell(row=row, column=4).value = self.format_change_for_excel(changes.get('5m'))
            ws.cell(row=row, column=5).value = self.format_change_for_excel(changes.get('15m'))
            ws.cell(row=row, column=6).value = self.format_change_for_excel(changes.get('30m'))
            ws.cell(row=row, column=7).value = self.format_change_for_excel(changes.get('60m'))
            ws.cell(row=row, column=8).value = self.format_change_for_excel(changes.get('240m'))
            ws.cell(row=row, column=9).value = f"{item.get('score', 0):.2f}"
            ws.cell(row=row, column=10).value = trend
            ws.cell(row=row, column=11).value = 'N/A'  # Volume would require additional data
            ws.cell(row=row, column=12).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 12



    def create_dashboard_sheet(self, wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data=None):
        """Create Dashboard sheet"""
        ws = wb.create_sheet("Dashboard")
        
        # Title
        ws['A1'] = 'MEXC FUTURES AUTO-UPDATE DASHBOARD'
        ws['A1'].font = Font(bold=True, size=14)
        
        # Get statistics
        unique_futures, exchange_stats = self.find_unique_futures_robust()
        working_exchanges = sum(1 for count in exchange_stats.values() if count > 0)
        total_exchanges = len(exchange_stats)
        
        # Calculate price coverage using analyzed_prices
        valid_prices = len([p for p in analyzed_prices if p.get('price') is not None]) if analyzed_prices else 0
        price_coverage = (valid_prices / len(unique_futures)) * 100 if unique_futures else 0
        
        # Statistics data
        stats_data = [
            ["Last Updated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Update Interval", f"{self.update_interval} minutes"],
            ["", ""],
            ["EXCHANGE STATISTICS", ""],
            ["Working Exchanges", f"{working_exchanges}/{total_exchanges}"],
            ["Total Unique Symbols", len(symbol_coverage)],
            ["Unique MEXC Futures", len(unique_futures)],
            ["", ""],
            ["PRICE ANALYSIS", ""],
            ["Symbols with Price Data", f"{valid_prices}/{len(unique_futures)}"],
            ["Price Coverage", f"{price_coverage:.1f}%"],
            ["MEXC Futures Count", len([f for f in all_futures_data if f['exchange'] == 'MEXC'])],
            ["", ""],
            ["PERFORMANCE", ""],
            ["Next Auto-Update", (datetime.now() + timedelta(minutes=self.update_interval)).strftime('%H:%M:%S')],
            ["Status", "RUNNING"],
        ]
        
        # Add data to sheet
        for i, (label, value) in enumerate(stats_data, 2):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            
            # Format headers
            if label and any(keyword in label for keyword in ["STATISTICS", "ANALYSIS", "PERFORMANCE"]):
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'A{i}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                ws[f'B{i}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25


    def create_unique_futures_sheet(self, wb, all_futures_data, symbol_coverage, analyzed_prices=None, historical_data=None):
        """Create Unique Futures sheet with historical data"""
        ws = wb.create_sheet("Unique Futures")
        
        # Headers matching Google Sheets
        headers = [
            'Symbol', 'Current Price', '5m Change %', '15m Change %', 
            '30m Change %', '1h Change %', '4h Change %', 'Score', 'Status', 'Last Updated'
        ]
        
        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Get unique futures
        unique_futures, _ = self.find_unique_futures_robust()
        
        # Add data with historical values
        row = 2
        for symbol in sorted(unique_futures):
            # Try to get historical data first, fall back to analyzed prices
            historical_info = historical_data.get(symbol) if historical_data else None
            price_info = next((p for p in analyzed_prices if p['symbol'] == symbol), None) if analyzed_prices else None
            
            if historical_info:
                # Use historical data from Google Sheets
                current_price = historical_info.get('current_price')
                change_5m = historical_info.get('change_5m')
                change_15m = historical_info.get('change_15m')
                change_30m = historical_info.get('change_30m')
                change_1h = historical_info.get('change_1h')
                change_4h = historical_info.get('change_4h')
                score = historical_info.get('score', 0)
                last_updated = historical_info.get('last_updated', '')
                status = historical_info.get('status', 'UNIQUE')
            elif price_info:
                # Fall back to analyzed prices
                current_price = price_info.get('price')
                changes = price_info.get('changes', {})
                change_5m = changes.get('5m')
                change_15m = changes.get('15m')
                change_30m = changes.get('30m')
                change_1h = changes.get('60m')
                change_4h = changes.get('240m')
                score = price_info.get('score', 0)
                last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                status = 'UNIQUE'
            else:
                # No data available
                current_price = None
                change_5m = change_15m = change_30m = change_1h = change_4h = None
                score = 0
                last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                status = 'UNIQUE'
            
            # Format price display
            if current_price is not None:
                if current_price >= 1:
                    price_display = f"${current_price:.4f}"
                elif current_price >= 0.01:
                    price_display = f"${current_price:.6f}"
                else:
                    price_display = f"${current_price:.8f}"
            else:
                price_display = 'N/A'
            
            # Add row data
            ws.cell(row=row, column=1).value = symbol
            ws.cell(row=row, column=2).value = price_display
            ws.cell(row=row, column=3).value = self.format_change_for_excel(change_5m)
            ws.cell(row=row, column=4).value = self.format_change_for_excel(change_15m)
            ws.cell(row=row, column=5).value = self.format_change_for_excel(change_30m)
            ws.cell(row=row, column=6).value = self.format_change_for_excel(change_1h)
            ws.cell(row=row, column=7).value = self.format_change_for_excel(change_4h)
            ws.cell(row=row, column=8).value = f"{score:.2f}"
            ws.cell(row=row, column=9).value = status
            ws.cell(row=row, column=10).value = last_updated
            
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 15

    def create_all_futures_sheet(self, wb, all_futures_data, symbol_coverage, historical_data=None):
        """Create All Futures sheet"""
        ws = wb.create_sheet("All Futures")
        
        # Headers
        headers = ['Symbol', 'Exchange', 'Normalized', 'Available On', 'Coverage', 'Timestamp', 'Unique']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add data
        row = 2
        for future in all_futures_data:
            normalized = self.normalize_symbol_for_comparison(future['symbol'])
            exchanges_list = symbol_coverage.get(normalized, set())
            available_on = ", ".join(sorted(exchanges_list)) if exchanges_list else "MEXC Only"
            coverage = f"{len(exchanges_list)} exchanges"
            is_unique = "✅" if len(exchanges_list) == 1 else ""
            
            ws.cell(row=row, column=1).value = future['symbol']
            ws.cell(row=row, column=2).value = future['exchange']
            ws.cell(row=row, column=3).value = normalized
            ws.cell(row=row, column=4).value = available_on
            ws.cell(row=row, column=5).value = coverage
            ws.cell(row=row, column=6).value = future['timestamp']
            ws.cell(row=row, column=7).value = is_unique
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 10

    def create_mexc_analysis_sheet(self, wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data=None):
        """Create MEXC Analysis sheet"""
        ws = wb.create_sheet("MEXC Analysis")
        
        # Headers
        headers = ['MEXC Symbol', 'Normalized', 'Available On', 'Exchanges Count', 'Current Price', '5m Change %', '1h Change %', '4h Change %', 'Status', 'Unique']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Get MEXC futures and price mapping
        mexc_futures = [f for f in all_futures_data if f['exchange'] == 'MEXC']
        price_map = {item['symbol']: item for item in analyzed_prices} if analyzed_prices else {}
        
        # Add data
        row = 2
        for future in mexc_futures:
            symbol = future['symbol']
            normalized = self.normalize_symbol_for_comparison(symbol)
            exchanges_list = symbol_coverage.get(normalized, set())
            available_on = ", ".join(sorted(exchanges_list)) if exchanges_list else "MEXC Only"
            exchange_count = len(exchanges_list)
            status = "Unique" if exchange_count == 1 else "Multi-exchange"
            unique_flag = "✅" if exchange_count == 1 else "🔸"
            
            price_info = price_map.get(symbol, {})
            changes = price_info.get('changes', {})
            
            ws.cell(row=row, column=1).value = symbol
            ws.cell(row=row, column=2).value = normalized
            ws.cell(row=row, column=3).value = available_on
            ws.cell(row=row, column=4).value = exchange_count
            ws.cell(row=row, column=5).value = price_info.get('price', 'N/A')
            ws.cell(row=row, column=6).value = self.format_change_for_excel(changes.get('5m'))
            ws.cell(row=row, column=7).value = self.format_change_for_excel(changes.get('60m'))
            ws.cell(row=row, column=8).value = self.format_change_for_excel(changes.get('240m'))
            ws.cell(row=row, column=9).value = status
            ws.cell(row=row, column=10).value = unique_flag
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 15

    def create_exchange_stats_sheet(self, wb, all_futures_data, historical_data=None):
        """Create Exchange Stats sheet"""
        ws = wb.create_sheet("Exchange Stats")
        
        # Headers
        headers = ['Exchange', 'Futures Count', 'Status', 'Last Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Count futures by exchange
        exchange_counts = {}
        for future in all_futures_data:
            exchange = future['exchange']
            exchange_counts[exchange] = exchange_counts.get(exchange, 0) + 1
        
        # Add data
        row = 2
        for exchange in sorted(exchange_counts.keys()):
            count = exchange_counts[exchange]
            status = "WORKING" if count > 0 else "FAILED"
            
            ws.cell(row=row, column=1).value = exchange
            ws.cell(row=row, column=2).value = count
            ws.cell(row=row, column=3).value = status
            ws.cell(row=row, column=4).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20


    def format_change_for_excel(self, change):
        """Format change for Excel with proper None handling"""
        if change is None:
            return 'N/A'
        
        # Add emoji based on change value for better visualization
        if change > 10:
            return f"🚀 {change:+.2f}%"
        elif change > 5:
            return f"🟢 {change:+.2f}%"
        elif change > 2:
            return f"📈 {change:+.2f}%"
        elif change < -10:
            return f"💥 {change:+.2f}%"
        elif change < -5:
            return f"🔴 {change:+.2f}%"
        elif change < -2:
            return f"📉 {change:+.2f}%"
        else:
            return f"{change:+.2f}%"


    def create_mexc_analysis_excel(self, all_futures_data, symbol_coverage, analyzed_prices=None):
        """Create comprehensive Excel file with historical data from Google Sheets"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Get historical data from Google Sheets
            historical_data = self.get_historical_data_from_sheets()
            
            # Create all sheets matching Google Sheets structure with historical data
            self.create_dashboard_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_unique_futures_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_all_futures_sheet(wb, all_futures_data, symbol_coverage, historical_data)
            self.create_mexc_analysis_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_price_analysis_sheet(wb, analyzed_prices, historical_data)
            self.create_exchange_stats_sheet(wb, all_futures_data, historical_data)
            self.create_historical_trends_sheet(wb, historical_data)  # New sheet for historical trends
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            output.close()
            
            logger.info("✅ Excel file created successfully with historical data")
            return excel_content
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return None




    def create_mexc_analysis_excel(self, all_futures_data, symbol_coverage, analyzed_prices=None):
        """Create comprehensive Excel file with historical data from Google Sheets"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Get historical data from Google Sheets
            historical_data = self.get_historical_data_from_sheets()
            
            # Create all sheets matching Google Sheets structure with historical data
            self.create_dashboard_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_unique_futures_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_all_futures_sheet(wb, all_futures_data, symbol_coverage, historical_data)
            self.create_mexc_analysis_sheet(wb, all_futures_data, symbol_coverage, analyzed_prices, historical_data)
            self.create_price_analysis_sheet(wb, analyzed_prices, historical_data)
            self.create_exchange_stats_sheet(wb, all_futures_data, historical_data)
            self.create_historical_trends_sheet(wb, historical_data)  # New sheet for historical trends
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_content = output.getvalue()
            output.close()
            
            logger.info("✅ Excel file created successfully with historical data")
            return excel_content
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return None



    def get_historical_data_from_sheets(self):
        """Extract historical data from Google Sheets"""
        try:
            historical_data = {}
            
            if not self.gs_client or not self.spreadsheet:
                logger.warning("Google Sheets not available for historical data")
                return historical_data
            
            # Read data from Unique Futures sheet
            try:
                worksheet = self.spreadsheet.worksheet('Unique Futures')
                sheet_data = worksheet.get_all_records()
                
                for row in sheet_data:
                    symbol = row.get('Symbol', '')
                    if symbol:
                        historical_data[symbol] = {
                            'current_price': self.parse_price_value(row.get('Current Price', '')),
                            'change_5m': self.parse_change_value(row.get('5m Change %', '')),
                            'change_15m': self.parse_change_value(row.get('15m Change %', '')),
                            'change_30m': self.parse_change_value(row.get('30m Change %', '')),
                            'change_1h': self.parse_change_value(row.get('1h Change %', '')),
                            'change_4h': self.parse_change_value(row.get('4h Change %', '')),
                            'score': self.parse_score_value(row.get('Score', '')),
                            'last_updated': row.get('Last Updated', ''),
                            'status': row.get('Status', '')
                        }
                
                logger.info(f"📊 Loaded historical data for {len(historical_data)} symbols from Google Sheets")
                
            except Exception as e:
                logger.error(f"Error reading historical data from Google Sheets: {e}")
            
            return historical_data
            
        except Exception as e:
            logger.error(f"Error getting historical data from sheets: {e}")
            return {}

    def get_historical_data_from_sheets(self):
        """Extract historical data from Google Sheets"""
        try:
            historical_data = {}
            
            if not self.gs_client or not self.spreadsheet:
                logger.warning("Google Sheets not available for historical data")
                return historical_data
            
            # Read data from Unique Futures sheet
            try:
                worksheet = self.spreadsheet.worksheet('Unique Futures')
                sheet_data = worksheet.get_all_records()
                
                for row in sheet_data:
                    symbol = row.get('Symbol', '')
                    if symbol:
                        historical_data[symbol] = {
                            'current_price': self.parse_price_value(row.get('Current Price', '')),
                            'change_5m': self.parse_change_value(row.get('5m Change %', '')),
                            'change_15m': self.parse_change_value(row.get('15m Change %', '')),
                            'change_30m': self.parse_change_value(row.get('30m Change %', '')),
                            'change_1h': self.parse_change_value(row.get('1h Change %', '')),
                            'change_4h': self.parse_change_value(row.get('4h Change %', '')),
                            'score': self.parse_score_value(row.get('Score', '')),
                            'last_updated': row.get('Last Updated', ''),
                            'status': row.get('Status', '')
                        }
                
                logger.info(f"📊 Loaded historical data for {len(historical_data)} symbols from Google Sheets")
                
            except Exception as e:
                logger.error(f"Error reading historical data from Google Sheets: {e}")
            
            return historical_data
            
        except Exception as e:
            logger.error(f"Error getting historical data from sheets: {e}")
            return {}

    def parse_price_value(self, price_str):
        """Parse price value from string format"""
        if not price_str or price_str == 'N/A':
            return None
        try:
            # Remove $ symbol and commas, then convert to float
            cleaned = str(price_str).replace('$', '').replace(',', '').strip()
            return float(cleaned)
        except:
            return None

    def parse_change_value(self, change_str):
        """Parse change percentage from string format"""
        if not change_str or change_str == 'N/A':
            return None
        try:
            # Remove emojis and % symbol, then convert to float
            cleaned = re.sub(r'[^\d.-]', '', str(change_str))
            return float(cleaned)
        except:
            return None

    def parse_score_value(self, score_str):
        """Parse score value from string format"""
        if not score_str or score_str == 'N/A':
            return 0
        try:
            return float(score_str)
        except:
            return 0
        

    # ==================== CORE UNIQUE FUTURES LOGIC ====================


    def normalize_symbol_for_comparison(self, symbol):
        """Robust symbol normalization - FIXED STOCK SYMBOLS"""
        if not symbol:
            return ""
        
        # Convert to uppercase
        symbol = symbol.upper()
        
        # DON'T remove stock-related suffixes - this is the main bug!
        # Keep STOCK, SHARE, etc. as they are important for stock symbols
        
        # Only remove common futures/perp suffixes
        patterns = [
            r'[-_/]?PERP(ETUAL)?$',
            r'[-_/]?SWAP$',
            r'[-_/]?FUTURES?$',
            r'[-_/]?CONTRACT$',
        ]
        
        normalized = symbol
        for pattern in patterns:
            normalized = re.sub(pattern, '', normalized, flags=re.IGNORECASE)
        
        # Remove separators but keep the symbol structure
        normalized = re.sub(r'[-_/]', '', normalized)
        
        # DON'T remove trailing numbers - stock symbols often have numbers
        # normalized = re.sub(r'\d+$', '', normalized)  # REMOVE THIS LINE
        
        return normalized.strip()

    def find_unique_futures_robust(self, timeout=60):
        """Find unique futures without threading to avoid thread errors"""
        try:
            logger.info("🔍 Starting unique futures search...")
            
            # Get MEXC futures
            mexc_futures = self.get_mexc_futures()
            if not mexc_futures:
                logger.error("❌ No MEXC futures found")
                return set(), {}
            
            logger.info(f"📊 MEXC futures to check: {len(mexc_futures)}")
            
            # Get futures from other exchanges
            all_other_futures, exchange_stats = self.get_all_exchanges_futures()
            logger.info(f"📊 Other exchanges futures: {len(all_other_futures)}")
            
            # Find unique futures (futures that are ONLY on MEXC)
            unique_futures = set()
            
            # Normalize all other futures for comparison
            logger.info("🔄 Normalizing symbols for comparison...")
            normalized_other_futures = set()
            for symbol in all_other_futures:
                try:
                    normalized = self.normalize_symbol_for_comparison(symbol)
                    if normalized:
                        normalized_other_futures.add(normalized)
                except Exception as e:
                    logger.debug(f"Could not normalize {symbol}: {e}")
            
            logger.info(f"📊 Normalized other futures: {len(normalized_other_futures)}")
            
            # Check each MEXC future against normalized other futures
            checked_count = 0
            for mexc_symbol in mexc_futures:
                try:
                    if checked_count % 100 == 0:
                        logger.info(f"🔍 Checked {checked_count}/{len(mexc_futures)} symbols...")
                    
                    normalized_mexc = self.normalize_symbol_for_comparison(mexc_symbol)
                    if normalized_mexc and normalized_mexc not in normalized_other_futures:
                        unique_futures.add(mexc_symbol)
                    
                    checked_count += 1
                    
                except Exception as e:
                    logger.error(f"Error checking {mexc_symbol}: {e}")
                    continue
            
            logger.info(f"🎯 Found {len(unique_futures)} unique futures")
            return unique_futures, exchange_stats
            
        except Exception as e:
            logger.error(f"❌ Unique futures search error: {e}")
            return set(), {}
        
    def format_change_with_emoji(self, change):
        """Format change with emoji and sign for Google Sheets"""
        if change is None:
            return 'N/A'
        
        # Use the same emoji logic as Telegram messages
        if change > 5:
            return f"🚀 +{change:.2f}%"
        elif change > 2:
            return f"🟢 +{change:.2f}%"
        elif change > 0.1:
            return f"📈 +{change:.2f}%"
        elif change < -5:
            return f"💥 {change:.2f}%"
        elif change < -2:
            return f"🔴 {change:.2f}%"
        elif change < -0.1:
            return f"📉 {change:.2f}%"
        else:
            return f"⚪ {change:.2f}%"

    def update_unique_futures_sheet_with_prices(self, unique_futures, analyzed_prices):
        """Update Unique Futures sheet with emoji formatting"""
        try:
            worksheet = self.spreadsheet.worksheet('Unique Futures')
            
            # Clear existing data
            worksheet.clear()
            
            headers = [
                'Symbol', 'Current Price', '5m Change %', '15m Change %', 
                '30m Change %', '1h Change %', '4h Change %', 'Score', 'Status', 'Last Updated'
            ]
            worksheet.update([headers], 'A1')
            
            sheet_data = []
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            price_map = {item['symbol']: item for item in analyzed_prices}
            
            for symbol in sorted(unique_futures):
                price_info = price_map.get(symbol)
                changes = price_info.get('changes', {}) if price_info else {}
                price = price_info.get('price') if price_info else None
                
                # Format price exactly like Telegram
                if price is not None:
                    if price >= 1000:
                        price_display = f"${price:,.2f}"
                    elif price >= 1:
                        price_display = f"${price:.2f}"
                    elif price >= 0.01:
                        price_display = f"${price:.4f}"
                    elif price >= 0.0001:
                        price_display = f"${price:.6f}"
                    else:
                        price_display = f"${price:.2e}"  # Scientific notation for very small numbers
                else:
                    price_display = 'N/A'
                
                # Format changes with emojis like Telegram
                row = [
                    symbol,
                    price_display,
                    self.format_change_with_emoji(changes.get('5m')),
                    self.format_change_with_emoji(changes.get('15m')),
                    self.format_change_with_emoji(changes.get('30m')),
                    self.format_change_with_emoji(changes.get('60m')),
                    self.format_change_with_emoji(changes.get('240m')),
                    f"{price_info.get('score', 0):.2f}" if price_info else 'N/A',
                    'UNIQUE',
                    current_time
                ]
                sheet_data.append(row)
            
            if sheet_data:
                worksheet.update(f'A2:J{len(sheet_data) + 1}', sheet_data)
                logger.info(f"✅ Updated Unique Futures with {len(sheet_data)} records (emoji format)")
                
        except Exception as e:
            logger.error(f"Error updating Unique Futures sheet: {e}")

    def update_price_analysis_sheet(self, analyzed_prices):
        """Update Price Analysis sheet with emoji formatting"""
        try:
            worksheet = self.spreadsheet.worksheet('Price Analysis')
            worksheet.clear()
            
            headers = [
                'Rank', 'Symbol', 'Current Price', '5m %', '15m %', '30m %', 
                '1h %', '4h %', 'Score', 'Trend', 'Last Updated'
            ]
            worksheet.update([headers], 'A1')
            
            sheet_data = []
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Take top 50 performers
            top_performers = analyzed_prices[:50] if analyzed_prices else []
            
            for i, item in enumerate(top_performers, 1):
                changes = item.get('changes', {})
                price = item.get('price')
                
                # Format price like Telegram
                if price is not None:
                    if price >= 1000:
                        price_display = f"${price:,.2f}"
                    elif price >= 1:
                        price_display = f"${price:.2f}"
                    elif price >= 0.01:
                        price_display = f"${price:.4f}"
                    elif price >= 0.0001:
                        price_display = f"${price:.6f}"
                    else:
                        price_display = f"${price:.2e}"
                else:
                    price_display = 'N/A'
                
                # Determine trend with emojis
                score = item.get('score', 0)
                if score > 5:
                    trend = "🚀 STRONG UP"
                elif score > 2:
                    trend = "🟢 UP"
                elif score < -5:
                    trend = "🔻 STRONG DOWN"
                elif score < -2:
                    trend = "🔴 DOWN"
                else:
                    trend = "⚪ FLAT"
                
                row = [
                    i,
                    item['symbol'],
                    price_display,
                    self.format_change_with_emoji(changes.get('5m')),
                    self.format_change_with_emoji(changes.get('15m')),
                    self.format_change_with_emoji(changes.get('30m')),
                    self.format_change_with_emoji(changes.get('60m')),
                    self.format_change_with_emoji(changes.get('240m')),
                    f"{score:.2f}",
                    trend,
                    current_time
                ]
                sheet_data.append(row)
            
            if sheet_data:
                worksheet.update(f'A2:K{len(sheet_data) + 1}', sheet_data)
                logger.info(f"✅ Updated Price Analysis with {len(sheet_data)} records (emoji format)")
                
        except Exception as e:
            logger.error(f"Error updating Price Analysis sheet: {e}")

    def update_mexc_analysis_sheet_with_prices(self, all_futures_data, symbol_coverage, analyzed_prices, timestamp):
        """Update MEXC Analysis sheet with emoji formatting"""
        try:
            worksheet = self.spreadsheet.worksheet('MEXC Analysis')
            worksheet.clear()
            
            headers = [
                'MEXC Symbol', 'Normalized', 'Available On', 'Exchanges Count', 
                'Current Price', '5m Change %', '1h Change %', '4h Change %', 
                'Status', 'Unique', 'Timestamp'
            ]
            worksheet.update([headers], 'A1')
            
            # Get only MEXC futures
            mexc_futures = [f for f in all_futures_data if f['exchange'] == 'MEXC']
            
            sheet_data = []
            price_map = {item['symbol']: item for item in analyzed_prices}
            
            for future in mexc_futures:
                symbol = future['symbol']
                normalized = self.normalize_symbol_for_comparison(symbol)
                exchanges_list = symbol_coverage.get(normalized, set())
                available_on = ", ".join(sorted(exchanges_list)) if exchanges_list else "MEXC Only"
                exchange_count = len(exchanges_list)
                status = "Unique" if exchange_count == 1 else "Multi-exchange"
                unique_flag = "✅" if exchange_count == 1 else "🔸"
                
                # Get price info
                price_info = price_map.get(symbol, {})
                changes = price_info.get('changes', {})
                price = price_info.get('price')
                
                # Format price like Telegram
                if price is not None:
                    if price >= 1000:
                        price_display = f"${price:,.2f}"
                    elif price >= 1:
                        price_display = f"${price:.2f}"
                    elif price >= 0.01:
                        price_display = f"${price:.4f}"
                    elif price >= 0.0001:
                        price_display = f"${price:.6f}"
                    else:
                        price_display = f"${price:.2e}"
                else:
                    price_display = 'N/A'
                
                row = [
                    symbol,
                    normalized,
                    available_on,
                    exchange_count,
                    price_display,
                    self.format_change_with_emoji(changes.get('5m')),
                    self.format_change_with_emoji(changes.get('60m')),
                    self.format_change_with_emoji(changes.get('240m')),
                    status,
                    unique_flag,
                    timestamp
                ]
                sheet_data.append(row)
            
            if sheet_data:
                worksheet.update(f'A2:K{len(sheet_data) + 1}', sheet_data)
                logger.info(f"✅ Updated MEXC Analysis with {len(sheet_data)} records (emoji format)")
                
        except Exception as e:
            logger.error(f"Error updating MEXC Analysis sheet: {e}")


    def format_price_for_display(self, price):
        """Format price for display exactly like Google Sheets"""
        if price is None:
            return "N/A"
        
        if price >= 1000:
            return f"${price:,.2f}"
        elif price >= 1:
            return f"${price:.2f}"
        elif price >= 0.01:
            return f"${price:.4f}"
        elif price >= 0.0001:
            return f"${price:.6f}"
        else:
            return f"${price:.2e}"  # Scientific notation for very small numbers

    def format_change_for_telegram(self, change):
        """Format change for Telegram messages"""
        if change is None:
            return "N/A"
        
        # Same emoji logic as Google Sheets
        if change > 5:
            return f"🚀 +{change:.2f}%"
        elif change > 2:
            return f"🟢 +{change:.2f}%"
        elif change > 0.1:
            return f"📈 +{change:.2f}%"
        elif change < -5:
            return f"💥 {change:.2f}%"
        elif change < -2:
            return f"🔴 {change:.2f}%"
        elif change < -0.1:
            return f"📉 {change:.2f}%"
        else:
            return f"⚪ {change:.2f}%"
        

    def send_new_unique_notification(self, new_futures, all_unique):
        """Send notification about new unique futures - UPDATED FORMATTING"""
        try:
            display_futures = list(new_futures)[:10]
            
            message = "🚀 <b>NEW UNIQUE FUTURES FOUND!</b>\n\n"
            
            # Get ALL prices
            all_price_data = self.get_all_mexc_prices()
            
            valid_count = 0
            for symbol in display_futures:
                price_info = all_price_data.get(symbol)
                
                if price_info and price_info.get('price') is not None and price_info['price'] >= 0:
                    # VALID PRICE (including micro-cap)
                    changes = price_info.get('changes', {})
                    change_5m = changes.get('5m', 0)
                    price = price_info['price']
                    
                    message += f"✅ <b>{symbol}</b>\n"
                    message += f"   Price: {self.format_price_for_display(price)}\n"
                    message += f"   5m: {self.format_change_for_telegram(change_5m)}\n\n"
                    valid_count += 1
                    
                else:
                    # TRULY MISSING PRICE
                    message += f"✅ <b>{symbol}</b> (price data unavailable)\n\n"
            
            if len(new_futures) > len(display_futures):
                message += f"... and {len(new_futures) - len(display_futures)} more symbols\n\n"
            
            message += f"📊 Total unique: <b>{len(all_unique)}</b>"
            message += f"\n💰 With prices: <b>{valid_count}/{len(display_futures)}</b> shown symbols"
            
            self.send_broadcast_message(message)
            
        except Exception as e:
            logger.error(f"Error sending new unique notification: {e}")





    def send_lost_unique_notification(self, lost_futures, remaining_unique):
        """Send notification about lost unique futures - OPTIMIZED"""
        try:
            # Limit the number of symbols to process
            display_futures = list(lost_futures)[:10]  # Show max 10 symbols
            
            message = "📉 <b>FUTURES NO LONGER UNIQUE:</b>\n\n"
            
            for symbol in display_futures:
                # Use fast verification with cached data
                try:
                    # For lost futures, we know they were previously unique
                    # Just show they're no longer unique without detailed coverage check
                    message += f"❌ <b>{symbol}</b>\n"
                    message += f"   No longer exclusive to MEXC\n\n"
                except Exception as e:
                    logger.error(f"Error checking coverage for {symbol}: {e}")
                    message += f"❌ <b>{symbol}</b> (verification failed)\n\n"
            
            if len(lost_futures) > len(display_futures):
                message += f"... and {len(lost_futures) - len(display_futures)} more symbols\n\n"
            
            message += f"📊 Remaining unique: <b>{len(remaining_unique)}</b>"
            
            self.send_broadcast_message(message)
            
        except Exception as e:
            logger.error(f"Error sending lost unique notification: {e}")

    def get_all_exchanges_futures(self):
        """Get futures from all exchanges except MEXC"""
        exchanges = {
            'Binance': self.get_binance_futures,
            'Bybit': self.get_bybit_futures,
            'OKX': self.get_okx_futures,
            'Gate.io': self.get_gate_futures,
            'KuCoin': self.get_kucoin_futures,
            'BingX': self.get_bingx_futures,
            'BitGet': self.get_bitget_futures
        }
        
        all_futures = set()
        exchange_stats = {}
        
        for name, method in exchanges.items():
            try:
                logger.info(f"🔍 Getting futures from {name}...")
                futures = method()
                if futures:
                    all_futures.update(futures)
                    exchange_stats[name] = len(futures)
                    logger.info(f"✅ {name}: {len(futures)} futures")
                else:
                    exchange_stats[name] = 0
                    logger.warning(f"❌ {name}: No futures found")
            except Exception as e:
                exchange_stats[name] = 0
                logger.error(f"🚨 Error getting {name} futures: {e}")
        
        logger.info(f"📊 Total futures from other exchanges: {len(all_futures)}")
        return all_futures, exchange_stats

    def verify_symbol_coverage(self, symbol, all_futures_cache=None, mexc_futures_cache=None):
        """FAST symbol coverage check using cached data - FIXED"""
        coverage = []
        
        # Use cached data if provided
        if mexc_futures_cache is not None and symbol in mexc_futures_cache:
            coverage.append('MEXC')
        
        # Check other exchanges using cached data
        # FIX: Use LESS aggressive normalization for stock symbols
        normalized_target = self.normalize_symbol_for_comparison(symbol)
        
        # For stock symbols, be more careful with normalization
        if 'STOCK' in symbol.upper():
            # For stock symbols, try multiple normalization approaches
            normalized_variations = [
                normalized_target,
                symbol.upper().replace('_', '').replace('-', '').replace('/', ''),
                symbol.upper().replace('STOCK', '').replace('_', '').replace('-', '').replace('/', ''),
            ]
        else:
            normalized_variations = [normalized_target]
        
        exchange_checks = {
            'Binance': all_futures_cache,
            'Bybit': all_futures_cache, 
            'OKX': all_futures_cache,
            'Gate.io': all_futures_cache,
            'KuCoin': all_futures_cache,
            'BingX': all_futures_cache,
            'BitGet': all_futures_cache
        }
        
        for exchange_name, futures_cache in exchange_checks.items():
            if futures_cache is None:
                continue
                
            found = False
            for fut in futures_cache:
                # Try all normalization variations
                for normalized_variation in normalized_variations:
                    normalized_fut = self.normalize_symbol_for_comparison(fut)
                    if normalized_variation == normalized_fut:
                        found = True
                        break
                if found:
                    break
            
            if found:
                coverage.append(exchange_name)
        
        return coverage

    # ==================== EXCHANGE API METHODS ====================

    def get_mexc_futures(self):
        """Get ALL futures from MEXC"""
        try:
            url = "https://contract.mexc.com/api/v1/contract/detail"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            futures = set()
            for contract in data.get('data', []):
                symbol = contract.get('symbol', '')
                if symbol:
                    futures.add(symbol)
            
            logger.info(f"MEXC: {len(futures)} futures")
            return futures
        except Exception as e:
            logger.error(f"MEXC error: {e}")
            return set()

    def get_binance_futures(self):
        """Get Binance futures with proxy support"""
        try:
            logger.info("🔄 Fetching Binance futures...")
            
            futures = set()
            
            # USDⓈ-M Futures - try multiple endpoints
            endpoints = [
                "https://testnet.binancefuture.com/fapi/v1/exchangeInfo"  # Fallback testnet
            ]
            
            for url in endpoints:
                logger.info(f"📡 Trying Binance URL: {url}")
                response = self._make_request_with_retry(url)
                
                if response and response.status_code == 200:
                    data = response.json()
                    symbols = data.get('symbols', [])
                    
                    usdt_futures = set()
                    for symbol in symbols:
                        if (symbol.get('contractType') == 'PERPETUAL' and 
                            symbol.get('status') == 'TRADING'):
                            usdt_futures.add(symbol.get('symbol'))
                    
                    futures.update(usdt_futures)
                    logger.info(f"✅ Binance USDⓈ-M perpetuals found: {len(usdt_futures)}")
                    break  # Success, no need to try other endpoints
                else:
                    logger.warning(f"❌ Failed to fetch from {url}")
            
            # If still no data, try alternative approach
            if not futures:
                logger.info("🔄 Trying alternative Binance endpoint...")
                alt_response = self._make_request_with_retry("https://api.binance.com/api/v3/exchangeInfo")
                if alt_response and alt_response.status_code == 200:
                    # This gives spot symbols, but we can use it as fallback
                    data = alt_response.json()
                    symbols = data.get('symbols', [])
                    spot_symbols = {s['symbol'] for s in symbols if s.get('status') == 'TRADING'}
                    # Filter for common futures symbols pattern
                    futures = {s for s in spot_symbols if s.endswith('USDT')}
                    logger.info(f"🔄 Using spot symbols as fallback: {len(futures)}")
            
            logger.info(f"🎯 Binance TOTAL: {len(futures)} futures")
            return futures
            
        except Exception as e:
            logger.error(f"❌ Binance error: {e}")
            return set()

    def get_bybit_futures(self):
        """Extremely simple Bybit implementation with caching to avoid 403 loops"""
        try:
            # Check cache first to avoid repeated failed requests
            cache_key = "bybit_futures_cache"
            cache_timeout = 300  # 5 minutes
            
            if hasattr(self, '_bybit_cache') and hasattr(self, '_bybit_cache_time'):
                if (datetime.now() - self._bybit_cache_time).seconds < cache_timeout:
                    logger.info("🔄 Using cached Bybit data")
                    return self._bybit_cache
            
            logger.info("🔄 Trying simplified Bybit request...")
            
            # Try the most basic endpoint with minimal headers
            url = "https://api.bybit.com/v5/market/tickers?category=linear"
            
            # Use minimal headers to avoid detection
            headers = {
                'User-Agent': 'Mozilla/5.0',
                'Accept': '*/*',
            }
            
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    if data.get('retCode') == 0:
                        futures = set()
                        for item in data.get('result', {}).get('list', []):
                            symbol = item.get('symbol')
                            if symbol:
                                futures.add(symbol)
                        
                        # Cache successful result
                        self._bybit_cache = futures
                        self._bybit_cache_time = datetime.now()
                        
                        logger.info(f"✅ Bybit simple: {len(futures)} symbols")
                        return futures
                except:
                    pass
            
            # If we get here, the request failed
            logger.warning("⚠️ Bybit simple method failed, using empty set")
            
            # Cache empty result to avoid repeated attempts
            self._bybit_cache = set()
            self._bybit_cache_time = datetime.now()
            
            return set()
            
        except Exception as e:
            logger.error(f"Bybit simple error: {e}")
            # Cache empty result on error too
            self._bybit_cache = set()
            self._bybit_cache_time = datetime.now()
            return set()
        
    def get_okx_futures(self):
        """Get ALL futures from OKX"""
        try:
            url = "https://www.okx.com/api/v5/public/instruments?instType=SWAP"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            futures = set()
            for item in data.get('data', []):
                inst_id = item.get('instId', '')
                if inst_id and 'SWAP' in inst_id:
                    futures.add(inst_id)
            
            logger.info(f"OKX: {len(futures)} futures")
            return futures
        except Exception as e:
            logger.error(f"OKX error: {e}")
            return set()

    def get_gate_futures(self):
        """Get ALL futures from Gate.io"""
        try:
            url = "https://api.gateio.ws/api/v4/futures/usdt/contracts"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            futures = set()
            for item in data:
                symbol = item.get('name', '')
                if symbol and item.get('in_delisting', False) is False:
                    futures.add(symbol)
            
            logger.info(f"Gate.io: {len(futures)} futures")
            return futures
        except Exception as e:
            logger.error(f"Gate.io error: {e}")
            return set()

    def get_kucoin_futures(self):
        """Get ALL futures from KuCoin"""
        try:
            url = "https://api-futures.kucoin.com/api/v1/contracts/active"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            futures = set()
            for item in data.get('data', []):
                symbol = item.get('symbol', '')
                if symbol:
                    futures.add(symbol)
            
            logger.info(f"KuCoin: {len(futures)} futures")
            return futures
        except Exception as e:
            logger.error(f"KuCoin error: {e}")
            return set()

    def get_bingx_futures(self):
        """Get ALL futures from BingX"""
        try:
            url = "https://open-api.bingx.com/openApi/swap/v2/quote/contracts"
            response = requests.get(url, timeout=10)
            data = response.json()
            
            futures = set()
            for item in data.get('data', []):
                symbol = item.get('symbol', '')
                if symbol:
                    futures.add(symbol)
            
            logger.info(f"BingX: {len(futures)} futures")
            return futures
        except Exception as e:
            logger.error(f"BingX error: {e}")
            return set()

    def get_bitget_futures(self):
        """Get Bitget perpetual futures"""
        try:
            futures = set()
            
            # USDT-FUTURES
            url1 = "https://api.bitget.com/api/v2/mix/market/contracts?productType=usdt-futures"
            response1 = requests.get(url1, timeout=10)
            
            if response1.status_code == 200:
                data = response1.json()
                if data.get('code') == '00000':
                    for item in data.get('data', []):
                        symbol_type = item.get('symbolType')
                        symbol_name = item.get('symbol')
                        if symbol_type == 'perpetual':
                            futures.add(symbol_name)
            
            # COIN-FUTURES
            url2 = "https://api.bitget.com/api/v2/mix/market/contracts?productType=coin-futures"
            response2 = requests.get(url2, timeout=10)
            
            if response2.status_code == 200:
                data = response2.json()
                if data.get('code') == '00000':
                    for item in data.get('data', []):
                        symbol_type = item.get('symbolType')
                        symbol_name = item.get('symbol')
                        if symbol_type == 'perpetual':
                            futures.add(symbol_name)
            
            logger.info(f"BitGet: {len(futures)} futures")
            return futures
            
        except Exception as e:
            logger.error(f"BitGet error: {e}")
            return set()

    # ==================== TELEGRAM COMMANDS ====================

    def excel_command(self, update: Update, context: CallbackContext):
        """Download comprehensive Excel report"""
        self.create_and_send_excel(update, context)

    def setup_handlers(self):
        """Setup command handlers"""
        self.dispatcher.add_handler(CommandHandler("start", self.start_command))
        self.dispatcher.add_handler(CommandHandler("status", self.status_command))
        self.dispatcher.add_handler(CommandHandler("check", self.check_command))
        self.dispatcher.add_handler(CommandHandler("help", self.help_command))
        self.dispatcher.add_handler(CommandHandler("stats", self.stats_command))
        self.dispatcher.add_handler(CommandHandler("exchanges", self.exchanges_command))
        self.dispatcher.add_handler(CommandHandler("analysis", self.analysis_command))
        self.dispatcher.add_handler(CommandHandler("findunique", self.find_unique_command))
        self.dispatcher.add_handler(CommandHandler("checksymbol", self.check_symbol_command))
        self.dispatcher.add_handler(CommandHandler("prices", self.prices_command))
        self.dispatcher.add_handler(CommandHandler("toppers", self.top_performers_command))
        self.dispatcher.add_handler(CommandHandler("forceupdate", self.force_update_command))
        self.dispatcher.add_handler(CommandHandler("excel", self.excel_command))
        self.dispatcher.add_handler(CommandHandler("download", self.excel_command))
        self.dispatcher.add_handler(CommandHandler("pricedebug", self.price_debug_command))
        self.dispatcher.add_handler(CommandHandler("symboldebug", self.symbol_debug_command))
        self.dispatcher.add_handler(CommandHandler("dataflow", self.data_flow_debug_command))
        self.dispatcher.add_handler(CommandHandler("qkctest", self.qkc_test_command))
        self.dispatcher.add_handler(CommandHandler("validateprices", self.validate_prices_command))
        self.dispatcher.add_handler(CommandHandler("symbolsearch", self.symbol_search_command))
        self.dispatcher.add_handler(CommandHandler("debugdatasources", self.debug_data_sources))
        self.dispatcher.add_handler(CommandHandler("growth", self.send_quick_growth_chart))
        self.dispatcher.add_handler(CommandHandler("growthreport", self.send_detailed_growth_report))
        self.dispatcher.add_handler(CommandHandler("4hchart", self.send_quick_growth_chart))
        self.dispatcher.add_handler(CommandHandler("trends", self.send_trend_analysis_command))
        self.dispatcher.add_handler(CommandHandler("datastatus", self.data_status_command))

    def symbol_search_command(self, update: Update, context: CallbackContext):
        """Search for symbols in MEXC API - CORRECTED"""
        try:
            if not context.args:
                update.message.reply_html("Usage: /symbolsearch SYMBOL\nExample: /symbolsearch BOBBSC")
                return
            
            search_term = context.args[0].upper()
            update.message.reply_html(f"🔍 <b>Searching for:</b> {search_term}")
            
            # Get batch data to see what's available
            batch_data = self.get_mexc_prices_batch_working()
            
            # Search for matching symbols
            matching_symbols = [s for s in batch_data.keys() if search_term in s]
            
            # Get MEXC futures to see what should be there
            mexc_futures = self.get_mexc_futures()
            mexc_matches = [s for s in mexc_futures if search_term in s]
            
            message = (
                f"🔍 <b>Symbol Search: {search_term}</b>\n\n"
                f"📊 <b>MEXC Futures List:</b> {len(mexc_matches)} matches\n"
            )
            
            if mexc_matches:
                message += "\n".join([f"• {s}" for s in mexc_matches[:10]]) + "\n"
            else:
                message += "• No matches found\n"
            
            message += f"\n📊 <b>Batch API Data:</b> {len(matching_symbols)} matches\n"
            
            if matching_symbols:
                for symbol in matching_symbols[:10]:
                    price = batch_data[symbol].get('price')
                    message += f"• {symbol}: ${price}\n"
            else:
                message += "• No matches in batch API\n"
            
            # FIXED: Create proper test symbols based on input
            message += f"\n🔧 <b>Direct API Tests:</b>\n"
            
            # If search_term already has _USDT, test it directly
            if search_term.endswith('_USDT'):
                test_symbols = [search_term]  # Just test the exact symbol
            else:
                # If it's a base symbol, add _USDT suffix
                test_symbols = [f"{search_term}_USDT"]
            
            for test_symbol in test_symbols:
                try:
                    url = f"https://contract.mexc.com/api/v1/contract/ticker?symbol={test_symbol}"
                    response = requests.get(url, timeout=5)
                    
                    if response.status_code == 200:
                        data = response.json()
                        if data.get('success') and data.get('data'):
                            # Handle both list and dict response formats
                            ticker_data = data['data']
                            if isinstance(ticker_data, list) and ticker_data:
                                price = ticker_data[0].get('lastPrice')
                            else:
                                price = ticker_data.get('lastPrice')
                                
                            if price:
                                message += f"• {test_symbol}: ✅ FOUND (${price})\n"
                            else:
                                message += f"• {test_symbol}: ✅ FOUND but no price data\n"
                        else:
                            message += f"• {test_symbol}: ❌ API returned success=False\n"
                    else:
                        message += f"• {test_symbol}: ❌ HTTP {response.status_code}\n"
                except Exception as e:
                    message += f"• {test_symbol}: ❌ ERROR: {str(e)[:50]}...\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Search error: {str(e)}")
            

    def debug_data_sources(self, update: Update, context: CallbackContext):
        """Debug where data is coming from"""
        try:
            update.message.reply_html("🔍 <b>Debugging data sources...</b>")
            
            # Source 1: MEXC futures list
            mexc_futures = self.get_mexc_futures()
            
            # Source 2: Batch price data
            batch_data = self.get_mexc_prices_batch_working()
            
            # Source 3: Unique futures
            unique_futures, _ = self.find_unique_futures_robust()
            
            # Test symbols
            test_symbols = ['BOBBSC_USDT', 'MANYU_USDT', 'RVV_USDT', 'AAPLSTOCK_USDT', 'LAZIO_USDT']
            
            message = "🔍 <b>Data Source Analysis</b>\n\n"
            
            for symbol in test_symbols:
                message += f"<b>{symbol}</b>\n"
                message += f"• MEXC Futures: {'✅' if symbol in mexc_futures else '❌'}\n"
                message += f"• Batch Prices: {'✅' if symbol in batch_data else '❌'}\n"
                message += f"• Unique Futures: {'✅' if symbol in unique_futures else '❌'}\n"
                message += "\n"
            
            # Check if there are symbol format differences
            message += "<b>🔍 Symbol Format Analysis</b>\n"
            batch_symbols_sample = list(batch_data.keys())[:5]
            message += f"Batch API symbols sample: {batch_symbols_sample}\n\n"
            
            mexc_futures_sample = list(mexc_futures)[:5]
            message += f"MEXC Futures sample: {mexc_futures_sample}\n\n"
            
            unique_sample = list(unique_futures)[:5]
            message += f"Unique Futures sample: {unique_sample}"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Debug error: {str(e)}")

    def validate_prices_command(self, update: Update, context: CallbackContext):
        """Validate prices for symbols with issues"""
        try:
            update.message.reply_html("🔍 <b>Validating price data quality...</b>")
            
            # Test problematic symbols
            test_symbols = [
                'BOBBSC_USDT',  # Shows $0.000000
                'MANYU_USDT',   # Shows $0.000000  
                'RVV_USDT',     # Shows unavailable
                'AAPLSTOCK_USDT', # Shows valid price
                'LAZIO_USDT'    # Shows valid price
            ]
            
            results = []
            all_price_data = self.get_all_mexc_prices()
            
            for symbol in test_symbols:
                price_info = all_price_data.get(symbol)
                
                if not price_info:
                    results.append(f"❌ {symbol}: NOT IN PRICE DATA")
                    continue
                    
                price = price_info.get('price')
                source = price_info.get('source', 'unknown')
                
                if price is None:
                    results.append(f"❌ {symbol}: PRICE IS NONE")
                elif price <= 0:
                    results.append(f"⚠️ {symbol}: ZERO/INVALID PRICE (${price}) - {source}")
                else:
                    results.append(f"✅ {symbol}: VALID ${price:.6f} - {source}")
            
            message = "🔍 <b>Price Validation Results</b>\n\n" + "\n".join(results)
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Validation error: {str(e)}")
            

    def qkc_test_command(self, update: Update, context: CallbackContext):
        """Test QKC_USDT specifically"""
        try:
            update.message.reply_html("🔍 <b>Testing QKC_USDT data flow...</b>")
            
            # Test each step
            batch_data = self.get_mexc_prices_batch_working()
            unique_futures, _ = self.find_unique_futures_robust()
            all_prices = self.get_all_mexc_prices()
            
            message = (
                f"🔍 <b>QKC_USDT Test Results</b>\n\n"
                f"📊 <b>Batch API:</b> {'✅ PRESENT' if 'QKC_USDT' in batch_data else '❌ MISSING'}\n"
                f"🎯 <b>Unique Futures:</b> {'✅ PRESENT' if 'QKC_USDT' in unique_futures else '❌ MISSING'}\n"
                f"💰 <b>Final Prices:</b> {'✅ PRESENT' if 'QKC_USDT' in all_prices else '❌ MISSING'}\n\n"
            )
            
            if 'QKC_USDT' in batch_data:
                message += f"• Batch Price: ${batch_data['QKC_USDT'].get('price')}\n"
            if 'QKC_USDT' in all_prices:
                message += f"• Final Price: ${all_prices['QKC_USDT'].get('price')}\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Test error: {str(e)}")


    def data_flow_debug_command(self, update: Update, context: CallbackContext):
        """Debug the complete data flow for a symbol"""
        try:
            if not context.args:
                update.message.reply_html("Usage: /dataflow SYMBOL\nExample: /dataflow QKC_USDT")
                return
            
            symbol = context.args[0].upper()
            update.message.reply_html(f"🔍 <b>Data Flow Debug:</b> {symbol}")
            
            # Step 1: Check batch API
            batch_data = self.get_mexc_prices_batch_working()
            in_batch = symbol in batch_data
            batch_price = batch_data.get(symbol, {}).get('price') if in_batch else None
            
            # Step 2: Check individual API
            individual_data = self.get_mexc_price_data_working(symbol)
            individual_works = individual_data is not None
            individual_price = individual_data.get('price') if individual_data else None
            
            # Step 3: Check get_all_mexc_prices
            all_prices = self.get_all_mexc_prices()
            in_all_prices = symbol in all_prices
            all_price_value = all_prices.get(symbol, {}).get('price') if in_all_prices else None
            
            # Step 4: Check analyzed_prices
            analyzed_prices = self.analyze_price_movements(all_prices)
            in_analyzed = any(p['symbol'] == symbol for p in analyzed_prices)
            analyzed_info = next((p for p in analyzed_prices if p['symbol'] == symbol), None)
            analyzed_price = analyzed_info.get('price') if analyzed_info else None
            
            message = (
                f"🔍 <b>Data Flow Debug: {symbol}</b>\n\n"
                f"📊 <b>Step 1 - Batch API:</b>\n"
                f"• Present: {'✅ YES' if in_batch else '❌ NO'}\n"
                f"• Price: ${batch_price if batch_price else 'N/A'}\n\n"
                
                f"📊 <b>Step 2 - Individual API:</b>\n"
                f"• Works: {'✅ YES' if individual_works else '❌ NO'}\n"
                f"• Price: ${individual_price if individual_price else 'N/A'}\n\n"
                
                f"📊 <b>Step 3 - get_all_mexc_prices:</b>\n"
                f"• Present: {'✅ YES' if in_all_prices else '❌ NO'}\n"
                f"• Price: ${all_price_value if all_price_value else 'N/A'}\n\n"
                
                f"📊 <b>Step 4 - analyze_price_movements:</b>\n"
                f"• Present: {'✅ YES' if in_analyzed else '❌ NO'}\n"
                f"• Price: ${analyzed_price if analyzed_price else 'N/A'}\n\n"
            )
            
            # Identify where the data is lost
            if in_all_prices and not in_analyzed:
                message += "❌ <b>ISSUE:</b> Data lost in analyze_price_movements()\n"
            elif in_analyzed and analyzed_price is None:
                message += "❌ <b>ISSUE:</b> Price is None in analyzed data\n"
            else:
                message += "✅ <b>DATA FLOW:</b> All steps working\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Data flow debug error: {str(e)}")



    def symbol_debug_command(self, update: Update, context: CallbackContext):
        """Debug why specific symbols don't have price data"""
        try:
            if not context.args:
                update.message.reply_html("Usage: /symboldebug SYMBOL\nExample: /symboldebug DEVVE_USDT")
                return
            
            symbol = context.args[0].upper()
            update.message.reply_html(f"🔧 <b>Debugging symbol:</b> {symbol}")
            
            # Get unique symbols
            unique_futures, _ = self.find_unique_futures_robust()
            
            # Check if symbol is unique
            is_unique = symbol in unique_futures
            batch_data = self.get_mexc_prices_batch_working()
            in_batch = symbol in batch_data
            
            # Try individual price fetch
            individual_price = self.get_mexc_price_data_working(symbol)
            
            # Try alternative formats
            alt_formats = self.get_alternative_symbol_formats(symbol)
            alt_matches = []
            for alt in alt_formats:
                if alt in batch_data:
                    alt_matches.append(f"{alt} (price: ${batch_data[alt].get('price', 'N/A')})")
            
            message = (
                f"🔧 <b>Symbol Debug: {symbol}</b>\n\n"
                f"📊 <b>Status:</b>\n"
                f"• In unique list: {'✅ YES' if is_unique else '❌ NO'}\n"
                f"• In batch API: {'✅ YES' if in_batch else '❌ NO'}\n"
                f"• Individual API: {'✅ WORKING' if individual_price else '❌ FAILED'}\n\n"
            )
            
            if individual_price:
                message += f"💰 <b>Price Data:</b>\n"
                message += f"• Price: ${individual_price['price']}\n"
                message += f"• 5m Change: {individual_price['changes'].get('5m', 0):.2f}%\n"
                message += f"• Source: {individual_price.get('source', 'unknown')}\n\n"
            
            if alt_matches:
                message += f"🔄 <b>Alternative format matches:</b>\n"
                message += "\n".join([f"• {match}" for match in alt_matches[:3]])
            else:
                message += f"🔄 <b>Alternative formats:</b> No matches found\n"
                message += f"   Tried: {', '.join(alt_formats[:3])}\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Debug error: {str(e)}")
            

    def price_debug_command(self, update: Update, context: CallbackContext):
        """Debug price fetching issues"""
        update.message.reply_html("🔧 <b>Debugging price data...</b>")
        
        try:
            # Test with a few known symbols
            test_symbols = [
                'BTC_USDT', 'ETH_USDT', 'ADA_USDT',  # Common symbols that should work
                'WIN_USDT', 'LAZIO_USDT', 'BOOST_USDT',  # Your unique symbols
                'FARM_USDT', 'ZKSYNC_USDT'  # Symbols that showed prices before
            ]
            
            results = []
            successful = 0
            
            for symbol in test_symbols:
                try:
                    price_info = self.get_mexc_price_data(symbol)
                    if price_info and price_info.get('price'):
                        results.append(f"✅ {symbol}: ${price_info['price']}")
                        successful += 1
                    else:
                        results.append(f"❌ {symbol}: No price data")
                    
                    time.sleep(0.1)  # Rate limiting
                    
                except Exception as e:
                    results.append(f"❌ {symbol}: Error - {str(e)}")
            
            # Test batch method
            batch_data = self.get_mexc_prices_batch()
            batch_count = len(batch_data)
            
            message = (
                f"🔧 <b>Price Debug Results</b>\n\n"
                f"Individual API: {successful}/{len(test_symbols)} successful\n"
                f"Batch API: {batch_count} symbols found\n\n"
                f"<b>Detailed Results:</b>\n" + "\n".join(results) +
                f"\n\n<b>Batch API Status:</b> {'✅ WORKING' if batch_count > 0 else '❌ FAILED'}"
            )
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Debug error: {str(e)}")



    def update_google_sheet(self):
        """Update the Google Sheet with fresh data including price analysis"""
        if not self.gs_client or not self.spreadsheet:
            logger.warning("Google Sheets not available for update")
            return
        
        try:
            logger.info("🔄 Starting comprehensive Google Sheet update...")
            
            # Collect fresh data from all exchanges
            all_futures_data = []
            exchanges = {
                'MEXC': self.get_mexc_futures,
                'Binance': self.get_binance_futures,
                'Bybit': self.get_bybit_futures,
                'OKX': self.get_okx_futures,
                'Gate.io': self.get_gate_futures,
                'KuCoin': self.get_kucoin_futures,
                'BingX': self.get_bingx_futures,
                'BitGet': self.get_bitget_futures
            }
            
            exchange_stats = {}
            symbol_coverage = {}
            current_time = datetime.now().astimezone().isoformat()
            
            # Get data from all exchanges
            for name, method in exchanges.items():
                try:
                    futures = method()
                    exchange_stats[name] = len(futures)
                    logger.info(f"{name}: {len(futures)} futures")
                    
                    for symbol in futures:
                        all_futures_data.append({
                            'symbol': symbol,
                            'exchange': name,
                            'timestamp': current_time
                        })
                        
                        # Track symbol coverage
                        normalized = self.normalize_symbol_for_comparison(symbol)
                        if normalized not in symbol_coverage:
                            symbol_coverage[normalized] = set()
                        symbol_coverage[normalized].add(name)
                    
                    time.sleep(0.5)  # Rate limiting
                    
                except Exception as e:
                    logger.error(f"Exchange {name} error during sheet update: {e}")
                    exchange_stats[name] = 0
            
            logger.info(f"Total futures collected: {len(all_futures_data)}")
            logger.info(f"Unique symbols: {len(symbol_coverage)}")
            
            # Get unique futures
            unique_futures, _ = self.find_unique_futures_robust()
            logger.info(f"Unique MEXC futures: {len(unique_futures)}")
            
            # Get price data for analysis
            logger.info("💰 Getting price data for analysis...")
            price_data = self.get_all_mexc_prices()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Update all sheets with fresh data
            self.update_unique_futures_sheet_with_prices(unique_futures, analyzed_prices)
            self.update_all_futures_sheet(self.spreadsheet, all_futures_data, symbol_coverage, current_time)
            self.update_mexc_analysis_sheet_with_prices(all_futures_data, symbol_coverage, analyzed_prices, current_time)
            self.update_price_analysis_sheet(analyzed_prices)
            self.update_exchange_stats_sheet(self.spreadsheet, exchange_stats, current_time)
            self.update_dashboard_with_comprehensive_stats(exchange_stats, len(symbol_coverage), len(unique_futures), analyzed_prices)
            
            logger.info("✅ Google Sheet update completed successfully")
            
        except Exception as e:
            logger.error(f"❌ Google Sheet update error: {e}")

    def update_all_futures_sheet(self, spreadsheet, all_futures_data, symbol_coverage, timestamp):
        """Update All Futures sheet"""
        try:
            worksheet = spreadsheet.worksheet('All Futures')
            
            # Clear existing data (keep headers)
            if worksheet.row_count > 1:
                worksheet.clear()
                # Re-add headers
                worksheet.update('A1', [[
                    'Symbol', 'Exchange', 'Normalized', 'Available On', 
                    'Coverage', 'Timestamp', 'Unique', 'Current Price'
                ]])
            
            all_data = []
            for future in all_futures_data:
                normalized = self.normalize_symbol_for_comparison(future['symbol'])
                exchanges_list = symbol_coverage.get(normalized, set())
                available_on = ", ".join(sorted(exchanges_list))
                coverage = f"{len(exchanges_list)} exchanges"
                is_unique = "✅" if len(exchanges_list) == 1 else ""
                
                all_data.append([
                    future['symbol'],
                    future['exchange'],
                    normalized,
                    available_on,
                    coverage,
                    timestamp,
                    is_unique,
                    'N/A'  # Price would be added separately
                ])
            
            # Write in batches
            if all_data:
                batch_size = 100
                for i in range(0, len(all_data), batch_size):
                    batch = all_data[i:i + batch_size]
                    worksheet.update(f'A{i+2}', batch)
                
                logger.info(f"Updated All Futures with {len(all_data)} records")
            
        except Exception as e:
            logger.error(f"Error updating All Futures sheet: {e}")


    def update_dashboard_stats(self, exchange_stats, unique_symbols_count, unique_futures_count, analyzed_prices):
        """Update dashboard statistics - simplified version"""
        if not self.spreadsheet:
            return
        
        try:
            worksheet = self.spreadsheet.worksheet("Dashboard")
            
            # Count working exchanges
            working_exchanges = sum(1 for count in exchange_stats.values() if count > 0)
            total_exchanges = len(exchange_stats)
            
            # Get current time for next update
            next_update = (datetime.now() + timedelta(minutes=self.update_interval)).strftime('%H:%M:%S')
            
            # Update only the statistics section (rows 23-27)
            stats_update = [
                ["Next Data Update", next_update],
                ["Next Price Update", (datetime.now() + timedelta(minutes=self.price_check_interval)).strftime('%H:%M:%S')],
                ["Unique Futures Count", unique_futures_count],
                ["Working Exchanges", f"{working_exchanges}/{total_exchanges}"],
                ["Total Symbols", unique_symbols_count]
            ]
            
            # Update stats section (starting at row 23)
            for i, (label, value) in enumerate(stats_update):
                worksheet.update(f'A{23+i}:B{23+i}', [[label, value]])
                    
        except Exception as e:
            logger.error(f"Error updating dashboard stats: {e}")


                          

    def update_price_analysis_sheet(self, analyzed_prices):
        """Update Price Analysis sheet with top performers"""
        try:
            # Get or create Price Analysis sheet
            try:
                worksheet = self.spreadsheet.worksheet('Price Analysis')
            except gspread.WorksheetNotFound:
                worksheet = self.spreadsheet.add_worksheet(title='Price Analysis', rows=1000, cols=12)
            
            # Clear existing data
            worksheet.clear()
            
            # Headers
            headers = [
                'Rank', 'Symbol', 'Current Price', '5m %', '15m %', '30m %', 
                '1h %', '4h %', 'Score', 'Trend', 'Volume', 'Last Updated'
            ]
            worksheet.update('A1', [headers])
            
            # Prepare data - top 50 performers
            sheet_data = []
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for i, item in enumerate(analyzed_prices[:50], 1):
                changes = item.get('changes', {})
                
                # Determine trend
                latest_change = item.get('latest_change', 0)
                if latest_change > 5:
                    trend = "🚀 STRONG UP"
                elif latest_change > 2:
                    trend = "🟢 UP"
                elif latest_change < -5:
                    trend = "🔻 STRONG DOWN"
                elif latest_change < -2:
                    trend = "🔴 DOWN"
                else:
                    trend = "⚪ FLAT"
                
                row = [
                    i,
                    item['symbol'],
                    item.get('price', 'N/A'),
                    self.format_change_for_sheet(changes.get('5m')),
                    self.format_change_for_sheet(changes.get('15m')),
                    self.format_change_for_sheet(changes.get('30m')),
                    self.format_change_for_sheet(changes.get('60m')),
                    self.format_change_for_sheet(changes.get('240m')),
                    f"{item.get('score', 0):.2f}",
                    trend,
                    'N/A',  # Volume would require additional API call
                    current_time
                ]
                sheet_data.append(row)
            
            # Update sheet
            if sheet_data:
                worksheet.update('A2', sheet_data)
                logger.info(f"✅ Updated Price Analysis with {len(sheet_data)} top performers")
            else:
                logger.warning("No price data to update")
            
        except Exception as e:
            logger.error(f"Error updating Price Analysis sheet: {e}")

    def get_prices_for_unique_symbols(self):
        """Get price data specifically for all unique symbols"""
        try:
            # Get unique symbols
            unique_futures, _ = self.find_unique_futures_robust()
            logger.info(f"🎯 Getting prices for {len(unique_futures)} unique symbols")
            
            # Start with batch data
            batch_data = self.get_mexc_prices_batch_working()
            price_data = batch_data.copy()
            
            # Check which unique symbols are missing from batch
            missing_symbols = [s for s in unique_futures if s not in batch_data]
            logger.info(f"🔍 {len(missing_symbols)} unique symbols missing from batch API")
            
            # Try alternative symbol formats for missing symbols
            found_with_alt_format = 0
            for symbol in missing_symbols[:]:
                # Try different symbol formats that might be in batch data
                alt_formats = self.get_alternative_symbol_formats(symbol)
                for alt_format in alt_formats:
                    if alt_format in batch_data:
                        price_data[symbol] = batch_data[alt_format]
                        missing_symbols.remove(symbol)
                        found_with_alt_format += 1
                        logger.debug(f"✅ Found {symbol} as {alt_format} in batch")
                        break
            
            if found_with_alt_format > 0:
                logger.info(f"🔄 Found {found_with_alt_format} symbols with alternative formats")
            
            # Get individual prices for remaining missing symbols
            successful_individual = 0
            logger.info(f"🔍 Getting individual prices for {len(missing_symbols)} remaining symbols")
            
            for symbol in missing_symbols[:80]:  # Limit to avoid timeout
                try:
                    price_info = self.get_mexc_price_data_working(symbol)
                    if price_info and price_info.get('price'):
                        price_data[symbol] = price_info
                        successful_individual += 1
                    
                    time.sleep(0.2)  # Conservative rate limiting
                    
                except Exception as e:
                    logger.debug(f"Individual price failed for {symbol}: {e}")
                    continue
            
            logger.info(f"✅ Unique symbols price coverage: {len([s for s in unique_futures if s in price_data])}/{len(unique_futures)}")
            return price_data
            
        except Exception as e:
            logger.error(f"Error getting prices for unique symbols: {e}")
            return {}

    def get_alternative_symbol_formats(self, symbol):
        """Generate alternative symbol formats that might match batch data"""
        alternatives = []
        
        # Common format variations
        if '_' in symbol:
            # Try without underscore
            alternatives.append(symbol.replace('_', ''))
            # Try with different case
            alternatives.append(symbol.lower())
            alternatives.append(symbol.upper())
        
        # Try common suffix variations
        if symbol.endswith('_USDT'):
            base = symbol[:-5]  # Remove _USDT
            alternatives.append(f"{base}USDT")
            alternatives.append(f"{base}-USDT")
            alternatives.append(f"{base}/USDT")
        
        elif symbol.endswith('_USDC'):
            base = symbol[:-5]  # Remove _USDC
            alternatives.append(f"{base}USDC")
            alternatives.append(f"{base}-USDC")
            alternatives.append(f"{base}/USDC")
        
        return alternatives





    def update_dashboard_with_comprehensive_stats(self, exchange_stats, unique_symbols_count, unique_futures_count, analyzed_prices):
        """Update the dashboard with comprehensive statistics including price coverage"""
        if not self.spreadsheet:
            return
        
        try:
            worksheet = self.spreadsheet.worksheet("Dashboard")
            
            # Count working exchanges
            working_exchanges = sum(1 for count in exchange_stats.values() if count > 0)
            total_exchanges = len(exchange_stats)
            
            # Calculate price statistics
            valid_prices = [p for p in analyzed_prices if p.get('price') is not None] if analyzed_prices else []
            top_performers = valid_prices[:10] if valid_prices else []
            strong_movers = [p for p in valid_prices if abs(p.get('latest_change', 0)) > 5]
            
            # Get unique futures for price coverage calculation
            unique_futures, _ = self.find_unique_futures_robust()
            price_coverage = len(valid_prices) / max(len(unique_futures), 1) * 100
            
            stats_update = [
                ["🤖 MEXC FUTURES AUTO-UPDATE DASHBOARD", ""],
                ["Last Updated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Update Interval", f"{self.update_interval} minutes"],
                ["", ""],
                ["📊 EXCHANGE STATISTICS", ""],
                ["Working Exchanges", f"{working_exchanges}/{total_exchanges}"],
                ["Total Unique Symbols", unique_symbols_count],
                ["Unique MEXC Futures", f"{unique_futures_count} 🎯"],
                ["", ""],
                ["💰 PRICE ANALYSIS", ""],
                ["Symbols with Price Data", f"{len(valid_prices)}/{unique_futures_count}"],
                ["Price Data Coverage", f"{price_coverage:.1f}%"],
                ["Top Performers Tracked", len(top_performers)],
                ["Strong Movers (>5%)", len(strong_movers)],
                ["", ""],
                ["⚡ PERFORMANCE", ""],
                ["Next Auto-Update", (datetime.now() + timedelta(minutes=self.update_interval)).strftime('%H:%M:%S')],
                ["Status", "🟢 RUNNING"],
                ["Price System", "✅ WORKING"],
            ]
            
            # Update dashboard
            worksheet.clear()
            worksheet.update(stats_update, 'A1')  # Fixed parameter order
            
            logger.info("✅ Dashboard updated with comprehensive stats")
            
        except Exception as e:
            logger.error(f"Error updating dashboard stats: {e}")

    def format_change_for_sheet(self, change):
        """Format change for Google Sheets with color indicators"""
        if change is None:
            return 'N/A'
        
        # Add emoji based on change value
        if change > 10:
            return f"🚀 {change:+.2f}%"
        elif change > 5:
            return f"🟢 {change:+.2f}%"
        elif change > 2:
            return f"📈 {change:+.2f}%"
        elif change < -10:
            return f"💥 {change:+.2f}%"
        elif change < -5:
            return f"🔴 {change:+.2f}%"
        elif change < -2:
            return f"📉 {change:+.2f}%"
        else:
            return f"{change:+.2f}%"

    # Also update the forceupdate command to use the new method
    def ensure_sheets_initialized(self):
        """Ensure all required sheets exist with comprehensive error handling"""
        try:
            # First, check if Google Sheets is properly configured
            if not self.gs_client or not self.spreadsheet:
                logger.error("❌ Google Sheets not properly configured")
                # Try to reinitialize
                if not self.setup_google_sheets():
                    logger.error("❌ Failed to reinitialize Google Sheets")
                    return False
            
            # Test connection by getting spreadsheet info
            try:
                spreadsheet_title = self.spreadsheet.title
                logger.info(f"📊 Working with spreadsheet: {spreadsheet_title}")
            except Exception as e:
                logger.error(f"❌ Cannot access spreadsheet: {e}")
                return False

            # Define required sheets
            required_sheets = [
                'Dashboard',
                'Unique Futures', 
                'All Futures',
                'MEXC Analysis',
                'Price Analysis',
                'Exchange Stats'
            ]

            # Get existing sheets
            try:
                existing_worksheets = self.spreadsheet.worksheets()
                existing_sheet_names = [sheet.title for sheet in existing_worksheets]
                logger.info(f"📋 Found {len(existing_worksheets)} existing sheets: {existing_sheet_names}")
            except Exception as e:
                logger.error(f"❌ Failed to get existing sheets: {e}")
                return False

            # Create missing sheets
            sheets_created = 0
            for sheet_name in required_sheets:
                try:
                    if sheet_name in existing_sheet_names:
                        logger.info(f"✅ Sheet exists: {sheet_name}")
                        continue
                    
                    # Create new sheet
                    logger.info(f"🆕 Creating new sheet: {sheet_name}")
                    new_sheet = self.spreadsheet.add_worksheet(
                        title=sheet_name, 
                        rows="1000", 
                        cols="20"
                    )
                    
                    # Set basic headers based on sheet type
                    headers = self.get_sheet_headers(sheet_name)
                    if headers:
                        new_sheet.update('A1', [headers])
                        logger.info(f"✅ Added headers to {sheet_name}")
                    
                    sheets_created += 1
                    time.sleep(1)  # Rate limiting
                    
                except Exception as e:
                    logger.error(f"❌ Failed to create sheet {sheet_name}: {e}")
                    continue

            # Setup Dashboard content
            try:
                dashboard = self.spreadsheet.worksheet('Dashboard')
                self.setup_dashboard_sheet(dashboard)
                logger.info("✅ Dashboard setup completed")
            except Exception as e:
                logger.error(f"❌ Failed to setup dashboard: {e}")

            logger.info(f"✅ Sheet initialization complete. Created {sheets_created} new sheets.")
            return True
            
        except Exception as e:
            logger.error(f"❌ Critical error in sheet initialization: {e}")
            return False

    def get_sheet_headers(self, sheet_name):
        """Get appropriate headers for each sheet type"""
        headers_map = {
            'Dashboard': ['Section', 'Value'],
            'Unique Futures': [
                'Symbol', 'Current Price', '5m Change %', '15m Change %', 
                '30m Change %', '1h Change %', '4h Change %', 'Score', 'Status', 'Last Updated'
            ],
            'All Futures': [
                'Symbol', 'Exchange', 'Normalized', 'Available On', 
                'Coverage', 'Timestamp', 'Unique', 'Current Price'
            ],
            'MEXC Analysis': [
                'MEXC Symbol', 'Normalized', 'Available On', 'Exchanges Count', 
                'Current Price', '5m Change %', '1h Change %', '4h Change %', 
                'Status', 'Unique', 'Timestamp'
            ],
            'Price Analysis': [
                'Rank', 'Symbol', 'Current Price', '5m %', '15m %', '30m %', 
                '1h %', '4h %', 'Score', 'Trend', 'Volume', 'Last Updated'
            ],
            'Exchange Stats': [
                'Exchange', 'Futures Count', 'Status', 'Last Updated', 
                'Success Rate', 'Price Data Available'
            ]
        }
        return headers_map.get(sheet_name, [])




    def cleanup_unexpected_sheets(self, existing_sheet_names, expected_sheets):
        """Remove sheets that are not in the expected list"""
        try:
            for sheet_name in existing_sheet_names:
                if sheet_name not in expected_sheets:
                    try:
                        worksheet = self.spreadsheet.worksheet(sheet_name)
                        self.spreadsheet.del_worksheet(worksheet)
                        logger.info(f"🗑️ Removed unexpected sheet: {sheet_name}")
                    except Exception as e:
                        logger.warning(f"Could not remove sheet {sheet_name}: {e}")
        except Exception as e:
            logger.error(f"Error cleaning up sheets: {e}")

    def setup_dashboard_sheet(self, worksheet):
        """Setup the dashboard sheet with comprehensive information"""
        try:
            # Clear existing data
            worksheet.clear()
            
            # Comprehensive dashboard data
            dashboard_data = [
                ["🤖 MEXC FUTURES AUTO-UPDATE DASHBOARD", ""],
                ["Last Updated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Update Interval", f"{self.update_interval} minutes"],
                ["Price Check Interval", f"{self.price_check_interval} minutes"],
                ["", ""],
                ["📊 EXCHANGE MONITORING", ""],
                ["Total Exchanges Tracked", "8"],
                ["Primary Exchange", "MEXC"],
                ["Comparison Exchanges", "Binance, Bybit, OKX, Gate.io, KuCoin, BingX, BitGet"],
                ["", ""],
                ["🎯 UNIQUE FUTURES TRACKING", ""],
                ["Auto Unique Detection", "✅ ENABLED"],
                ["Price Monitoring", "✅ ENABLED"],
                ["Telegram Alerts", "✅ ENABLED"],
                ["Google Sheets Sync", "✅ ENABLED"],
                ["", ""],
                ["💰 PRICE ANALYSIS FEATURES", ""],
                ["Timeframes Tracked", "5m, 15m, 30m, 1h, 4h"],
                ["Top Performers", "Top 50 ranked by score"],
                ["Trend Analysis", "🚀🟢📈🔴📉 emoji indicators"],
                ["Volume Tracking", "⚡ Coming soon"],
                ["", ""],
                ["⚡ REAL-TIME STATS", ""],
                ["Next Data Update", "Will update automatically"],
                ["Next Price Update", "Will update automatically"],
                ["Unique Futures Count", "Will update automatically"],
                ["Top Performer", "Will update automatically"],
                ["", ""],
                ["🔧 SHEETS OVERVIEW", ""],
                ["Dashboard", "This overview and real-time stats"],
                ["Unique Futures", "Futures only on MEXC with prices"],
                ["All Futures", "All futures from all exchanges"],
                ["MEXC Analysis", "Detailed MEXC coverage with prices"],
                ["Price Analysis", "Top 50 performers with trends"],
                ["Exchange Stats", "Exchange performance metrics"],
                ["", ""],
                ["💡 QUICK START", ""],
                ["1. Check /status", "Current bot status"],
                ["2. Use /findunique", "Find unique MEXC futures"],
                ["3. Check /toppers", "Top performing futures"],
                ["4. Use /verifyunique", "Verify symbol uniqueness"],
                ["", ""],
                ["🆘 SUPPORT", ""],
                ["Use /help", "Complete command list"],
                ["Use /check", "Force immediate data update"],
                ["Check logs", "For detailed debugging info"]
            ]
            
            # Update the dashboard
            worksheet.update('A1', dashboard_data)
            
            # Apply formatting
            try:
                # Format title row
                worksheet.format('A1:B1', {
                    'textFormat': {'bold': True, 'fontSize': 14},
                    'backgroundColor': {'red': 0.8, 'green': 0.9, 'blue': 1.0},
                    'horizontalAlignment': 'CENTER'
                })
                
                # Format section headers
                section_rows = [5, 11, 17, 23, 29, 35, 41]
                for row in section_rows:
                    worksheet.format(f'A{row}:B{row}', {
                        'textFormat': {'bold': True},
                        'backgroundColor': {'red': 0.95, 'green': 0.95, 'blue': 0.95}
                    })
                    
            except Exception as format_error:
                logger.warning(f"⚠️ Could not format dashboard: {format_error}")
            
            logger.info("✅ Dashboard content updated")
            
        except Exception as e:
            logger.error(f"❌ Error setting up dashboard: {e}")
            raise


    def update_google_sheet_dashboard(self):
        """Update Google Sheet dashboard with current statistics"""
        if not self.gs_client or not self.spreadsheet:
            logger.warning("Google Sheets not available for dashboard update")
            return
        
        try:
            # Get current data
            unique_futures, exchange_stats = self.find_unique_futures_robust()
            
            # Get all futures data for statistics
            all_futures_data = []
            exchanges = {
                'MEXC': self.get_mexc_futures,
                'Binance': self.get_binance_futures,
                'Bybit': self.get_bybit_futures,
                'OKX': self.get_okx_futures,
                'Gate.io': self.get_gate_futures,
                'KuCoin': self.get_kucoin_futures,
                'BingX': self.get_bingx_futures,
                'BitGet': self.get_bitget_futures
            }
            
            symbol_coverage = {}
            for name, method in exchanges.items():
                try:
                    futures = method()
                    for symbol in futures:
                        all_futures_data.append({
                            'symbol': symbol,
                            'exchange': name,
                            'timestamp': datetime.now().isoformat()
                        })
                        
                        # Track symbol coverage
                        normalized = self.normalize_symbol_for_comparison(symbol)
                        if normalized not in symbol_coverage:
                            symbol_coverage[normalized] = set()
                        symbol_coverage[normalized].add(name)
                except Exception as e:
                    logger.error(f"Error getting {name} data: {e}")
            
            # Get price data
            price_data = self.get_all_mexc_prices()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Update dashboard
            self.update_dashboard_with_comprehensive_stats(exchange_stats, len(symbol_coverage), len(unique_futures), analyzed_prices)
            
            logger.info("✅ Google Sheet dashboard updated")
            
        except Exception as e:
            logger.error(f"Error updating Google Sheet dashboard: {e}")
            

    def force_update_command(self, update: Update, context: CallbackContext):
        """Force immediate Google Sheet update with comprehensive data"""
        try:
            update.message.reply_html("🔄 <b>Force updating Google Sheet with comprehensive data...</b>")
            
            # Step 1: Initialize Google Sheets connection
            update.message.reply_html("🔧 <b>Step 1:</b> Initializing Google Sheets connection...")
            
            if not self.gs_client or not self.spreadsheet:
                update.message.reply_html("🔄 Reinitializing Google Sheets connection...")
                if not self.setup_google_sheets():
                    update.message.reply_html(
                        "❌ <b>Failed to initialize Google Sheets.</b>\n\n"
                        "Please check:\n"
                        "1. GOOGLE_CREDENTIALS_JSON environment variable\n"
                        "2. GOOGLE_SHEET_EMAIL environment variable\n"  # CHANGED: ID → EMAIL
                        "3. Spreadsheet sharing permissions\n"
                        "4. Service account has editor access to the spreadsheet"
                    )
                    return False
            
            # Step 2: Ensure sheets are initialized
            update.message.reply_html("📋 <b>Step 2:</b> Ensuring all sheets are initialized...")
            if not self.ensure_sheets_initialized():
                update.message.reply_html("❌ <b>Failed to initialize sheets.</b>\n\nPlease check if the Google Sheet exists and is accessible.")
                return False
            
            # Step 3: Run the comprehensive update
            update.message.reply_html("📊 <b>Step 3:</b> Running comprehensive data update...")
            self.update_google_sheet()
            
            # Get spreadsheet URL for the message
            sheet_url = self.spreadsheet.url if self.spreadsheet else 'Not available'
            
            update.message.reply_html(
                f"✅ <b>Google Sheet updated successfully!</b>\n\n"
                f"📊 <a href='{sheet_url}'>Open Your Sheet</a>\n\n"
                f"<b>Sheets Updated:</b>\n"
                f"• 📈 Dashboard - Overview and stats\n"
                f"• 🎯 Unique Futures - MEXC-only symbols\n"
                f"• 📋 All Futures - All exchange data\n"
                f"• 🔍 MEXC Analysis - Detailed coverage\n"
                f"• 💰 Price Analysis - Top performers\n"
                f"• 📊 Exchange Stats - Performance metrics",
                reply_markup=ReplyKeyboardRemove()
            )
            return True
            
        except Exception as e:
            error_msg = (
                f"❌ <b>Force update failed:</b>\n\n"
                f"<b>Error:</b> {str(e)}\n\n"
                f"<b>Debugging steps:</b>\n"
                f"1. Check GOOGLE_CREDENTIALS_JSON is valid JSON\n"
                f"2. Verify GOOGLE_SHEET_EMAIL is correct\n"  # CHANGED: ID → EMAIL
                f"3. Ensure service account has edit permissions\n"
                f"4. Check if spreadsheet exists and is accessible"
            )
            update.message.reply_html(error_msg)
            logger.error(f"Force update command error: {e}")
            return False
        
    def _make_request_with_retry(self, url: str, timeout: int = 15, max_retries: int = 3) -> Optional[requests.Response]:
        """Make request with retry logic and proxy rotation"""
        for attempt in range(max_retries):
            try:
                proxy = random.choice(self.proxies) if self.proxies else {}
                response = self.session.get(url, timeout=timeout, proxies=proxy if proxy else None)
                
                if response.status_code == 200:
                    return response
                elif response.status_code in [403, 429]:
                    logger.warning(f"⚠️  Blocked on attempt {attempt + 1} for {url}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # Exponential backoff
                        continue
                else:
                    logger.error(f"❌ HTTP {response.status_code} for {url}")
                    break
                    
            except requests.exceptions.RequestException as e:
                logger.warning(f"⚠️  Request failed on attempt {attempt + 1}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
        
        return None



    def format_start_time(self, start_time):
        """Format start time for display"""
        if start_time:
            try:
                # Handle both string and datetime objects
                if isinstance(start_time, str):
                    dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                else:
                    dt = start_time
                return dt.strftime("%Y-%m-%d %H:%M")
            except:
                pass
        return "Unknown"
    
    def get_uptime(self):
        """Calculate bot uptime"""
        data = self.load_data()
        start_time = data.get('statistics', {}).get('start_time')
        if start_time:
            try:
                if isinstance(start_time, str):
                    start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                else:
                    start_dt = start_time
                uptime = datetime.now() - start_dt
                days = uptime.days
                hours = uptime.seconds // 3600
                minutes = (uptime.seconds % 3600) // 60
                return f"{days}d {hours}h {minutes}m"
            except:
                pass
        return "Unknown"            

    def start_command(self, update: Update, context: CallbackContext):
        """Send welcome message"""
        user = update.effective_user
        welcome_text = (
            f"🤖 Hello {user.mention_html()}!\n\n"
            "I'm <b>MEXC Unique Futures Tracker</b>\n\n"
            "<b>Features:</b>\n"
            "• Real-time monitoring of 8 exchanges\n"
            "• Unique futures detection\n"
            "• Price movement analysis\n"
            "• Automatic alerts\n"
            "• Google Sheets integration\n\n"
            "<b>Commands:</b>\n"
            "/start - Welcome message\n"
            "/status - Current status\n"
            "/check - Immediate check\n"
            "/pricedebug - Price debug\n"
            "/symboldebug - Symbol debug\n"
            "/excel - Download excel\n"
            "/datastatus - Data status\n"
            "/analysis - Full analysis\n"
            "/growth - Growth chart\n"
            "/growthreport - Growthreport chart\n"
            "/4hchart - 4h chart\n"
            "/trends - Trends chart\n"
            "/dataflow - Dataflow Symbol\n"
            "/exchanges - Exchange info\n"
            "/stats - Bot statistics\n"
            "/help - Help information\n"
            "/findunique - Find unique futures\n"
            "/forceupdate - Force update Google Sheet\n"
            "/checksymbol SYMBOL - Check specific symbol\n"
            "/prices - Check current prices\n"
            "/toppers - Top performing futures\n\n"
            "⚡ <i>Happy trading!</i>"
        )
        update.message.reply_html(welcome_text)

    def prices_command(self, update: Update, context: CallbackContext):
        """Get current price information for unique futures"""
        update.message.reply_html("📊 <b>Getting current prices...</b>")
        
        try:
            unique_futures, _ = self.find_unique_futures_robust()
            price_data = self.get_all_mexc_prices()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Filter to only unique futures
            unique_prices = [p for p in analyzed_prices if p['symbol'] in unique_futures]
            
            if not unique_prices:
                update.message.reply_html("❌ No price data available for unique futures")
                return
            
            message = "💰 <b>Unique Futures Prices</b>\n\n"
            
            for i, item in enumerate(unique_prices[:10], 1):  # Show top 10
                changes = item.get('changes', {})
                message += f"{i}. <b>{item['symbol']}</b>\n"
                message += f"   Price: ${item['price']:.4f}\n"
                
                if '5m' in changes:
                    message += f"   5m: {self.format_change(changes['5m'])}\n"
                if '1h' in changes:
                    message += f"   1h: {self.format_change(changes.get('60m', 0))}\n"
                if '4h' in changes:
                    message += f"   4h: {self.format_change(changes.get('240m', 0))}\n"
                
                message += "\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error getting prices: {str(e)}")

    def top_performers_command(self, update: Update, context: CallbackContext):
        """Show top performing futures"""
        update.message.reply_html("🚀 <b>Analyzing top performers...</b>")
        
        try:
            price_data = self.get_all_mexc_prices()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            if not analyzed_prices:
                update.message.reply_html("❌ No price data available")
                return
            
            message = "🏆 <b>Top Performing Futures</b>\n\n"
            
            for i, item in enumerate(analyzed_prices[:15], 1):
                changes = item.get('changes', {})
                message += f"{i}. <b>{item['symbol']}</b>\n"
                message += f"   Price: ${item['price']:.4f}\n"
                
                change_5m = changes.get('5m', 0)
                change_1h = changes.get('60m', 0)
                change_4h = changes.get('240m', 0)
                
                message += f"   5m: {self.format_change(change_5m)}"
                message += f" | 1h: {self.format_change(change_1h)}"
                message += f" | 4h: {self.format_change(change_4h)}\n"
                
                # Add emoji for very strong performers
                if change_5m > 10 or change_1h > 20:
                    message += "   🚀 <b>STRONG UPTREND!</b>\n"
                elif change_5m > 5 or change_1h > 10:
                    message += "   📈 <b>Uptrend</b>\n"
                
                message += "\n"
            
            update.message.reply_html(message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error analyzing performers: {str(e)}")

    def check_command(self, update: Update, context: CallbackContext):
        """Perform immediate check with colorful visual progress bar - FIXED PRICE DATA FLOW"""
        try:
            # Send initial message
            progress_message = update.message.reply_html(
                "🚀 <b>Starting Comprehensive Exchange Analysis</b>\n\n"
                "⚡ Initializing tracking systems...\n"
                "▰▱▱▱▱▱▱▱▱▱ 10%"
            )
            
            def update_progress(step, total_steps, status, current_exchange=None, count=None):
                """Update progress bar with colors"""
                try:
                    # Calculate progress
                    progress_percent = (step / total_steps) * 100
                    filled_blocks = int(progress_percent / 10)
                    empty_blocks = 10 - filled_blocks
                    
                    # Colorful progress bar based on completion
                    if progress_percent < 30:
                        progress_bar = "🟦" * filled_blocks + "⬜" * empty_blocks
                    elif progress_percent < 70:
                        progress_bar = "🟨" * filled_blocks + "⬜" * empty_blocks
                    else:
                        progress_bar = "🟩" * filled_blocks + "⬜" * empty_blocks
                    
                    # Build animated status
                    spinner = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"][step % 10]
                    
                    message = f"🚀 <b>Comprehensive Exchange Analysis</b>\n\n"
                    message += f"{spinner} <b>Progress:</b> {progress_bar} {progress_percent:.0f}%\n"
                    message += f"📝 <b>Status:</b> {status}\n"
                    
                    if current_exchange and count is not None:
                        if count > 0:
                            message += f"✅ <b>{current_exchange}:</b> {count} futures found\n"
                        else:
                            message += f"❌ <b>{current_exchange}:</b> Failed\n"
                    
                    message += f"\n⏰ Step {step+1} of {total_steps}"
                    
                    # Update the message
                    context.bot.edit_message_text(
                        chat_id=update.effective_chat.id,
                        message_id=progress_message.message_id,
                        text=message,
                        parse_mode='HTML'
                    )
                except Exception as e:
                    logger.debug(f"Progress update failed: {e}")

            # Define check steps with more detail
            steps = [
                ("Initializing systems", "⚡ Starting tracking systems..."),
                ("Checking MEXC", "🔍 Scanning MEXC futures database..."),
                ("Checking Binance", "🌐 Connecting to Binance API..."),
                ("Checking Bybit", "🔄 Accessing Bybit perpetuals..."),
                ("Checking OKX", "📊 Analyzing OKX swap contracts..."),
                ("Checking Gate.io", "🔍 Scanning Gate.io futures..."),
                ("Checking KuCoin", "📈 Checking KuCoin derivatives..."),
                ("Checking BingX", "🔄 Accessing BingX futures..."),
                ("Checking BitGet", "🔍 Analyzing BitGet perpetuals..."),
                ("Finding unique symbols", "🎯 Calculating unique futures..."),
                ("Collecting price data", "💰 Fetching current prices..."),
                ("Analyzing results", "📊 Compiling comprehensive report..."),
                ("Finalizing", "✅ Completing analysis...")
            ]

            exchange_results = {}
            data_before = self.load_data()
            unique_before = set(data_before.get('unique_futures', []))
            
            # VARIABLES TO STORE RESULTS
            unique_after = set()
            new_futures = set()
            lost_futures = set()
            price_data = {}
            
            # Execute each step with progress updates
            for i, (step_name, status_text) in enumerate(steps):
                try:
                    current_count = 0
                    current_exchange = None
                    
                    if step_name.startswith("Checking "):
                        current_exchange = step_name.replace("Checking ", "")
                    
                    # Update progress
                    update_progress(i, len(steps), status_text, current_exchange, current_count)
                    time.sleep(0.8)  # Smooth animation
                    
                    # Execute the actual step
                    if step_name == "Checking MEXC":
                        mexc_futures = self.get_mexc_futures()
                        exchange_results['MEXC'] = len(mexc_futures)
                        current_count = len(mexc_futures)
                        
                    elif step_name == "Checking Binance":
                        binance_futures = self.get_binance_futures()
                        exchange_results['Binance'] = len(binance_futures)
                        current_count = len(binance_futures)
                        
                    elif step_name == "Checking Bybit":
                        bybit_futures = self.get_bybit_futures()
                        exchange_results['Bybit'] = len(bybit_futures)
                        current_count = len(bybit_futures)
                        
                    elif step_name == "Checking OKX":
                        okx_futures = self.get_okx_futures()
                        exchange_results['OKX'] = len(okx_futures)
                        current_count = len(okx_futures)
                        
                    elif step_name == "Checking Gate.io":
                        gate_futures = self.get_gate_futures()
                        exchange_results['Gate.io'] = len(gate_futures)
                        current_count = len(gate_futures)
                        
                    elif step_name == "Checking KuCoin":
                        kucoin_futures = self.get_kucoin_futures()
                        exchange_results['KuCoin'] = len(kucoin_futures)
                        current_count = len(kucoin_futures)
                        
                    elif step_name == "Checking BingX":
                        bingx_futures = self.get_bingx_futures()
                        exchange_results['BingX'] = len(bingx_futures)
                        current_count = len(bingx_futures)
                        
                    elif step_name == "Checking BitGet":
                        bitget_futures = self.get_bitget_futures()
                        exchange_results['BitGet'] = len(bitget_futures)
                        current_count = len(bitget_futures)
                        
                    elif step_name == "Finding unique symbols":
                        # Get unique futures directly
                        unique_after, exchange_stats = self.find_unique_futures_robust()
                        
                        # Calculate changes
                        new_futures = unique_after - unique_before
                        lost_futures = unique_before - unique_after
                        current_count = len(unique_after)
                        
                    elif step_name == "Collecting price data":
                        # CRITICAL FIX: Use the EXACT SAME method as symbolsearch
                        # Get batch data directly (what symbolsearch uses)
                        batch_data = self.get_consistent_price_data()
                        logger.info(f"📊 Batch data collected: {len(batch_data)} symbols")
                        
                        # Create price_data by matching unique symbols with batch data
                        price_data = {}
                        matched_symbols = 0
                        
                        for symbol in unique_after:
                            # Try exact match first
                            if symbol in batch_data:
                                price_data[symbol] = batch_data[symbol]
                                matched_symbols += 1
                            else:
                                # Try alternative formats (what symbolsearch does)
                                alt_formats = [
                                    symbol.replace('_', ''),
                                    symbol.replace('_', '-'), 
                                    symbol.replace('_', '/'),
                                ]
                                
                                found = False
                                for alt_format in alt_formats:
                                    if alt_format in batch_data:
                                        price_data[symbol] = batch_data[alt_format].copy()
                                        price_data[symbol]['symbol'] = symbol  # Fix symbol name
                                        matched_symbols += 1
                                        found = True
                                        break
                                
                                if not found:
                                    # Symbol not found in batch, add with None price
                                    price_data[symbol] = {
                                        'symbol': symbol,
                                        'price': None,
                                        'changes': {},
                                        'timestamp': datetime.now(),
                                        'source': 'not_found'
                                    }
                        
                        logger.info(f"💰 Price data: {matched_symbols}/{len(unique_after)} symbols matched")
                        current_count = matched_symbols
                        
                    elif step_name == "Analyzing results":
                        # Analysis is already done, just update progress
                        pass
                        
                    # Update progress with results
                    update_progress(i, len(steps), status_text, current_exchange, current_count)
                        
                except Exception as e:
                    logger.error(f"Step {step_name} failed: {e}")
                    if step_name.startswith("Checking "):
                        exchange_name = step_name.replace("Checking ", "")
                        exchange_results[exchange_name] = 0
                        update_progress(i, len(steps), f"❌ {status_text}", exchange_name, 0)

            # Final progress update
            update_progress(len(steps), len(steps), "✅ Check complete!", exchange_results)
            time.sleep(1)

            # Build final results message
            working_exchanges = [name for name, count in exchange_results.items() if count > 0]
            total_futures = sum(exchange_results.values())
            
            # Calculate price coverage statistics
            unique_with_prices = len([s for s in unique_after if s in price_data and price_data[s].get('price') is not None])
            price_coverage_percent = (unique_with_prices / len(unique_after)) * 100 if unique_after else 0

            # DEBUG: Log what we found for specific symbols
            debug_symbols = ['METASTOCK_USDT', 'TRY_USDT', 'BOBBSC_USDT']
            logger.info("🔍 DEBUG - Checking specific symbols in price_data:")
            for symbol in debug_symbols:
                if symbol in price_data:
                    price_info = price_data[symbol]
                    logger.info(f"  {symbol}: ${price_info.get('price')} (source: {price_info.get('source')})")
                else:
                    logger.info(f"  {symbol}: NOT in price_data")

            # Create final report WITH PRICE DATA
            final_message = "🎯 <b>COMPREHENSIVE CHECK COMPLETE</b>\n\n"
            
            # Exchange Statistics
            final_message += "📊 <b>EXCHANGE STATISTICS</b>\n"
            final_message += f"✅ Working: {len(working_exchanges)}/{len(exchange_results)} exchanges\n"
            final_message += f"📈 Total Futures: {total_futures}\n"
            final_message += f"🎯 MEXC Unique: {len(unique_after)}\n"
            final_message += f"💰 Price Coverage: {unique_with_prices}/{len(unique_after)} ({price_coverage_percent:.1f}%)\n\n"
            
            # Detailed Exchange Results
            final_message += "🔍 <b>DETAILED RESULTS</b>\n"
            for exchange in ['MEXC', 'Binance', 'Bybit', 'OKX', 'Gate.io', 'KuCoin', 'BingX', 'BitGet']:
                count = exchange_results.get(exchange, 0)
                status = "✅" if count > 0 else "❌"
                final_message += f"{status} {exchange}: {count} futures\n"
            
            # Changes detected
            final_message += f"\n🔄 <b>CHANGES DETECTED</b>\n"
            if new_futures:
                final_message += f"🆕 New Unique: {len(new_futures)}\n"
                # Show first 3 new symbols
                for i, symbol in enumerate(list(new_futures)[:3], 1):
                    final_message += f"   {i}. {symbol}\n"
                if len(new_futures) > 3:
                    final_message += f"   ... and {len(new_futures) - 3} more\n"
            else:
                final_message += "🆕 New Unique: None\n"
                
            if lost_futures:
                final_message += f"📉 Lost Unique: {len(lost_futures)}\n"
            else:
                final_message += "📉 Lost Unique: None\n"
            
            # Performance summary
            final_message += f"\n⚡ <b>SUMMARY</b>\n"
            final_message += f"📊 MEXC Coverage: {len(unique_after)}/{exchange_results.get('MEXC', 0)} unique\n"
            final_message += f"🔄 Unique Ratio: {(len(unique_after)/exchange_results.get('MEXC', 1)*100):.1f}%\n"
            final_message += f"⏰ Next Auto-check: {self.update_interval} minutes\n\n"
            
            final_message += "✅ <i>Check completed successfully!</i>"

            # ADD NEW SECTION: SHOW NEW UNIQUE FUTURES WITH PRICES
            if new_futures:
                final_message += f"\n\n🚀 <b>NEW UNIQUE FUTURES FOUND!</b>\n\n"
                
                priced_count = 0
                for symbol in list(new_futures)[:10]:  # Show first 10
                    price_info = price_data.get(symbol)
                    
                    if price_info and price_info.get('price') is not None:
                        price = price_info['price']
                        changes = price_info.get('changes', {})
                        change_5m = changes.get('5m', 0)
                        
                        final_message += f"✅ <b>{symbol}</b>\n"
                        final_message += f"   Price: ${price}\n"
                        final_message += f"   5m: {self.format_change(change_5m)}\n\n"
                        priced_count += 1
                    else:
                        final_message += f"✅ <b>{symbol}</b> (price data unavailable)\n\n"
                
                if len(new_futures) > 10:
                    final_message += f"... and {len(new_futures) - 10} more symbols\n\n"
                
                final_message += f"📊 Total unique: <b>{len(unique_after)}</b>\n"
                final_message += f"💰 With prices: <b>{priced_count}/10</b> shown symbols"

            # Send final message
            context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=progress_message.message_id,
                text=final_message,
                parse_mode='HTML'
            )

        except Exception as e:
            # Error handling
            error_message = (
                "❌ <b>CHECK FAILED</b>\n\n"
                f"<b>Error:</b> {str(e)}\n\n"
                "🔧 <i>The check encountered an unexpected error. "
                "This might be due to network issues or exchange API problems. "
                "Try again in a few moments.</i>"
            )
            
            try:
                context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=progress_message.message_id,
                    text=error_message,
                    parse_mode='HTML'
                )
            except:
                # If we can't edit, send new message
                update.message.reply_html(error_message)
            
            logger.error(f"Check command failed: {e}")
                

    def find_unique_command(self, update: Update, context: CallbackContext):
        """Find and display currently unique symbols with prices"""
        update.message.reply_html("🔍 Scanning for unique MEXC symbols with prices...")
        
        try:
            unique_futures, exchange_stats = self.find_unique_futures_robust()
            price_data = self.get_all_mexc_prices()
            
            if not unique_futures:
                update.message.reply_html("❌ No unique symbols found on MEXC")
                return
            
            # Get price info for unique futures
            unique_with_prices = []
            for symbol in unique_futures:
                price_info = price_data.get(symbol)
                if price_info:
                    unique_with_prices.append({
                        'symbol': symbol,
                        'price': price_info['price'],
                        'changes': price_info.get('changes', {})
                    })
                else:
                    unique_with_prices.append({
                        'symbol': symbol,
                        'price': None,
                        'changes': {}
                    })
            
            # Sort by 5m change if available
            unique_with_prices.sort(key=lambda x: x['changes'].get('5m', 0), reverse=True)
            
            response = f"🎯 <b>Unique MEXC Symbols: {len(unique_futures)}</b>\n\n"
            
            for i, item in enumerate(unique_with_prices[:15], 1):
                response += f"{i}. <b>{item['symbol']}</b>"
                if item['price']:
                    response += f" - ${item['price']:.4f}"
                    if '5m' in item['changes']:
                        response += f" {self.format_change(item['changes']['5m'])}"
                response += "\n"
            
            if len(unique_with_prices) > 15:
                response += f"\n... and {len(unique_with_prices) - 15} more symbols"
            
            response += f"\n\n💡 Use /prices for detailed price info"
            
            update.message.reply_html(response)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error finding unique symbols: {str(e)}")

    def check_symbol_command(self, update: Update, context: CallbackContext):
        """Check if a symbol is unique to MEXC"""
        if not context.args:
            update.message.reply_html("Usage: /checksymbol SYMBOL\nExample: /checksymbol BTC")
            return
        
        symbol = context.args[0].upper()
        update.message.reply_html(f"🔍 Checking symbol: {symbol}")
        
        try:
            coverage = self.verify_symbol_coverage(symbol)
            
            if not coverage:
                response = f"❌ Symbol not found on any exchange: {symbol}"
            elif len(coverage) == 1 and 'MEXC' in coverage:
                response = f"🎯 <b>UNIQUE TO MEXC!</b>\n\n{symbol} - Only available on: <b>MEXC</b>"
            elif 'MEXC' in coverage:
                other_exchanges = [e for e in coverage if e != 'MEXC']
                response = (f"📊 <b>{symbol} - Multi-Exchange</b>\n\n"
                        f"✅ Available on MEXC\n"
                        f"🔸 Also on: {', '.join(other_exchanges)}\n"
                        f"📈 Total exchanges: {len(coverage)}")
            else:
                response = f"📊 <b>{symbol}</b>\n\nNot on MEXC, available on:\n• " + "\n• ".join(coverage)
            
            update.message.reply_html(response)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error checking symbol: {str(e)}")

    def status_command(self, update: Update, context: CallbackContext):
        """Send current status"""
        data = self.load_data()
        unique_count = len(data.get('unique_futures', []))
        last_check = data.get('last_check', 'Never')
        exchange_stats = data.get('exchange_stats', {})
        
        if last_check != 'Never':
            try:
                last_dt = datetime.fromisoformat(last_check.replace('Z', '+00:00'))
                last_check = last_dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
        
        status_text = (
            "📈 <b>Bot Status</b>\n\n"
            f"🎯 Current unique: <b>{unique_count}</b>\n"
            f"📅 Last check: {last_check}\n"
            f"⚡ Auto-check: {self.update_interval}min\n"
        )
        
        # Show exchange status
        working_exchanges = [name for name, count in exchange_stats.items() if count > 0]
        status_text += f"✅ Working exchanges: {len(working_exchanges)}/7\n"
        
        # Show unique futures if any
        if unique_count > 0:
            status_text += "\n<b>🎯 Unique futures:</b>\n"
            for symbol in sorted(data['unique_futures'])[:5]:
                status_text += f"• {symbol}\n"
            if unique_count > 5:
                status_text += f"• ... and {unique_count - 5} more"
        
        update.message.reply_html(status_text)

    def analysis_command(self, update: Update, context: CallbackContext):
        """Create comprehensive analysis with both Google Sheet and Excel updates"""
        update.message.reply_html("📈 <b>Creating comprehensive analysis...</b>")
        
        try:
            # Step 1: Update Google Sheet dashboard first
            update.message.reply_html("🔄 <b>Step 1:</b> Updating Google Sheet dashboard...")
            self.update_google_sheet_dashboard()
            
            # Step 2: Get fresh data for reports
            update.message.reply_html("📊 <b>Step 2:</b> Gathering latest data...")
            unique_futures, exchange_stats = self.find_unique_futures_robust()
            
            # Step 3: Create and send text report
            update.message.reply_html("📄 <b>Step 3:</b> Creating text report...")
            report = self.create_analysis_report(unique_futures, exchange_stats)
            file_obj = io.BytesIO(report.encode('utf-8'))
            file_obj.name = f'mexc_analysis_{datetime.now().strftime("%Y%m%d_%H%M")}.txt'
            
            update.message.reply_document(
                document=file_obj,
                caption=f"📊 <b>MEXC Analysis Complete</b>\n\n"
                    f"🎯 Unique futures: {len(unique_futures)}\n"
                    f"🏢 Exchanges: {len(exchange_stats) + 1}\n"
                    f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                parse_mode='HTML'
            )
            
            # Step 4: Create and send Excel file
            update.message.reply_html("📁 <b>Step 4:</b> Creating Excel report...")
            self.create_and_send_excel(update, context)
            
        except Exception as e:
            update.message.reply_html(f"❌ <b>Analysis error:</b>\n{str(e)}")

    def create_analysis_report(self, unique_futures, exchange_stats):
        """Create comprehensive analysis report"""
        report = []
        report.append("=" * 60)
        report.append("🎯 MEXC UNIQUE FUTURES ANALYSIS REPORT")
        report.append("=" * 60)
        report.append(f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        
        # Exchange statistics
        report.append("🏭 EXCHANGE STATISTICS:")
        total_futures = sum(exchange_stats.values())
        report.append(f"  MEXC: {len(self.get_mexc_futures())} futures")
        for exchange, count in exchange_stats.items():
            status = "✅" if count > 0 else "❌"
            report.append(f"  {status} {exchange}: {count} futures")
        
        report.append(f"  Total futures from other exchanges: {total_futures}")
        report.append("")
        
        # Unique futures
        report.append(f"🎯 UNIQUE MEXC FUTURES ({len(unique_futures)}):")
        if unique_futures:
            for i, symbol in enumerate(sorted(unique_futures), 1):
                report.append(f"  {i:2d}. {symbol}")
        else:
            report.append("  No unique futures found")
        
        report.append("")
        report.append("📊 ANALYSIS SUMMARY:")
        report.append(f"  MEXC futures analyzed: {len(self.get_mexc_futures())}")
        report.append(f"  Unique ratio: {len(unique_futures)}/{len(self.get_mexc_futures())}")
        report.append(f"  Market coverage: {len(exchange_stats) + 1} exchanges")
        
        report.append("=" * 60)
        
        return "\n".join(report)

    def exchanges_command(self, update: Update, context: CallbackContext):
        """Show exchange information"""
        data = self.load_data()
        exchange_stats = data.get('exchange_stats', {})
        
        exchanges_text = "🏢 <b>Supported Exchanges</b>\n\n"
        exchanges_text += "🎯 <b>MEXC</b> (source)\n"
        
        if exchange_stats:
            exchanges_text += "\n<b>Other exchanges:</b>\n"
            for exchange, count in sorted(exchange_stats.items()):
                status = "✅" if count > 0 else "❌"
                exchanges_text += f"{status} {exchange}: {count} futures\n"
        else:
            exchanges_text += "\nNo data. Use /check first."
        
        exchanges_text += f"\n🔍 Monitoring <b>{len(exchange_stats) + 1}</b> exchanges"
        
        update.message.reply_html(exchanges_text)

    def stats_command(self, update: Update, context: CallbackContext):
        """Show statistics"""
        data = self.load_data()
        stats = data.get('statistics', {})
        exchange_stats = data.get('exchange_stats', {})
        
        stats_text = (
            "📈 <b>Bot Statistics</b>\n\n"
            f"🔄 Checks performed: <b>{stats.get('checks_performed', 0)}</b>\n"
            f"🎯 Max unique found: <b>{stats.get('unique_found_total', 0)}</b>\n"
            f"⏰ Current unique: <b>{len(data.get('unique_futures', []))}</b>\n"
            f"🏢 Exchanges: <b>{len(exchange_stats) + 1}</b>\n"
            f"📅 Running since: {self.format_start_time(stats.get('start_time'))}\n"
            f"🤖 Uptime: {self.get_uptime()}\n"
            f"⚡ Auto-check: {self.update_interval}min"
        )
        
        update.message.reply_html(stats_text)

    def help_command(self, update: Update, context: CallbackContext):
        """Show help information"""
        help_text = (
            "🆘 <b>MEXC Futures Tracker - Help</b>\n\n"
            "<b>Monitoring 8 exchanges:</b>\n"
            "MEXC, Binance, Bybit, OKX,\n"
            "Gate.io, KuCoin, BingX, BitGet\n\n"
            "<b>Main commands:</b>\n"
            "/check - Quick check for unique futures\n"
            "/pricedebug - Price debug\n"
            "/symboldebug - Symbol debug\n"
            "/excel - Download excel\n"
            "/datastatus - Data status\n"
            "/analysis - Full analysis report\n"
            "/growth - Growth chart\n"
            "/growthreport - Growthreport chart\n"
            "/4hchart - 4h chart\n"
            "/trends - Trends chart\n"
            "/dataflow - Dataflow Symbol\n"
            "/status - Current status\n"
            "/exchanges - Exchange information\n"
            "/stats - Bot statistics\n"
            "/findunique - Find currently unique symbols\n"
            "/forceupdate - Force update Google Sheet\n"
            "/checksymbol SYMBOL - Check specific symbol\n\n"
            f"⚡ Auto-checks every {self.update_interval} minutes\n"
            "🎯 Alerts for new unique futures\n"
            "📊 Comprehensive analysis available\n\n"
            "⚡ <i>Happy trading!</i>"
        )
        update.message.reply_html(help_text)


    # ==================== SCHEDULER ====================

    def setup_scheduler(self):
        """Setup scheduled tasks with rate limiting and error handling"""
        try:
            # Clear any existing schedules
            schedule.clear()
            
            # Initialize rate limiting attributes
            self.last_sheets_update = 0
            self.sheets_update_interval = 300  # 5 minutes minimum between Sheets updates
            self.sheets_retry_count = 0
            self.sheets_circuit_breaker = CircuitBreaker(failure_threshold=3, recovery_timeout=300)
            
            # Unique futures monitoring (less frequent to reduce API calls)
            schedule.every(self.update_interval).minutes.do(
                self.safe_monitor_unique_futures_changes
            )
            
            # Price monitoring (more frequent but doesn't use Sheets API)
            schedule.every(self.price_check_interval).minutes.do(
                self.safe_run_price_monitoring
            )
            
            # Google Sheets update with rate limiting (increased to 5 minutes)
            schedule.every(5).minutes.do(
                self.safe_update_google_sheet_with_prices
            )
            
            # 4-hour chart reporting
            schedule.every(4).hours.do(
                self.safe_send_4h_growth_chart
            )
            
            # Data cleanup (once per day)
            schedule.every().day.at("02:00").do(
                self.safe_cleanup_old_price_data
            )
            
            # Health check every 30 minutes
            schedule.every(30).minutes.do(
                self.health_check
            )
            
            logger.info(f"✅ Optimized scheduler setup complete:")
            logger.info(f"   - Unique check: every {self.update_interval} minutes")
            logger.info(f"   - Price check: every {self.price_check_interval} minutes") 
            logger.info(f"   - Google Sheets: every 5 minutes (rate limited)")
            logger.info(f"   - 4h charts: every 4 hours")
            logger.info(f"   - Cleanup: daily at 02:00")
            logger.info(f"   - Health check: every 30 minutes")
            
        except Exception as e:
            logger.error(f"Error setting up scheduler: {e}")
            # Don't raise the exception, just log it so the bot can continue
    def run_scheduler(self):
        """Run the scheduler with better error handling"""
        while True:
            try:
                schedule.run_pending()
                time.sleep(1)
            except Exception as e:
                logger.error(f"Scheduler error: {e}")
                time.sleep(60)  # Wait a minute before retrying


    def send_4h_growth_chart(self):
        """Send 4-hour growth chart with historical trends from Google Sheets"""
        try:
            logger.info("📊 Generating 4-hour growth chart with historical trends...")
            
            # Get historical data from Google Sheets
            historical_data = self.get_historical_data_from_sheets()
            
            if not historical_data:
                logger.warning("No historical data available for growth chart")
                # Fallback to current data
                self.send_4h_growth_chart_fallback()
                return
            
            # Analyze growth trends from historical data
            growth_analysis = self.analyze_growth_trends(historical_data)
            
            if not growth_analysis:
                logger.warning("No growth analysis data available")
                self.send_4h_growth_chart_fallback()
                return
            
            # Create the chart message with historical trends
            chart_message = self.create_historical_growth_chart(growth_analysis)
            
            # Send to Telegram
            self.send_broadcast_message(chart_message)
            
            logger.info("✅ 4-hour growth chart with historical trends sent successfully")
            
        except Exception as e:
            logger.error(f"Error sending 4h growth chart with historical data: {e}")
            # Fallback to basic chart
            self.send_4h_growth_chart_fallback()

    def send_4h_growth_chart_fallback(self):
        """Fallback method using current data only"""
        try:
            price_data = self.get_consistent_price_data()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            if not analyzed_prices:
                return
            
            symbols_with_4h_growth = []
            for item in analyzed_prices:
                changes = item.get('changes', {})
                change_4h = changes.get('240m')
                price = item.get('price')
                
                if change_4h is not None and price is not None:
                    symbols_with_4h_growth.append({
                        'symbol': item['symbol'],
                        'price': price,
                        'change_4h': change_4h,
                        'changes': changes
                    })
            
            symbols_with_4h_growth.sort(key=lambda x: x['change_4h'], reverse=True)
            top_10_growth = symbols_with_4h_growth[:10]
            
            if top_10_growth:
                chart_message = self.create_growth_chart_message(top_10_growth)
                self.send_broadcast_message(chart_message)
                
        except Exception as e:
            logger.error(f"Error in fallback growth chart: {e}")


    def create_growth_chart_message(self, top_growth_symbols):
        """Create formatted growth chart message"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
            
            message = f"📈 <b>4-HOUR GROWTH LEADERS</b>\n\n"
            message += f"🕒 <i>As of {current_time}</i>\n\n"
            
            # Create ASCII-style bar chart
            max_change = max(abs(item['change_4h']) for item in top_growth_symbols)
            
            for i, item in enumerate(top_growth_symbols, 1):
                symbol = item['symbol']
                price = item['price']
                change_4h = item['change_4h']
                
                # Create progress bar
                bar_length = 20
                normalized_length = int((abs(change_4h) / max_change) * bar_length) if max_change > 0 else 0
                bar = "█" * normalized_length
                
                # Format based on change direction
                if change_4h > 0:
                    emoji = "🚀" if change_4h > 20 else "🟢" if change_4h > 10 else "📈"
                    bar_display = f"{emoji} {bar}"
                else:
                    emoji = "💥" if change_4h < -20 else "🔴" if change_4h < -10 else "📉"
                    bar_display = f"{bar} {emoji}"
                
                # Format price based on value
                if price >= 1000:
                    price_display = f"${price:,.2f}"
                elif price >= 1:
                    price_display = f"${price:.2f}"
                elif price >= 0.01:
                    price_display = f"${price:.4f}"
                else:
                    price_display = f"${price:.6f}"
                
                message += f"{i}. <b>{symbol}</b>\n"
                message += f"   {price_display} | {change_4h:+.1f}%\n"
                message += f"   {bar_display}\n\n"
            
            # Add summary statistics
            avg_growth = sum(item['change_4h'] for item in top_growth_symbols) / len(top_growth_symbols)
            best_growth = top_growth_symbols[0]['change_4h']
            worst_growth = top_growth_symbols[-1]['change_4h'] if len(top_growth_symbols) > 1 else best_growth
            
            message += f"📊 <b>Summary:</b>\n"
            message += f"• Best: <b>{best_growth:+.1f}%</b> ({top_growth_symbols[0]['symbol']})\n"
            message += f"• Average: <b>{avg_growth:+.1f}%</b>\n"
            message += f"• Count: {len(top_growth_symbols)} symbols\n\n"
            
            message += "⚡ <i>Next update in 4 hours</i>"
            
            return message
            
        except Exception as e:
            logger.error(f"Error creating growth chart message: {e}")
            return "📊 <b>4-Hour Growth Chart</b>\n\nError generating chart."

    def send_detailed_growth_report(self, update: Update, context: CallbackContext):
        """Manual command to get detailed growth report"""
        try:
            update.message.reply_html("📊 <b>Generating detailed growth report...</b>")
            
            # Get price data
            price_data = self.get_consistent_price_data()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            if not analyzed_prices:
                update.message.reply_html("❌ No price data available")
                return
            
            # Create reports for different timeframes
            timeframes = [
                ('5m', '5-minute'),
                ('60m', '1-hour'), 
                ('240m', '4-hour'),
                ('24h', '24-hour')
            ]
            
            for timeframe_key, timeframe_name in timeframes:
                # Filter and sort by timeframe growth
                symbols_with_growth = []
                for item in analyzed_prices:
                    changes = item.get('changes', {})
                    change = changes.get(timeframe_key)
                    price = item.get('price')
                    
                    if change is not None and price is not None:
                        symbols_with_growth.append({
                            'symbol': item['symbol'],
                            'price': price,
                            'change': change,
                            'changes': changes
                        })
                
                # Sort by growth and take top 5
                symbols_with_growth.sort(key=lambda x: x['change'], reverse=True)
                top_growth = symbols_with_growth[:5]
                
                if not top_growth:
                    continue
                
                message = f"🏆 <b>{timeframe_name.upper()} GROWTH LEADERS</b>\n\n"
                
                for i, item in enumerate(top_growth, 1):
                    symbol = item['symbol']
                    price = item['price']
                    change = item['change']
                    
                    # Format price
                    if price >= 1000:
                        price_display = f"${price:,.2f}"
                    elif price >= 1:
                        price_display = f"${price:.2f}"
                    elif price >= 0.01:
                        price_display = f"${price:.4f}"
                    else:
                        price_display = f"${price:.6f}"
                    
                    # Format change with emoji
                    if change > 20:
                        change_emoji = "🚀"
                    elif change > 10:
                        change_emoji = "🟢" 
                    elif change > 5:
                        change_emoji = "📈"
                    elif change < -20:
                        change_emoji = "💥"
                    elif change < -10:
                        change_emoji = "🔴"
                    elif change < -5:
                        change_emoji = "📉"
                    else:
                        change_emoji = "⚪"
                    
                    message += f"{i}. <b>{symbol}</b>\n"
                    message += f"   {price_display} | {change_emoji} {change:+.1f}%\n\n"
                
                # Send each timeframe report
                update.message.reply_html(message)
                time.sleep(1)  # Rate limiting
            
            update.message.reply_html("✅ <b>Growth reports completed!</b>")
            
        except Exception as e:
            update.message.reply_html(f"❌ Error generating growth report: {str(e)}")

    def send_quick_growth_chart(self, update: Update, context: CallbackContext):
        """Quick command to get current 4h growth chart"""
        try:
            update.message.reply_html("📊 <b>Generating quick growth chart...</b>")
            
            # Get price data
            price_data = self.get_consistent_price_data()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            if not analyzed_prices:
                update.message.reply_html("❌ No price data available")
                return
            
            # Filter and sort by 4h growth
            symbols_with_4h_growth = []
            for item in analyzed_prices:
                changes = item.get('changes', {})
                change_4h = changes.get('240m')
                price = item.get('price')
                
                if change_4h is not None and price is not None:
                    symbols_with_4h_growth.append({
                        'symbol': item['symbol'],
                        'price': price,
                        'change_4h': change_4h,
                        'changes': changes
                    })
            
            # Sort by 4h growth and take top 10
            symbols_with_4h_growth.sort(key=lambda x: x['change_4h'], reverse=True)
            top_10_growth = symbols_with_4h_growth[:10]
            
            if not top_10_growth:
                update.message.reply_html("❌ No 4h growth data available")
                return
            
            # Create and send chart
            chart_message = self.create_growth_chart_message(top_10_growth)
            update.message.reply_html(chart_message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error generating quick growth chart: {str(e)}")



    def create_historical_growth_chart(self, growth_analysis):
        """Create growth chart with historical trend analysis"""
        try:
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
            
            # Sort by 4h growth for the main chart
            sorted_by_4h = sorted(
                [item for item in growth_analysis.values() if item['change_4h'] is not None],
                key=lambda x: x['change_4h'],
                reverse=True
            )[:10]
            
            if not sorted_by_4h:
                return "📊 <b>4-Hour Growth Chart</b>\n\nNo growth data available."
            
            message = f"📈 <b>4-HOUR GROWTH LEADERS</b>\n\n"
            message += f"🕒 <i>Historical Trends Analysis - {current_time}</i>\n\n"
            
            # Create chart with trend indicators
            max_change = max(abs(item['change_4h']) for item in sorted_by_4h)
            
            for i, item in enumerate(sorted_by_4h, 1):
                symbol = item['symbol']
                price = item['current_price']
                change_4h = item['change_4h']
                momentum = item['momentum']
                consistency = item['trend_consistency']
                
                # Create enhanced progress bar with trend indicator
                bar_length = 15
                normalized_length = int((abs(change_4h) / max_change) * bar_length) if max_change > 0 else 0
                bar = "█" * normalized_length
                
                # Determine trend emoji based on multiple factors
                trend_emoji = self.get_trend_emoji(change_4h, momentum, consistency)
                
                # Momentum indicator
                if momentum > 5:
                    momentum_indicator = "⚡"  # Accelerating
                elif momentum > 2:
                    momentum_indicator = "📈"  # Growing
                elif momentum < -5:
                    momentum_indicator = "🔻"  # Decelerating fast
                elif momentum < -2:
                    momentum_indicator = "📉"  # Decelerating
                else:
                    momentum_indicator = "➡️"  # Stable
                
                # Consistency indicator
                if consistency >= 80:
                    consistency_indicator = "🟢"  # Very consistent
                elif consistency >= 60:
                    consistency_indicator = "🟡"  # Somewhat consistent
                else:
                    consistency_indicator = "🔴"  # Inconsistent
                
                # Format price
                price_display = self.format_price_for_display(price)
                
                message += f"{i}. <b>{symbol}</b>\n"
                message += f"   {price_display} | {trend_emoji} {change_4h:+.1f}%\n"
                message += f"   {bar} {momentum_indicator}{consistency_indicator}\n"
                
                # Add trend details for top 3
                if i <= 3:
                    message += f"   └─ 1h: {item['change_1h']:+.1f}% | 30m: {item['change_30m']:+.1f}% | 5m: {item['change_5m']:+.1f}%\n"
                
                message += "\n"
            
            # Add comprehensive summary
            message += self.create_growth_summary(sorted_by_4h)
            
            return message
            
        except Exception as e:
            logger.error(f"Error creating historical growth chart: {e}")
            return "📊 <b>Growth Chart</b>\n\nError generating chart."

    def get_trend_emoji(self, change_4h, momentum, consistency):
        """Get appropriate emoji based on trend analysis"""
        if change_4h > 20 and momentum > 5 and consistency >= 70:
            return "🚀"  # Strong, accelerating, consistent uptrend
        elif change_4h > 15 and momentum > 0:
            return "🟢"  # Strong uptrend with positive momentum
        elif change_4h > 10:
            return "📈"  # Good uptrend
        elif change_4h > 5:
            return "↗️"  # Moderate uptrend
        elif change_4h < -20 and momentum < -5:
            return "💥"  # Strong, accelerating downtrend
        elif change_4h < -15:
            return "🔴"  # Strong downtrend
        elif change_4h < -10:
            return "📉"  # Good downtrend
        elif change_4h < -5:
            return "↘️"  # Moderate downtrend
        else:
            return "➡️"  # Sideways


    def create_growth_summary(self, top_growth):
        """Create comprehensive growth summary"""
        try:
            if not top_growth:
                return ""
            
            avg_4h_growth = sum(item['change_4h'] for item in top_growth) / len(top_growth)
            best_growth = top_growth[0]['change_4h']
            best_symbol = top_growth[0]['symbol']
            
            # Count symbols by trend strength
            strong_uptrend = len([item for item in top_growth if item['change_4h'] > 15])
            uptrend = len([item for item in top_growth if 5 < item['change_4h'] <= 15])
            sideways = len([item for item in top_growth if -5 <= item['change_4h'] <= 5])
            downtrend = len([item for item in top_growth if item['change_4h'] < -5])
            
            # Average momentum
            avg_momentum = sum(item['momentum'] for item in top_growth) / len(top_growth)
            
            # Most consistent symbol
            most_consistent = max(top_growth, key=lambda x: x['trend_consistency'])
            
            summary = "📊 <b>Market Summary:</b>\n"
            summary += f"• Best Performer: <b>{best_growth:+.1f}%</b> ({best_symbol})\n"
            summary += f"• Average 4h Growth: <b>{avg_4h_growth:+.1f}%</b>\n"
            summary += f"• Market Momentum: <b>{avg_momentum:+.1f}</b>\n"
            summary += f"• Most Consistent: {most_consistent['symbol']} ({most_consistent['trend_consistency']:.0f}%)\n"
            summary += f"• Trend Distribution: 🚀{strong_uptrend} ↗️{uptrend} ➡️{sideways} ↘️{downtrend}\n"
            summary += f"• Total Analyzed: {len(top_growth)} symbols\n\n"
            summary += "⚡ <i>Based on historical data from Google Sheets</i>\n"
            summary += "🔄 <i>Next update in 4 hours</i>"
            
            return summary
            
        except Exception as e:
            logger.error(f"Error creating growth summary: {e}")
            return ""
    def analyze_consistency_leaders(self, historical_data):
        """Analyze symbols with most consistent trends"""
        growth_analysis = self.analyze_growth_trends(historical_data)
        if not growth_analysis:
            return None
        
        consistency_leaders = sorted(
            [item for item in growth_analysis.values() if item['trend_consistency'] is not None],
            key=lambda x: x['trend_consistency'],
            reverse=True
        )[:5]
        
        if not consistency_leaders:
            return None
        
        message = "🎯 <b>MOST CONSISTENT TRENDS</b>\n\n"
        message += "<i>Symbols with stable directional movement</i>\n\n"
        
        for i, item in enumerate(consistency_leaders, 1):
            # Determine consistency level
            if item['trend_consistency'] >= 80:
                consistency_level = "🟢 VERY CONSISTENT"
            elif item['trend_consistency'] >= 60:
                consistency_level = "🟡 CONSISTENT"
            else:
                consistency_level = "🔴 MIXED"
            
            message += f"{i}. <b>{item['symbol']}</b>\n"
            message += f"   {consistency_level} ({item['trend_consistency']:.0f}%)\n"
            message += f"   4h: {item['change_4h']:+.1f}% | Volatility: {item['volatility']:.1f}%\n"
            message += f"   Timeframes: {item['timeframes_available']}/5 available\n\n"
        
        return message

    def analyze_volatility_leaders(self, historical_data):
        """Analyze symbols with highest volatility"""
        growth_analysis = self.analyze_growth_trends(historical_data)
        if not growth_analysis:
            return None
        
        volatility_leaders = sorted(
            [item for item in growth_analysis.values() if item['volatility'] is not None],
            key=lambda x: x['volatility'],
            reverse=True
        )[:5]
        
        if not volatility_leaders:
            return None
        
        message = "🌊 <b>HIGH VOLATILITY SYMBOLS</b>\n\n"
        message += "<i>Symbols with largest price swings</i>\n\n"
        
        for i, item in enumerate(volatility_leaders, 1):
            # Determine volatility level
            if item['volatility'] > 30:
                volatility_level = "🌪️ EXTREME"
            elif item['volatility'] > 20:
                volatility_level = "⚡ HIGH"
            elif item['volatility'] > 10:
                volatility_level = "📊 MODERATE"
            else:
                volatility_level = "💤 LOW"
            
            message += f"{i}. <b>{item['symbol']}</b>\n"
            message += f"   {volatility_level} Volatility: {item['volatility']:.1f}%\n"
            message += f"   Range: {item['min_growth']:+.1f}% to {item['max_growth']:+.1f}%\n"
            message += f"   Current 4h: {item['change_4h']:+.1f}%\n\n"
        
        return message

    def analyze_reversal_candidates(self, historical_data):
        """Analyze symbols that might be reversing trends"""
        growth_analysis = self.analyze_growth_trends(historical_data)
        if not growth_analysis:
            return None
        
        reversal_candidates = []
        
        for symbol, item in growth_analysis.items():
            # Look for symbols with negative momentum but positive recent changes
            if (item['momentum'] is not None and item['change_5m'] is not None and 
                item['change_4h'] is not None):
                
                # Potential reversal: negative overall momentum but positive recent movement
                if (item['momentum'] < -2 and item['change_5m'] > 2 and 
                    item['change_4h'] < 0):
                    reversal_candidates.append(item)
                
                # Or positive momentum but negative recent movement (potential top)
                elif (item['momentum'] > 2 and item['change_5m'] < -2 and 
                    item['change_4h'] > 0):
                    reversal_candidates.append(item)
        
        # Sort by momentum strength (absolute value)
        reversal_candidates.sort(key=lambda x: abs(x['momentum']), reverse=True)
        top_reversals = reversal_candidates[:5]
        
        if not top_reversals:
            return None
        
        message = "🔄 <b>POTENTIAL REVERSAL CANDIDATES</b>\n\n"
        message += "<i>Symbols showing trend change signals</i>\n\n"
        
        for i, item in enumerate(top_reversals, 1):
            momentum = item['momentum']
            
            if momentum < 0 and item['change_5m'] > 0:
                signal = "🟢 POTENTIAL BOTTOM"
                explanation = "Negative momentum but recent buying"
            else:
                signal = "🔴 POTENTIAL TOP"
                explanation = "Positive momentum but recent selling"
            
            message += f"{i}. <b>{item['symbol']}</b>\n"
            message += f"   {signal}\n"
            message += f"   Momentum: {momentum:+.1f} | 5m: {item['change_5m']:+.1f}%\n"
            message += f"   4h: {item['change_4h']:+.1f}% | 1h: {item['change_1h']:+.1f}%\n"
            message += f"   {explanation}\n\n"
        
        return message

    def analyze_growth_trends(self, historical_data):
        """Analyze growth trends from historical Google Sheets data - ENHANCED"""
        try:
            growth_analysis = {}
            
            for symbol, data in historical_data.items():
                current_price = data.get('current_price')
                change_4h = data.get('change_4h')
                change_1h = data.get('change_1h')
                change_30m = data.get('change_30m')
                change_15m = data.get('change_15m')
                change_5m = data.get('change_5m')
                score = data.get('score', 0)
                
                if current_price is None or change_4h is None:
                    continue
                
                # Calculate trend consistency
                changes = [change_5m, change_15m, change_30m, change_1h, change_4h]
                valid_changes = [ch for ch in changes if ch is not None]
                
                if len(valid_changes) < 3:  # Need at least 3 timeframes for trend analysis
                    continue
                
                # Calculate trend metrics
                avg_growth = sum(valid_changes) / len(valid_changes)
                max_growth = max(valid_changes)
                min_growth = min(valid_changes)
                volatility = max_growth - min_growth
                
                # Trend direction analysis
                positive_changes = len([ch for ch in valid_changes if ch > 0])
                negative_changes = len([ch for ch in valid_changes if ch < 0])
                trend_consistency = positive_changes / len(valid_changes) * 100
                
                # Momentum analysis (recent vs older changes)
                recent_changes = [ch for ch in [change_5m, change_15m, change_30m] if ch is not None]
                older_changes = [ch for ch in [change_1h, change_4h] if ch is not None]
                
                recent_avg = sum(recent_changes) / len(recent_changes) if recent_changes else 0
                older_avg = sum(older_changes) / len(older_changes) if older_changes else 0
                momentum = recent_avg - older_avg  # Positive = accelerating growth
                
                growth_analysis[symbol] = {
                    'symbol': symbol,
                    'current_price': current_price,
                    'change_4h': change_4h,
                    'change_1h': change_1h,
                    'change_30m': change_30m,
                    'change_15m': change_15m,
                    'change_5m': change_5m,
                    'avg_growth': avg_growth,
                    'max_growth': max_growth,
                    'min_growth': min_growth,
                    'volatility': volatility,
                    'trend_consistency': trend_consistency,
                    'momentum': momentum,
                    'score': score,
                    'timeframes_available': len(valid_changes)
                }
            
            return growth_analysis
            
        except Exception as e:
            logger.error(f"Error analyzing growth trends: {e}")
            return {}

    def check_historical_data_availability(self):
        """Check if we have sufficient historical data for analysis"""
        try:
            historical_data = self.get_historical_data_from_sheets()
            if not historical_data:
                return False, "No historical data found"
            
            # Check if we have data for multiple timeframes
            symbols_with_multiple_timeframes = 0
            for symbol, data in historical_data.items():
                changes = [data.get('change_5m'), data.get('change_15m'), data.get('change_30m'), 
                        data.get('change_1h'), data.get('change_4h')]
                valid_changes = len([ch for ch in changes if ch is not None])
                if valid_changes >= 3:
                    symbols_with_multiple_timeframes += 1
            
            if symbols_with_multiple_timeframes < 5:
                return False, f"Only {symbols_with_multiple_timeframes} symbols have sufficient data"
            
            return True, f"Good data coverage: {symbols_with_multiple_timeframes} symbols"
            
        except Exception as e:
            return False, f"Error checking data: {str(e)}"

    def data_status_command(self, update: Update, context: CallbackContext):
        """Check historical data status"""
        try:
            has_data, message = self.check_historical_data_availability()
            
            status_message = "📊 <b>Historical Data Status</b>\n\n"
            
            if has_data:
                status_message += "✅ <b>Status:</b> Data available for analysis\n"
            else:
                status_message += "❌ <b>Status:</b> Insufficient data\n"
            
            status_message += f"📝 <b>Details:</b> {message}\n\n"
            status_message += "💡 <i>The bot needs at least 5 symbols with 3+ timeframes of data for proper trend analysis.</i>"
            
            update.message.reply_html(status_message)
            
        except Exception as e:
            update.message.reply_html(f"❌ Error checking data status: {str(e)}")


    def send_trend_analysis_command(self, update: Update, context: CallbackContext):
        """Manual command for detailed trend analysis - FIXED"""
        try:
            update.message.reply_html("📈 <b>Generating detailed trend analysis...</b>")
            
            # Get historical data
            historical_data = self.get_historical_data_from_sheets()
            
            if not historical_data:
                update.message.reply_html("❌ No historical data available in Google Sheets")
                return
            
            # Analyze different aspects
            analyses = [
                ("Momentum Leaders", self.analyze_momentum_leaders(historical_data)),
                ("Consistency Leaders", self.analyze_consistency_leaders(historical_data)),
                ("Volatility Analysis", self.analyze_volatility_leaders(historical_data)),
                ("Reversal Candidates", self.analyze_reversal_candidates(historical_data))
            ]
            
            analyses_sent = 0
            for analysis_name, analysis_result in analyses:
                if analysis_result:
                    update.message.reply_html(analysis_result)
                    analyses_sent += 1
                    time.sleep(1)  # Rate limiting
            
            if analyses_sent == 0:
                update.message.reply_html("❌ No trend analysis data available. The bot may need more historical data to generate trends.")
            else:
                update.message.reply_html(f"✅ <b>Trend analysis completed! ({analyses_sent}/4 reports generated)</b>")
            
        except Exception as e:
            update.message.reply_html(f"❌ Error in trend analysis: {str(e)}")




    def analyze_momentum_leaders(self, historical_data):
        """Analyze symbols with strongest momentum"""
        growth_analysis = self.analyze_growth_trends(historical_data)
        if not growth_analysis:
            return None
        
        momentum_leaders = sorted(
            [item for item in growth_analysis.values() if item['momentum'] is not None],
            key=lambda x: x['momentum'],
            reverse=True
        )[:5]
        
        if not momentum_leaders:
            return None
        
        message = "⚡ <b>MOMENTUM LEADERS</b>\n\n"
        message += "<i>Symbols with accelerating growth</i>\n\n"
        
        for i, item in enumerate(momentum_leaders, 1):
            message += f"{i}. <b>{item['symbol']}</b>\n"
            message += f"   Momentum: {item['momentum']:+.1f}\n"
            message += f"   4h: {item['change_4h']:+.1f}% | 1h: {item['change_1h']:+.1f}%\n"
            message += f"   5m: {item['change_5m']:+.1f}% | Consistency: {item['trend_consistency']:.0f}%\n\n"
        
        return message




    def run_price_monitoring(self):
        """Run price monitoring and alert on significant movements - FIXED"""
        try:
            logger.info("💰 Running price monitoring...")
            
            price_data = self.get_consistent_price_data()
            analyzed_prices = self.analyze_price_movements(price_data)
            
            # Check for significant movers
            significant_movers = []
            for item in analyzed_prices[:20]:  # Check top 20
                changes = item.get('changes', {})
                change_5m = changes.get('5m', 0)
                change_1h = changes.get('60m', 0)
                
                # Alert criteria - lowered thresholds for testing
                if abs(change_5m) > 2 or abs(change_1h) > 5:  # 2% in 5min or 5% in 1h
                    significant_movers.append(item)
            
            # Send alerts for significant movers
            if significant_movers:
                self.send_price_alert(significant_movers)
                logger.info(f"🚨 Sent alerts for {len(significant_movers)} significant movers")
            else:
                logger.info("💰 No significant price movements detected")
                
        except Exception as e:
            logger.error(f"Price monitoring error: {e}")

    def send_price_alert(self, significant_movers):
        """Send alert for significant price movements - FIXED"""
        try:
            message = "🚨 <b>SIGNIFICANT PRICE MOVEMENTS!</b>\n\n"
            
            for item in significant_movers[:5]:  # Max 5 alerts
                changes = item.get('changes', {})
                message += f"📈 <b>{item['symbol']}</b>\n"
                message += f"   Price: {self.format_price_for_display(item.get('price'))}\n"
                
                change_5m = changes.get('5m', 0)
                change_1h = changes.get('60m', 0)
                
                if abs(change_5m) > 2:
                    message += f"   🚀 5m: {self.format_change_for_telegram(change_5m)}\n"
                if abs(change_1h) > 5:
                    message += f"   📊 1h: {self.format_change_for_telegram(change_1h)}\n"
                
                message += "\n"
            
            self.send_broadcast_message(message)
            logger.info("✅ Price alert sent to Telegram")
            
        except Exception as e:
            logger.error(f"Error sending price alert: {e}")

    # ==================== DATA MANAGEMENT ====================

    def init_data_file(self):
        """Initialize data in memory"""
        self.data = self.get_default_data()

    def load_data(self):
        """Load data from memory"""
        return self.data
    
    def save_data(self, data):
        """Save data to memory"""
        self.data = data
        logger.info("Data saved to memory")

    def get_default_data(self):
        """Return default data structure"""
        return {
            "unique_futures": [],
            "last_check": None,
            "statistics": {
                "checks_performed": 0,
                "unique_found_total": 0,
                "start_time": datetime.now().isoformat()
            },
            "exchange_stats": {},
            "google_sheet_url": None,
            "price_alerts_sent": {}
        }

    def send_broadcast_message(self, message):
        """Send message to configured chat"""
        try:
            if self.chat_id:
                self.bot.send_message(
                    chat_id=self.chat_id,
                    text=message,
                    parse_mode='HTML'
                )
        except Exception as e:
            logger.error(f"Broadcast error: {e}")

    def run(self):
        """Start the bot"""
        try:
            # Load initial data
            data = self.load_data()
            self.last_unique_futures = set(data.get('unique_futures', []))
            
            # Setup Google Sheets historical storage
            if self.gs_client and self.spreadsheet:
                self.setup_google_sheets_historical_storage()
                logger.info("✅ Historical data storage initialized")
            
            # Setup scheduler
            self.setup_scheduler()
            
            # Start scheduler in background
            scheduler_thread = threading.Thread(target=self.run_scheduler, daemon=True)
            scheduler_thread.start()
            
            # Start the bot
            self.updater.start_polling()
            
            logger.info("Bot started successfully with historical data tracking")
            
            # Send startup message
            startup_msg = (
                "🤖 <b>MEXC Futures Tracker Started</b>\n\n"
                "✅ Monitoring 8 exchanges\n"
                f"⏰ Unique check: {self.update_interval} minutes\n"
                f"💰 Price check: {self.price_check_interval} minutes\n"
                f"📊 Historical data: ENABLED\n"
                "🎯 Unique futures detection\n"
                "🚀 Price movement alerts\n"
                "💬 Use /help for commands"
            )
            
            self.send_broadcast_message(startup_msg)
            
            logger.info("Bot is now running with historical data tracking...")
            
            # Keep running
            self.updater.idle()
                
        except Exception as e:
            logger.error(f"Bot run error: {e}")
            raise



def main():
    tracker = MEXCTracker()
    tracker.run()

if __name__ == "__main__":
    print("Starting MEXC Futures Tracker...")
    main()